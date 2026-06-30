from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.service import Service as EdgeService
from fcm_intake.config import EDGE_DRIVER_PATH as CONFIG_EDGE_DRIVER_PATH
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
import time

# Set up the WebDriver (make sure you have the correct driver installed, e.g., ChromeDriver)
# driver = webdriver.Chrome()

edgeOptions = Options()
edgeOptions.add_argument('--ignore-certificate-errors')
edgeOptions.add_argument('--allow-insecure-localhost') 
edgeOptions.add_argument('--disable-popup-blocking')

service = EdgeService(executable_path=CONFIG_EDGE_DRIVER_PATH)
driver = webdriver.Edge(service=service, options=edgeOptions)   

#Employer
# EMPLOYER_NAME = "Global Building Services Inc"
# EMPLOYER_NAME = "Ruan Transportation Management"
EMPLOYER_CODE = ""#"EMG291"
# EMPLOYER_ADDRESS1 = "427 W. Colorado St. Ste 106"
EMPLOYER_ADDRESS1 = ""
EMPLOYER_CITY = ""
EMPLOYER_STATE = ""
EMPLOYER_ZIP = ""
CASE_NUM = "ZS01KY"

# CUSTOMER_NAME = "First Regional Library"
# CUSTOMER_CODE = "C02283"
# CUSTOMER_ADDRESS1 = "427 W. Colorado St. Ste 106"
# CUSTOMER_CITY = "Glendale"
# CUSTOMER_STATE = "California"
# CUSTOMER_ZIP = "91204"

CUSTOMER_NAME = ""
CUSTOMER_CODE = ""
CUSTOMER_ADDRESS1 = ""
CUSTOMER_CITY = ""
CUSTOMER_STATE = ""
CUSTOMER_ZIP = ""





def employerSearch(CustomerName,ClaimID,ClaimantName):
    driver.get("https://uat-gnx-cem-portal-eus-as.azurewebsites.net/employerSearch")
    
    if elementExist(By.XPATH, "/html/body/div[3]/div/div/div/div/div/div[2]/ul/li/button"):
        loginbtn = driver.find_element(By.XPATH,"/html/body/div[3]/div/div/div/div/div/div[2]/ul/li/button")
        loginbtn.click()
    time.sleep(5)

    eCode = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[2]/div[2]/input")
    eCode.send_keys(EMPLOYER_CODE)

    eName = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[1]/div[2]/input")
    eName.send_keys(CustomerName)

    btnSearch = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[10]/div[2]/button[1]")
    btnSearch.click()

    time.sleep(5)

    
    #Add if no result in employer
    if not elementExist(By.XPATH, "/html/body/app/div[2]/div[2]/table"):
        addEmployerBtn = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[10]/div[2]/button[2]")
        addEmployerBtn.click()
        time.sleep(5)
        driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[1]/div[2]/input").send_keys(CustomerName)
        customerName = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[2]/div[2]/div[1]/div/input")
        customerName.send_keys("Liberty Mutual Commercial Market (C02283)")
        time.sleep(5)
        customerName.send_keys(Keys.RETURN)
        time.sleep(2)
        # no ADdress
        if not EMPLOYER_ADDRESS1 or EMPLOYER_ADDRESS1 == "":
            
            address1 = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[5]/div[2]/input")
            address1.send_keys("Unknown")

            city = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[6]/div[2]/input")
            city.send_keys("Unknown")

            dropdownState = Select(driver.find_element(By.ID, "stateId"))
            dropdownState.select_by_visible_text("Unknown")
        
            zip = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[8]/div[2]/input")
            zip.send_keys("00000")

            dropdowCountry = Select(driver.find_element(By.ID, "countryId"))
            dropdowCountry.select_by_visible_text("USA")

            time.sleep(5)
        else:
            address1 = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[5]/div[2]/input")
            address1.send_keys(EMPLOYER_ADDRESS1)

            city = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[6]/div[2]/input")
            city.send_keys(EMPLOYER_CITY)

            dropdownState = Select(driver.find_element(By.ID, "stateId"))
            dropdownState.select_by_visible_text(EMPLOYER_STATE)
        
            zip = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[8]/div[2]/input")
            zip.send_keys(EMPLOYER_ZIP)

            dropdowCountry = Select(driver.find_element(By.ID, "countryId"))
            dropdowCountry.select_by_visible_text("USA")

            time.sleep(5)
        
        notify("Validation","Kindly validate if information is correct. Manually click Submit in CEM before clicking Ok.")
        add_claim_record(ClaimantName,ClaimID)
    else:
        tableResult = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/table")
        tbRows = tableResult.find_elements(By.XPATH, ".//tbody/tr")

        for row in tbRows:
            tdResult = row.find_elements(By.TAG_NAME, "td")
            tdEmpName = tdResult[0].text.strip()
            if CustomerName in tdEmpName:
                tdResult[0].click() 
                time.sleep(5)

                tableLoc = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[8]/div/div[2]/div[2]/div[1]/div/table")
                trsLoc = tableLoc.find_elements(By.TAG_NAME, ".//tbody/tr")
                found = False
                for locRow in trsLoc:
                    locAddress = locRow.find_element(By.XPATH, ".td[1]").text.strip()

                    if EMPLOYER_ADDRESS1.lower() in locAddress.lower():
                        found = True
                        break
                
                if not found:
                    #add new location
                    driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[9]/div[2]/button").click()
                    time.sleep(2)
                    
                    driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[2]/div[2]/input").send_keys(EMPLOYER_ADDRESS1)
                    driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[3]/div[2]/input").send_keys(EMPLOYER_CITY)
                    dropdownState = Select(driver.find_element(By.ID, "stateId"))
                    dropdownState.select_by_visible_text(EMPLOYER_STATE)
                
                    zip = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[5]/div[2]/input")
                    zip.send_keys(EMPLOYER_ZIP)

                    dropdowCountry = Select(driver.find_element(By.ID, "countryId"))
                    dropdowCountry.select_by_visible_text("USA")

                    case = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[7]/div[2]/input")
                    case.send_keys(EMPLOYER_ZIP)
                    

def custSearch():
    driver.get("https://uat-gnx-cem-portal-eus-as.azurewebsites.net/customerSearch")
    
    if elementExist(By.XPATH, "/html/body/div[3]/div/div/div/div/div/div[2]/ul/li/button"):
        loginbtn = driver.find_element(By.XPATH,"/html/body/div[3]/div/div/div/div/div/div[2]/ul/li/button")
        loginbtn.click()
    time.sleep(5)
    if not CUSTOMER_CODE:       
        driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[1]/div[2]/input").send_keys(CUSTOMER_NAME)
    else:
        driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[2]/div[2]/input").send_keys(CUSTOMER_CODE)

    btnSearch = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[10]/div[2]/button[1]")
    btnSearch.click()
    
    time.sleep(2)
    
    #Add if no result in Customer
    if not elementExist(By.XPATH, "/html/body/app/div[2]/div[2]/table"):
        addCustBtn = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[10]/div[2]/button[2]")
        addCustBtn.click()
        time.sleep(5)
        driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[2]/div[2]/input").send_keys(CUSTOMER_NAME)
        time.sleep(2)
        if not EMPLOYER_ADDRESS1 or EMPLOYER_ADDRESS1 == "":
            
            address1 = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[4]/div[2]/input")
            address1.send_keys(CUSTOMER_ADDRESS1)

            city = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[5]/div[2]/input")
            city.send_keys(CUSTOMER_CITY)

            dropdownState = Select(driver.find_element(By.ID, "stateId"))
            dropdownState.select_by_visible_text(CUSTOMER_STATE)
        
            zip = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[7]/div[2]/input")
            zip.send_keys(CUSTOMER_ZIP)

            dropdowCountry = Select(driver.find_element(By.ID, "countryId"))
            dropdowCountry.select_by_visible_text("USA")

            caseNum = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[10]/div[2]/input")
            caseNum.send_keys(CASE_NUM)

            time.sleep(5)
    else:
        tableResult = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/table")
        tbRows = tableResult.find_elements(By.XPATH, ".//tbody/tr")

        for row in tbRows:
            tdResult = row.find_elements(By.TAG_NAME, "td")
            tdEmpName = tdResult[0].text.strip()
            if CUSTOMER_NAME.lower() in tdEmpName.lower():
                tdResult[0].click() 
                time.sleep(5)

                if CUSTOMER_STATE:
                    driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[7]/div/div[1]/div/table/thead/tr[2]/td[3]/div/div/span/input").send_keys(CUSTOMER_STATE)

                tableLoc = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[8]/div/div[2]/div[2]/div[1]/div/table")
                trsLoc = tableLoc.find_elements(By.TAG_NAME, ".//tbody/tr")
                found = False
                for locRow in trsLoc:
                    locAddress = locRow.find_element(By.XPATH, ".td[1]").text.strip()

                    if CUSTOMER_ADDRESS1.lower() in locAddress.lower():
                        found = True
                        break
                
                if not found:
                    #add new location
                    driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[8]/div[2]/button").click()
                    time.sleep(2)
                    
                    driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[2]/div[2]/input").send_keys(CUSTOMER_ADDRESS1)
                    driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[3]/div[2]/input").send_keys(CUSTOMER_CITY)
                    dropdownState = Select(driver.find_element(By.ID, "stateId"))
                    dropdownState.select_by_visible_text(CUSTOMER_STATE)
                
                    zip = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[5]/div[2]/input")
                    zip.send_keys(CUSTOMER_ZIP)

                    dropdowCountry = Select(driver.find_element(By.ID, "countryId"))
                    dropdowCountry.select_by_visible_text("USA")

                    case = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[8]/div[2]/input")
                    case.send_keys(CASE_NUM)
                    
                btnSubmit = driver.find_element(By.XPATH, "/html/body/app/div[2]/div[2]/form/div/div/div[12]/div[2]/button[1]")
                # btnSubmit.click()

def notify(title: str, text: str):
    """Show a messagebox if Tk is available; otherwise print."""
    try:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost',True)
        messagebox.showinfo(title, text,parent=root)
        root.destroy()
    except Exception:
        print(f"[{title}] {text}")       


def elementExist(by, value, timeout=10):
    try:
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))
        return True
    except:
        return False
    
def main(CustomerName:str,ClaimID,ClaimantName):
    try:
        employerSearch(CustomerName.title(),ClaimID,ClaimantName)
    except Exception as ex:
        print("error occurred: ", ex)
    finally:
        time.sleep(10)
        driver.quit()

import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
 
 
EXCEL_FILE_NAME = "FCM CEM Checher.xlsx"
HEADER_ROW = ["Claimant", "Claim ID", "Status"]
 
 
def get_script_dir():
    """
    Returns the directory where the script or exe is located.
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))
 
 
def get_excel_path():
    """
    Full path of the workbook in the same folder as the script.
    """
    return os.path.join(get_script_dir(), EXCEL_FILE_NAME)
 
 
def get_today_sheet_name():
    """
    Returns today's date in mmddyyyy format.
    Example: 03092026
    """
    return datetime.now().strftime("%m%d%Y")
 
 
def create_workbook_if_missing(excel_path):
    """
    Create the workbook if it does not exist.
    """
    if not os.path.exists(excel_path):
        wb = Workbook()
 
        # Default sheet from openpyxl
        ws = wb.active
        ws.title = get_today_sheet_name()
 
        # Add headers to first row
        for col_idx, header in enumerate(HEADER_ROW, start=1):
            ws.cell(row=1, column=col_idx, value=header)
 
        wb.save(excel_path)
        wb.close()
 
 
def get_or_create_today_worksheet(wb):
    """
    Returns today's worksheet. If missing, create it and add headers.
    """
    sheet_name = get_today_sheet_name()
 
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
 
    ensure_headers(ws)
    return ws
 
 
def ensure_headers(ws):
    """
    Make sure row 1 contains:
    A1 = Claimant
    B1 = Claim ID
    C1 = Status
    """
    for col_idx, header in enumerate(HEADER_ROW, start=1):
        current_value = ws.cell(row=1, column=col_idx).value
        if current_value != header:
            ws.cell(row=1, column=col_idx, value=header)
 
 
def get_next_empty_row(ws):
    """
    Finds the next empty row after the header.
    """
    row = ws.max_row
 
    # If only header exists, write on row 2
    if row < 2:
        return 2
 
    # Move downward until a truly empty A/B/C row is found
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if (a in [None, ""]) and (b in [None, ""]) and (c in [None, ""]):
            return r
 
    return ws.max_row + 1
 
 
def add_claim_record(ClaimantName, ClaimID, Status=""):
    """
    Creates/opens the workbook in the same folder as the script,
    checks today's worksheet in mmddyyyy format,
    creates it if missing,
    ensures headers exist,
    appends ClaimantName and ClaimID,
    saves and closes the workbook.
    """
    excel_path = get_excel_path()
 
    create_workbook_if_missing(excel_path)
 
    wb = load_workbook(excel_path)
    try:
        ws = get_or_create_today_worksheet(wb)
        next_row = get_next_empty_row(ws)
 
        ws.cell(row=next_row, column=1, value=ClaimantName)
        ws.cell(row=next_row, column=2, value=ClaimID)
        ws.cell(row=next_row, column=3, value=Status)
 
        wb.save(excel_path)
    finally:
        wb.close()
 
 
    




if __name__ == "__main__":
    main("UNIVERSAL STUDIOS HOLLYWOOD")

