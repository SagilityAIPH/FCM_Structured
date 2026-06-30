import pyodbc
from fcm_intake.config import CMS_CASESEARCH_URL as CONFIG_CMS_CASESEARCH_URL, CMS_LOGIN_URL as CONFIG_CMS_LOGIN_URL, EDGE_DRIVER_PATH as CONFIG_EDGE_DRIVER_PATH, EDGE_PATH as CONFIG_EDGE_PATH, IE_DRIVER_PATH as CONFIG_IE_DRIVER_PATH
import openpyxl
from openpyxl import load_workbook
import difflib
 
from selenium import webdriver
from selenium.webdriver.ie.service import Service as IeService
from selenium.webdriver.ie.options import Options as IeOptions
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import WebDriverException,NoSuchWindowException,StaleElementReferenceException
 
from pywinauto import Desktop, uia_defines
from pywinauto.uia_element_info import UIAElementInfo
from pywinauto.controls.uiawrapper import UIAWrapper
from pywinauto.application import Application

import time
import re
import os
import subprocess
from dataclasses import dataclass
import sys
from datetime import datetime

# import ProviderResultChecker as PRC
from legacy import legacy_providerresultchecker as PRC
 
# ===================== CONFIG =====================
 
CMS_LOGIN_URL = CONFIG_CMS_LOGIN_URL
CMS_CASESEARCH_URL = CONFIG_CMS_CASESEARCH_URL
 
EDGE_PATH = CONFIG_EDGE_PATH
IE_DRIVER_PATH = CONFIG_IE_DRIVER_PATH
 
LIBERTY_OFFICE_MATRIX = r"Liberty Mutual Claims Office Matrix_08-29-19.xlsx"
 
# When called by another bot, normally set INTERACTIVE = False
INTERACTIVE = False
 
# ===================== TK ROOT =====================
 
# root = tk.Tk()
# root.withdraw()
 
driver = None
 
# ===================== DATACLASS FOR SP RESULT =====================
 
@dataclass
class CaseData:
    claim_number: str
    xlReferralSource: str
    xlCustomerName: str
    dxCode: str
    bodyParts : str
    officeNumber: str
    xlProvider: str
    provType: str
    xlCustomerEmail: str
    xlInsuredAddr1: str
    xlAttorney: str
    xlAttorneyAdd1: str
    xlAttorneyAdd2: str
    xlAttorneyCity: str
    xlAttorneyState: str
    xlAttorneyZip: str
    xlAttorneyPhone: str
    apptDate: str
    refType: str
    xlProviderPhone:str
    Special_Instructions:str
 
# ===================== BASIC HELPERS =====================
 
def notify(title: str, msg: str):
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost',True)
    if INTERACTIVE:
        messagebox.showinfo(title, msg)
        root.destroy()
    else:
        print(f"[{title}] {msg}")
 
 
def ensure_readable_path(path: str, description: str = ""):
    if not os.path.exists(path):
        raise FileNotFoundError(f"{description or 'Path'} not found: {path}")
    if not os.access(path, os.R_OK):
        raise PermissionError(f"Access denied (no read permission) for {description or path}")
 
 
def ensure_executable_path(path: str, description: str = ""):
    ensure_readable_path(path, description)
    if not os.access(path, os.X_OK):
        print(f"⚠ Warning: {description or path} might not be executable. On Windows this is usually fine.")
 
 
def formatStr(s: str) -> str:
    return s.replace('\n', ' ').replace('\r', '').strip().lower() if s else ""
 
 
# ===================== DB: FETCH CASE DATA FROM SP =====================
 
def fetch_case_data_from_db(
    connection_string: str,
    claimant_name: str,
    claim_number: str,
    market: str = "Liberty Mutual Commercial Market"
) -> CaseData:
    connection = None
    cursor = None
    try:
        connection = pyodbc.connect(connection_string)
        cursor = connection.cursor()
 
        print(f"Executing RRS_Sagility_Details_Lookup for {claimant_name} / {claim_number}")
        cursor.execute("{CALL RRS_Sagility_Details_Lookup(?,?,?)}",
                       (claimant_name, claim_number, market))
        row = cursor.fetchone()
        if not row:
            raise ValueError(f"No records found for claimant={claimant_name}, claim={claim_number}")
 
        col_index = {desc[0]: i for i, desc in enumerate(cursor.description)}
 
        def get(col_name: str):
            idx = col_index.get(col_name)
            if idx is None:
                return None
            val = row[idx]
            if isinstance(val, str):
                return val.strip()
            return val
 
        def first_nonempty(*cols: str):
            for c in cols:
                val = get(c)
                if val not in (None, "", "None"):
                    return val
            return None
 
        # ---------------- core mapping ----------------
        claim_num = get("Claim_Number") or claim_number

        #Appt Date
        apptDate = get("ApptDate")

         #Special_Instructions
        Special_Instructions = get("Special_Instructions")

        #Ref Type
        refType = get("RefType")
 
        # Referral source -> "Que: <Que>"
        # que_val = get("Que")
        # if que_val:
        #     xlReferralSource = f"{que_val}"
        # else:
        #     xlReferralSource = ""


        que_val = get("Que")
        if que_val:
            xlReferralSource = f"{que_val}"
        else:
            xlReferralSource = ""

        # Customer / employer name
        xlCustomerName = get("Employer") or ""
 
        # Diagnosis code – Book1 says N/A; leave blank or map to a diagnosis column if you have one migz
        dxCode = get("DxCode") #migz

        bodyParts = get("Unity_Body_Part")
 
        # Office number = MID(ClaimNum,3,3) in Book1.xlsx
        if claim_num and len(claim_num) >= 5:
            officeNumber = claim_num[2:5]
        else:
            officeNumber = ""
 
        # Provider
        # xlProvider = first_nonempty("Unity_Provider_Facility", "Unity_Provider_Name") or ""
        xlProvider = get("Unity_Provider_Name") or ""
        # Provider type – adjust this to real column name if you have one
        FacProv=""
        FacProv = get("Unity_Provider_Facility") or ""
        if FacProv:
            provType = "facility"
        else:
            provType = ""
  
        #Provider Phone
        xlProviderPhone = get("Unity_Provider_Phone") or ""

        # Customer email & insured address1 (if you mapped these in Book1)
        xlCustomerEmail = get("Customer_Email") or ""
        xlInsuredAddr1 = get("Customer_Address1") or ""
 
        # Attorney
        atty_first = first_nonempty(
            "Unity_Claimant_Attorney_New_First_Name",
            "Unity_Defense_Attorney_New_First_Name",
        )
        atty_last = first_nonempty(
            "Unity_Claimant_Attorney_New_Last_Name",
            "Unity_Defense_Attorney_New_Last_Name",
        )
        atty_firm = first_nonempty(
            "Unity_Claimant_Attorney_New_Firm_Name",
            "Unity_Defense_Attorney_New_Firm_Name",
        )
 
        if atty_first or atty_last:
            xlAttorney = " ".join(filter(None, [atty_first, atty_last]))
        else:
            xlAttorney = atty_firm or ""
 
        xlAttorneyAdd1 = first_nonempty(
            "Unity_Claimant_Attorney_New_Address1",
            "Unity_Defense_Attorney_New_Address1",
        ) or ""
 
        xlAttorneyAdd2 = first_nonempty(
            "Unity_Claimant_Attorney_New_Address2",
            "Unity_Defense_Attorney_New_Address2",
        ) or ""
 
        xlAttorneyCity = first_nonempty(
            "Unity_Claimant_Attorney_New_City",
            "Unity_Defense_Attorney_New_City",
        ) or ""
 
        xlAttorneyState = first_nonempty(
            "Unity_Claimant_Attorney_New_State",
            "Unity_Defense_Attorney_New_State",
        ) or ""
 
        xlAttorneyZip = first_nonempty(
            "Unity_Claimant_Attorney_New_Zip",
            "Unity_Defense_Attorney_New_Zip",
        ) or ""
 
        xlAttorneyPhone = first_nonempty(
            "Unity_Claimant_Attorney_New_Phone",
            "Unity_Defense_Attorney_New_Phone",
        ) or ""
 
        return CaseData(
            claim_number=claim_num,
            xlReferralSource=xlReferralSource,
            xlCustomerName=xlCustomerName,
            dxCode=dxCode,
            bodyParts=bodyParts,
            officeNumber=officeNumber,
            xlProvider=xlProvider,
            provType=provType,
            xlCustomerEmail=xlCustomerEmail,
            xlInsuredAddr1=xlInsuredAddr1,
            xlAttorney=xlAttorney,
            xlAttorneyAdd1=xlAttorneyAdd1,
            xlAttorneyAdd2=xlAttorneyAdd2,
            xlAttorneyCity=xlAttorneyCity,
            xlAttorneyState=xlAttorneyState,
            xlAttorneyZip=xlAttorneyZip,
            xlAttorneyPhone=xlAttorneyPhone,
            apptDate=apptDate,
            refType=refType,
            xlProviderPhone=xlProviderPhone,
            Special_Instructions = Special_Instructions,
        )
 
    finally:
        if cursor is not None:
            cursor.close()
        if connection is not None:
            connection.close()
 
 
# ===================== SELENIUM SETUP =====================
 
def create_ie_driver() -> webdriver.Ie:
    ensure_executable_path(IE_DRIVER_PATH, "IE Driver")
    ensure_readable_path(EDGE_PATH, "Edge executable")
 
    options = IeOptions()
    options.add_additional_option("ie.edgechromium", True)
    options.add_additional_option("ie.edgepath", EDGE_PATH)
    options.add_additional_option("ignoreProtectedModeSettings", True)
    options.add_additional_option("requireWindowFocus", True)
    options.add_additional_option("nativeEvents", False)
    options.ensure_clean_session = True
 
    service = IeService(executable_path=IE_DRIVER_PATH)
 
    try:
        drv = webdriver.Ie(service=service, options=options)
        return drv
    except WebDriverException as e:
        raise RuntimeError(f"Failed to start IE driver (check permissions / IE settings): {e}") from e
 
 
# ===================== SELENIUM HELPERS =====================
 
def legacy_safe_type(by, sel, text, timeout=20, click_first=True):
    global driver
    wait = WebDriverWait(driver, timeout)
    el = wait.until(EC.presence_of_element_located((by, sel)))
 
    driver.execute_script("arguments[0].scrollIntoView(true);", el)
    if click_first:
        try:
            el.click()
        except WebDriverException:
            pass
 
    try:
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.DELETE)
        el.send_keys(text)
        if el.get_attribute("value") == text:
            return True
    except WebDriverException:
        pass
 
    try:
        ok = driver.execute_script(r"""
            var el = arguments[0], val = arguments[1];
            try { el.focus && el.focus(); } catch(e){}
            try { el.value = ''; } catch(e){}
            try { el.value = val; } catch(e){}
 
            if (document.createEventObject && el.fireEvent) {
                try {
                    var ev = document.createEventObject();
                    ev.propertyName = 'value';
                    el.fireEvent('onpropertychange', ev);
                } catch(e){}
                try { el.fireEvent('oninput', document.createEventObject()); } catch(e){}
                try { el.fireEvent('onchange', document.createEventObject()); } catch(e){}
                try { el.fireEvent('onkeyup', document.createEventObject()); } catch(e){}
                try { el.fireEvent('onblur', document.createEventObject()); } catch(e){}
                return el.value === val;
            } else if (document.createEvent) {
                var types = ['input','change','keyup','keydown','blur'];
                for (var i=0;i<types.length;i++){
                    try {
                        var evt = document.createEvent('HTMLEvents');
                        evt.initEvent(types[i], true, false);
                        el.dispatchEvent(evt);
                    } catch(e){}
                }
                return el.value === val;
            } else {
                return el.value === val;
            }
        """, el, text)
        if ok:
            return True
    except WebDriverException:
        pass
 
    return False
 
 
def legacy_clear(by, sel, timeout=20):
    global driver
    el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, sel)))
    try:
        el.clear()
        if not el.get_attribute("value"):
            return True
    except Exception:
        pass
 
    try:
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.DELETE)
        if not el.get_attribute("value"):
            return True
    except Exception:
        pass
 
    driver.execute_script("""
        var el = arguments[0];
        try { el.focus(); } catch(e){}
        try { el.value=''; } catch(e){}
        if (document.createEventObject && el.fireEvent) {
            var ev = document.createEventObject();
            el.fireEvent('onchange', ev);
        } else if (document.createEvent) {
            var evt = document.createEvent('HTMLEvents');
            evt.initEvent('change', true, false);
            el.dispatchEvent(evt);
        }
    """, el)
    return el.get_attribute("value") == ""
 
 
def elementExist(by, value, timeout=10):
    global driver
    try:
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))
        return True
    except Exception:
        return False
 
 
def element_check(el):
    global driver
    return driver.execute_script(r"""
        (function(el){
            el.checked = true;
            if (document.createEventObject && el.fireEvent) {
                el.fireEvent('onchange', document.createEventObject());
            } else if (document.createEvent) {
                var e = document.createEvent('Event');
                e.initEvent('change', true, false);
                el.dispatchEvent(e);
            }
        })(arguments[0]);
    """, el)
 
 
def element_click(el):
    global driver
    return driver.execute_script(r"""
        (function(el){
            function idToName(id){ return id.replace(/_/g,'$'); }
            try {
                if (el && el.click) { el.click(); return; }
            } catch(e){}
 
            if (typeof window.__doPostBack === 'function') {
                window.__doPostBack(idToName(el.id), '');
                return;
            }
            if (el.form) {
                try { el.form.submit(); return; } catch(e){}
            }
            if (document.createEventObject && el.fireEvent) {
                try {
                    var ev = document.createEventObject();
                    el.fireEvent('onmousedown', ev);
                    el.fireEvent('onmouseup', ev);
                    el.fireEvent('onclick', ev);
                    return;
                } catch(e){}
            }
            try { if (typeof el.onclick === 'function') el.onclick(); } catch(e){}
        })(arguments[0]);
    """, el)

def newClick(el):
    try:
        driver.execute_script("""
            (function(el){ 
                function idToName(id){ return id.replace(/_/g,'$'); } 
                try { el.click(); return; } catch(e){} 
                try { el.fireEvent && el.fireEvent('onclick'); return; } catch(e){} 
                try { if (typeof window.__doPostBack === 'function') 
                    window.__doPostBack(idToName(el.id), ''); 
                    } catch(e){} 
                })(arguments[0]); 
        """, el)
    except Exception as e: 
        print("JS click failed:", e) 
 
def newClickTry(el):
    global driver
    try:
        driver.execute_script("""
            var target = arguments[0];
            setTimeout(function(){
                try { target.click(); } catch(e) {}
            },0);
        """, el)
    except Exception as e:
        print("JS click failed:", e)
 
 
def ie_select_by_visible_text(by, locator, text, timeout=20, partial_ok=False):
    global driver
    el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, locator)))
    try:
        Select(el).select_by_visible_text(text)
        return True
    except Exception:
        pass
 
    js = r"""
     (function(sel, want, partial){
         function norm(s){return (s||"").replace(/\u00A0/g,' ').replace(/\s+/g,' ').trim();}
         var nWant = norm(want), hit = -1, exact = -1;
         for (var i=0; i<sel.options.length; i++){
             var t = norm(sel.options[i].text);
             if (t === nWant){ exact = i; break; }
             if (partial && hit === -1 && t.indexOf(nWant) > -1) hit = i;
         }
         var idx = (exact >= 0 ? exact : hit);
         if (idx < 0) return false;
         sel.selectedIndex = idx;
 
         try {
             if (document.createEventObject && sel.fireEvent) {
                 sel.fireEvent('onchange', document.createEventObject());
             } else if (document.createEvent) {
                 var evt = document.createEvent('HTMLEvents');
                 evt.initEvent('change', true, false);
                 sel.dispatchEvent(evt);
             } else if (typeof sel.onchange === 'function') {
                 sel.onchange();
             }
         } catch(_){}
         try {
             if (typeof window.__doPostBack === 'function' && sel.id){
                 function idToName(id){ return id.replace(/_/g,'$'); }
                 window.__doPostBack(idToName(sel.id), '');
             }
         } catch(_){}
         return true;
     })(arguments[0], arguments[1], arguments[2]);
    """
    return driver.execute_script(js, el, text, partial_ok)

def select_unknown_contact(full_address): 
    wait = WebDriverWait(driver, 30) 

    normalized_address = re.sub(r'\bP.?\s*O.?', 'PO ', full_address, flags=re.IGNORECASE) 
    
    # Keep only the part before the first comma (street line only) 
    first_line = normalized_address.split(',')[0].strip() 
    
    # Normalize spaces 
    first_line = re.sub(r'\s+', ' ', first_line) 
    
    # Find all location blocks 
    locations = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "Location"))) 
    
    for loc in locations: 
        address_text = loc.find_element(By.TAG_NAME, "pre").text.strip().replace("\n"," ") 
        scroll = loc.find_element(By.TAG_NAME, "pre")
        try: driver.execute_script("arguments[0].scrollIntoView(true);", scroll)
        except WebDriverException: pass
        
        # if first_line.lower() in address_text.lower(): 
        if PRC.addressMatch(full_address,address_text):
            # Found the right location 
            print(f"Matched address: {address_text}") 
    
            # Find sibling UL element with contacts 
        
            try: 
                contacts_ul = loc.find_element( By.XPATH, ".//following::ul[contains(@id,'Contacts')][1]" ) 
            except Exception: 
                print("Contact list not immediately found, retrying...") 
                time.sleep(2) 
                try: 
                    contacts_ul = driver.find_element(By.XPATH, "//ul[contains(@id,'Contacts')]") 
                except: 
                    print("Could not locate Contacts list for this address.") 
                    continue 

            contact_items = contacts_ul.find_elements(By.XPATH, ".//li[contains(@class,'Contact')]") 

            for contact in contact_items: 
                contact_name = contact.find_element(By.TAG_NAME, "pre").text.strip() 
                scroll = contact.find_element(By.TAG_NAME, "pre")
                try: driver.execute_script("arguments[0].scrollIntoView(true);", scroll)
                except WebDriverException: pass
                if contact_name.lower() == "unknown, unknown": 
                    # Found the Unknown contact — click its link 
                    try: 
                        link = contact.find_element(By.TAG_NAME, "a") 
                        # element_click(link)
                        newClickTry(link)
                        print("Clicked Unknown contact link.") 
                        return True 
                    except Exception: 
                        print("No clickable link found inside Unknown contact.") 
                        return False 
            
            # attorneysTable = driver.find_element(By.ID, "ctl00_Body_SearchResultsTable_Table")
            # attorneyTbody = attorneysTable.find_elements(By.XPATH, ".//tbody/tr")
            # for rowAtt in attorneyTbody:
            #     cells = rowAtt.find_elements(By.TAG_NAME, "td")
            #     if not cells:
            #         continue

            #     cmsAttorneyName = cells[1].text.strip()
            #     xlAttorneyFullName = f"{xlAttorneyLname}, {xlAttorneyFname}"

            #     if xlAttorneyFullName in cmsAttorneyName:
            #         links = rowAtt.find_elements(By.TAG_NAME, "a")
            #         element_click(links[-1])
    
    print("No matching address or Unknown contact found.") 
    return False 

def switch_to_new_window(old_handles, timeout=10):
    WebDriverWait(driver, timeout).until(
        lambda d: len(d.window_handles) > len(old_handles)
    )

    new_handles = set(driver.window_handles) - old_handles
    new_handles = new_handles.pop()
    driver.switch_to.window(new_handles)

def search_address( address: dict,  mailAdd: str):
    
    a1 = address["Address1"]
    a2 = address["Address2"]
    city = address["City"]
    State = address["State"]
    zip = address["Zip"]
    original = driver.current_window_handle
    old_handles = set(driver.window_handles)
    company = driver.find_element(By.NAME, "ctl00$Body$ctl00$ctl23")
    element_click(company)
    # newClick(company)
    time.sleep(3)

    
    switch_to_new_window(old_handles, timeout=15)
    time.sleep(2)

    legacy_safe_type(By.ID, "ctl00_Body__Name", "Liberty Mutual Commercial Market")
    search_btn = driver.find_element(By.XPATH, "//input[@type='submit' and @value='Search']") 
    element_click(search_btn)

    time.sleep(2)
    result_table = driver.find_element(By.ID, "ctl00_Body__ResultTable_Table")
    result_trs = result_table.find_elements(By.XPATH, ".//tbody/tr")

    for i_row in result_trs:
        result_cells = i_row.find_elements(By.TAG_NAME, "td")
        if len(result_cells) < 2:
            continue
        if result_cells[1].text.strip() ==  "Liberty Mutual Commercial Market":
            select_btn = result_cells[7].find_element(By.TAG_NAME, "a")
            element_click(select_btn)
            break
        
    time.sleep(3)

    legacy_safe_type(By.ID, "ctl00_Body__FilterCity", city)

    filter_btn = driver.find_element(By.XPATH, "//input[@type='submit' and @value='Filter List']")
    element_click(filter_btn)

    time.sleep(3)

    location_table = driver.find_element(By.ID, "ctl00_Body__LocationTable_Table")
    location_trs = location_table.find_elements(By.XPATH, ".//tbody/tr")

    for i_row in location_trs:
        location_cells = i_row.find_elements(By.TAG_NAME, "td")
        if len(location_cells) < 2:
            continue
        try: driver.execute_script("arguments[0].scrollIntoView(true);", i_row)
        except WebDriverException: pass
        location_Add1 = location_cells[0].text.strip().lower()

        print(f"{location_Add1} {mailAdd.lower()}")  
        import ProviderResultChecker as PSC
        TempAddress = PSC.expand_suffix_long(location_Add1)
        RealAddress = PSC.expand_suffix_long(mailAdd)
        ComparisonResult = PSC.addressMatch(TempAddress,RealAddress)
        if location_Add1 in mailAdd.lower() or ComparisonResult:
            select_btn = location_cells[5].find_element(By.TAG_NAME, "a")
            element_click(select_btn)
            break

    time.sleep(2)

    driver.switch_to.window(original)

    save_btn = driver.find_element(By.XPATH, "//input[@value='Save']") 
    element_click(save_btn)
    time.sleep(2)
 
def ie_select_by_value(by, locator, value, timeout=20):
    global driver
    el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, locator)))
    try:
        Select(el).select_by_value(value)
        return True
    except Exception:
        return driver.execute_script(
            """
            return (function(sel,v){
                for(var i=0;i<sel.options.length;i++){
                    if(sel.options[i].value==v||sel.options[i].text==v){
                        sel.selectedIndex=i;
                        sel.onchange&&sel.onchange();
                        return true;
                    }
                }
                return false;
            })(arguments[0], arguments[1]);
            """,
            el,
            value
        )
 
 
def splitMailingAdd(address):
    if not address or not address.strip():
        return ""
    s = " ".join(address.split())
    s = re.sub(r"P\.?\s*O\.?", "PO", s, flags=re.IGNORECASE)
    m = re.match(r"(PO Box \d+)", s, flags=re.IGNORECASE)
    if m:
        return m.group(1)
    return s.replace(".", "").replace(",", "").strip()
 
CITY_PREFIXES = {
    "new", "los", "las", "san", "santa",
    "fort", "ft", "st", "saint", "port",
    "white", "lake", "north", "south", "east", "west",
    "mt", "mount"
}
 
def split_address(address: str):
    if not address:
        return "", "", "", "", ""
 
    # Normalize spaces and strip trailing commas/spaces
    s = re.sub(r"\s+", " ", address.replace("\u00A0", " ")).strip(" ,")
 
    # Allow optional comma after state and no hard requirement that ZIP is last char
    m_city = re.search(
        r"([A-Za-z .'-]+)\s*,\s*([A-Z]{2})\s*,?\s*(\d{5})(?:-\d{4})?$",
        s,
        re.I,
    )
    if not m_city:
        return "", "", "", "", ""
 
    before = s[:m_city.start(1)].rstrip(" ,")
    city_block = m_city.group(1).strip()
    state = m_city.group(2).upper()
    zip5 = m_city.group(3)[:5]
 
    parts = city_block.split()
    if not parts:
        return "", "", "", "", ""
 
    # 3-word city
    if len(parts) >= 3 and parts[-3].lower() in CITY_PREFIXES:
        city = " ".join(parts[-3:])
        rest_parts = parts[:-3]
    # 2-word city
    elif len(parts) >= 2 and parts[-2].lower() in CITY_PREFIXES:
        city = " ".join(parts[-2:])
        rest_parts = parts[:-2]
    else:
        city = parts[-1]
        rest_parts = parts[:-1]
 
    rest_for_address = " ".join(rest_parts).strip()
    addr_all = (before + " " + rest_for_address).strip()
 
    addr1, addr2 = addr_all, ""
    if "," in addr_all:
        a1, a2 = addr_all.rsplit(",", 1)
        addr1, addr2 = a1.strip(), a2.strip()
 
    # return addr1, addr2, city, state, zip5
    return {
        "Address1":addr1,
        "Address2":addr2,
        "City":city,
        "State":state,
        "Zip":zip5
    }

 
def formataddrSearchStr (s: str) -> str:
    # your existing formatter – placeholder
    return " ".join(str(s).split())
 
def extract_address_by_identifier(filename, target_id):
    ensure_readable_path(filename, "Office matrix Excel")
    wb = openpyxl.load_workbook(filename)
    ws1 = wb['Liberty and Middle Market']
    max_row = ws1.max_row
 
    current_id = None
    collecting = False
    physical_lines = []
    mailing_lines = []
 
    for i in range(6, max_row + 1):
        cell_id = ws1.cell(row=i, column=1).value
 
        if cell_id not in (None, ""):
            current_id = str(cell_id).strip()
            collecting =  str(target_id).upper() in current_id.upper()
            if collecting:
                physical_lines = []
                mailing_lines = []
 
        if collecting:
            physical1 = ws1.cell(row=i, column=7).value
            mailing1 = ws1.cell(row=i, column=8).value
 
            physical_line = " ".join(filter(None, [physical1]))
            mailing_line = " ".join(filter(None, [mailing1]))
 
            if physical_line:
                physical_lines.append(physical_line)
            if mailing_line:
                mailing_lines.append(mailing_line)
 
    physAddress = mailAddress = ""
    if physical_lines or mailing_lines:
        physAddress = formatStr(" ".join(physical_lines))
        mailAddress = formatStr(" ".join(mailing_lines))
    else:
        print(f"No address found for ID {target_id}")
    
    if physAddress:
        wb.close()
        return physAddress , mailAddress
    

    # 2) Sheet: "LM Physical Address"
    try:
        ws2 = wb['LM Physical Address']
        max_row2 = ws2.max_row
 
        for i in range(2, max_row2 + 1):  # starts row 2
            cell_id = ws2.cell(row=i, column=1).value  # ID in col A
            if cell_id is None:
                continue
 
            if target_id.upper() in str(cell_id).strip().upper() :
                addr_value = ws2.cell(row=i, column=6).value  # address in col F
                if addr_value:
                    physAddress = formataddrSearchStr (str(addr_value))
                    mailAddress = physAddress
                break
    except KeyError:
        pass
 
    if physAddress:
        wb.close()
        return physAddress , mailAddress
 
    # 3) Sheet: "Manual Tracking LM Claim Office"
    try:
        ws3 = wb['Manual Tracking LM Claim Office']
        max_row3 = ws3.max_row
 
        for i in range(1, max_row3 + 1):  # starts row 1
            cell_val = ws3.cell(row=i, column=1).value  # ID + address in col A
            if not cell_val:
                continue
            
            text = str(cell_val).strip()
            # if "-" not in text or "–" not in text:
            #     continue
            
            text = text.replace("–", "-").replace("—", "-")

            if "-" in text:
                id_part, addr_part = text.split("-", 1)

            
            if target_id.upper() in id_part.strip().upper():
                physAddress = formataddrSearchStr (addr_part.strip())
                mailAddress = physAddress
                break
    except KeyError:
        pass
 
    wb.close()
 
    if not physAddress:
        print(f"No physical address found for ID {target_id}")

    # wb1.close()
    return physAddress, mailAddress
 
 
# def send_address_via_ahk_v2(address: dict, mailAdd: str):
#     ensure_executable_path(AHK_EXE, "AutoHotkey")
#     ensure_readable_path(AHK_CLAIMTAB_SCRIPT, "ClaimTab AHK script")
 
#     a1 = address["Address1"]
#     a2 = address["Address2"]
#     city = address["City"]
#     State = address["State"]
#     zip_code = address["Zip"]
 
#     subprocess.run([AHK_EXE, AHK_CLAIMTAB_SCRIPT, a1, a2, city, State, zip_code, mailAdd], check=False)
 
 
# def searchEmployer(custName):
#     ensure_executable_path(AHK_EXE, "AutoHotkey")
#     ensure_readable_path(AHK_EMPSEARCH_SCRIPT, "empSearch AHK script")
#     subprocess.run([AHK_EXE, AHK_EMPSEARCH_SCRIPT, custName], check=False)
 
 
def SplitName(full_name):
    parts = full_name.strip().split()
    if len(parts) >= 2:
        return parts[0], parts[-1]
    elif len(parts) == 1:
        return parts[0], ""
    else:
        return "", ""
 
 
# ===================== LOGIN =====================
 
def login(cms_user: str, cms_pass: str):
    global driver
    driver.get(CMS_LOGIN_URL)
    driver.maximize_window()
    notify("Login", "Please Login your genex account")
    legacy_safe_type(By.ID, "ctl00_Body_UserName", cms_user)
    time.sleep(1)
    legacy_safe_type(By.ID, "ctl00_Body_Password", cms_pass)
    time.sleep(2)
 
    iAccept = driver.find_element(By.ID, "ctl00_Body_Accept")
    element_check(iAccept)
    driver.execute_script("AcceptChanged();")
 
    loginBtn = driver.find_element(By.ID, "ctl00_Body_Login")
    element_click(loginBtn)
 
    time.sleep(8)
 
 
# ===================== MAIN CASE LOGIC =====================
 
def searchCase(connection_string: str, claimant_name: str, claim_number: str,caseIDSelect:str):
    """
    Uses stored procedure to get case data, then drives CMS UI to update the case.
    """
    global driver
    global caseID
    dxCode_list = []
    bodyParts = []
 
    # 1) Get data from DB instead of Excel
    case_data = fetch_case_data_from_db(connection_string, claimant_name, claim_number)
 
    claimNUm = case_data.claim_number
    dxCode = case_data.dxCode
    bodyParts = case_data.bodyParts
    officeNumber = case_data.officeNumber
    xlReferralSource = case_data.xlReferralSource
    xlCustomerName = case_data.xlCustomerName
    xlProvider = case_data.xlProvider
    provType = case_data.provType
    xlCustomerEmail = case_data.xlCustomerEmail
    xlInsuredAddr1 = case_data.xlInsuredAddr1
    xlAttorney = case_data.xlAttorney
    xlAttorneyAdd1 = case_data.xlAttorneyAdd1
    xlAttorneyAdd2 = case_data.xlAttorneyAdd2
    xlAttorneyCity = case_data.xlAttorneyCity
    xlAttorneyState = case_data.xlAttorneyState
    xlAttorneyZip = case_data.xlAttorneyZip
    xlAttorneyPhone = case_data.xlAttorneyPhone
    apptDate = case_data.apptDate
    refType = case_data.refType
    xlProviderPhone = case_data.xlProviderPhone
    Special_Instructions = case_data.Special_Instructions
    try:
        # --- Go to Case Management → Claimant Profile ---
        driver.execute_script("h$(0);")
        time.sleep(1)
        main_window = driver.current_window_handle
        driver.execute_script("h$(10);")
        time.sleep(1)
 
        claimantPfElem = driver.find_element(By.ID, "mmlink1")
        element_click(claimantPfElem)
        time.sleep(4)
 
        WebDriverWait(driver, 20).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, "/html/body/form/div[5]/iframe"))
        )
        
        legacy_safe_type(By.ID, "ctl00_Body_CaseNumber", caseIDSelect)
        time.sleep(1)
        legacy_safe_type(By.ID, "ctl00_Body_ClaimNumber", claimNUm)
        searchClaimBtn = driver.find_element(By.ID, "ctl00_Body_SearchButton")
        element_click(searchClaimBtn)

        ######################-----------------------------------------------
        # latest_date = None
        # latest_ow_index = None
        # time.sleep(4)
 
        # # --- Search result table, pick latest AssignedDate ---
        # SelectedCaseID = ""
        # if elementExist(By.XPATH, "/html/body/form/div[3]/table[2]/tbody/tr/td[6]/a"):
        #     tableClaim = driver.find_element(By.ID, "ctl00_Body_SearchResultsTable_Table")
        #     claimTrs = tableClaim.find_elements(By.TAG_NAME, "tr")
 
        #     for row_index in range(1, len(claimTrs)):
        #         tableClaim = driver.find_element(By.ID, "ctl00_Body_SearchResultsTable_Table")
        #         claimTrs = tableClaim.find_elements(By.TAG_NAME, "tr")
        #         claimLink = claimTrs[row_index].find_element(By.TAG_NAME, "a")
        #         element_click(claimLink)
        #         time.sleep(3)
 
        #         table = driver.find_element(By.ID, "ctl00_Body_ClaimantProfileTable_Table")
        #         try:
        #             driver.execute_script("arguments[0].scrollIntoView(true);", table)
        #         except WebDriverException:
        #             pass
        #         rows = table.find_elements(By.TAG_NAME, "tr")
 
        #         for i in range(len(rows)):
        #             claimTxt = rows[i].text.strip()
        #             if claimNUm in claimTxt:

        #                 CaseRows = i +1
        #                 while CaseRows < len(rows):
        #                     try:
        #                         caseRow = rows[CaseRows]
        #                         caseTxt = caseRow.text.strip()

        #                         caseSplits = " ".join((caseTxt or "").split())
        #                         caseParts = caseSplits.split()
        #                         if "case" in caseTxt[:4].lower():
        #                             caseID = caseParts[2]
        #                     except StaleElementReferenceException:
        #                         table = driver.find_element(By.ID, "ctl00_Body_ClaimantProfileTable_Table")
        #                         rows = table.find_elements(By.TAG_NAME, "tr")
        #                         caseRow = rows[CaseRows]
        #                         caseTxt = caseRow.text.strip()

        #                         caseSplits = " ".join((caseTxt or "").split())
        #                         caseParts = caseSplits.split()
        #                         if "case" in caseTxt[:4].lower():
        #                             caseID = caseParts[2]

        #                     if not caseTxt.lower().startswith("case"):
        #                         break

        #                     # notify("Next", "Press ok to Continue")
        #                     # aclick = driver.find_element(
        #                     #     By.XPATH,
        #                     #     f"/html/body/form/div[3]/table/tbody/tr[{i + 2}]/td[3]/a"
        #                     # )

        #                     aclick = caseRow.find_element(
        #                         By.XPATH,
        #                         ".//a[normalize-space()='Select Case']"
        #                     )

        #                     element_click(aclick)
        #                     time.sleep(3)
        #                     caseAssignedDatee = driver.find_element(
        #                         By.ID,
        #                         "ctl00_Body_AssignedDateField_DatePickerTextBox"
        #                     )
        #                     driver.execute_script("arguments[0].scrollIntoView(true);", caseAssignedDatee)
        #                     caseAssignedDate_str = driver.find_element(
        #                         By.ID,
        #                         "ctl00_Body_AssignedDateField_DatePickerTextBox"
        #                     ).get_attribute("value")
                            
        #                     caseAssignedDate = datetime.strptime(
        #                         caseAssignedDate_str, "%d-%b-%Y"
        #                     )
        #                     if latest_date is None or caseAssignedDate > latest_date:
        #                         latest_date = caseAssignedDate
        #                         latest_ow_index = row_index
        #                         SelectedCaseID = caseID

                            
                            
                            
        #                     if latest_date == None:
        #                         return

        #                     scroll = driver.find_element(By.ID, "udf_2_1329_2436")
        #                     try:
        #                         driver.execute_script("arguments[0].scrollIntoView(true);", scroll)
        #                     except WebDriverException:
        #                         pass
        #                     claimaintProfile = driver.find_element(By.ID, "ctl00_Body_ClaimantProfileLink")
        #                     element_click(claimaintProfile)
        #                     time.sleep(2)
        #                     CaseRows += 1
        #                 claimaintListing = driver.find_element(By.XPATH, "/html/body/form/div[3]/div[2]/a")
        #                 element_click(claimaintListing)
        #                 time.sleep(4)
        #                 break

                            
 
        #     if latest_ow_index is not None:
        #         tableClaim = driver.find_element(By.ID, "ctl00_Body_SearchResultsTable_Table")
        #         rowsClaim = tableClaim.find_elements(By.TAG_NAME, "tr")
        #         target_row = rowsClaim[latest_ow_index]
        #         claim_link = target_row.find_element(By.TAG_NAME, "a")
        #         element_click(claim_link)
        # else:
        #     print("No claim found")
        #     return
 
        # time.sleep(5)
        
 
        # --- Claimant profile again to find “Case” row ---
        # table = driver.find_element(By.ID, "ctl00_Body_ClaimantProfileTable_Table")
        # try:
        #     driver.execute_script("arguments[0].scrollIntoView(true);", table)
        # except WebDriverException:
        #     pass
        # rows = table.find_elements(By.TAG_NAME, "tr")
 
        # for i in range(len(rows)):
        #     claimTxt = rows[i].text.strip()
        #     if caseIDSelect in claimTxt:
        #         caseRow = rows[i]
        #         aclick = caseRow.find_element(
        #                         By.XPATH,
        #                         ".//a[normalize-space()='Select Case']"
        #                     )
                
        #         element_click(aclick)
        #         time.sleep(3)
        ######################-----------------------------------------------
        # ================== CLAIM TAB / DX CODE ==================
        claim_tab = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable(
                (By.XPATH,
                    "/html/body/form/div[3]/div[1]/table/tbody/tr[1]/td[2]/table[2]/tbody/tr/td[2]/a")
            )
        )
        element_click(claim_tab)
        #- Add Prompter Yes No
        # notify("Notify","Dx Code for Manual adding. Click Ok once Dx Code is added or N/A")
        if dxCode:
            dxclick = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.ID, "ctl00_Body_ClaimDiagnosisTable_ctl00_AddLinkButton")
                )
            )
            element_click(dxclick)
            legacy_safe_type(By.ID, "ctl00_Body_ClaimDiagnosis1_DiagnosisCode", dxCode)
            time.sleep(5)
            saveScroll = driver.find_element(
                By.XPATH, "/html/body/form/div[3]/div[3]/div/input[1]"
            )
            try:
                driver.execute_script("arguments[0].scrollIntoView(true);", saveScroll)
                newClickTry(saveScroll)
                time.sleep(5)
            except WebDriverException:
                pass

        # ================== BODY PARTS ==================
        time.sleep(3)
        tblScroll = driver.find_element(
            By.XPATH, "/html/body/form/div[3]/div[3]/div[2]/div/div/div[2]/a"
        )
        try:
            driver.execute_script("arguments[0].scrollIntoView(true);", tblScroll)
        except WebDriverException:
            pass

        tableBdPrts = driver.find_element(
            By.ID, "ctl00_Body_ClaimBodyParts1_Section_Table_Table"
        )
        try:
            driver.execute_script("arguments[0].scrollIntoView(true);", tableBdPrts)
        except WebDriverException:
            pass

        # trBdPrts = tableBdPrts.find_elements(By.XPATH, ".//tbody/tr")
        # for rowBdprts in trBdPrts:
        #     bdPrtsTr = rowBdprts.text.strip()
        #     bdPrtsTr = bdPrtsTr.split()[0] if bdPrtsTr else None
        #     if bdPrtsTr:
        #         bodyParts.append(bdPrtsTr)
        if bodyParts:
            addBodyParts = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "ctl00_Body_ClaimBodyParts1_Section_AddLinkButton")))
            driver.execute_script("arguments[0].scrollIntoView(true);", addBodyParts)
            element_click(addBodyParts)
            time.sleep(5)
            # bodyPartDropdown = Select(driver.find_element(By.ID,"ctl00_Body_BodyPartDropDown"))
            ie_select_by_visible_text(By.ID, "ctl00_Body_BodyPartDropDown", bodyParts)
            # bodyPartDropdown.select_by_visible_text(bodyParts)
            time.sleep(5)
            saveScroll = driver.find_element(
                By.ID, "ctl00_Body_ButtonSave"
            )
            try:
                driver.execute_script("arguments[0].scrollIntoView(true);", saveScroll)
                newClickTry(saveScroll)
                time.sleep(5)
            except WebDriverException:
                pass


        # ================== OFFICE ADDRESS COMPARISON ==================
        try:
            # officeNumber = "868"
            physAdd, mailAdd = extract_address_by_identifier(LIBERTY_OFFICE_MATRIX, officeNumber)
            
            mailAdd_CMS = driver.find_element(
                By.ID, "ctl00_Body_ClaimAdjuster_ctl00_MailingAddressCell"
            ).text.strip()
            
            physAdd_CMS = driver.find_element(
                By.ID, "ctl00_Body_ClaimAdjuster_ctl00_PhysicalAddressCell"
            ).text.strip()

            mailAdd_CMS = formatStr(mailAdd_CMS)
            physAdd_CMS = formatStr(physAdd_CMS)
            similarity = difflib.SequenceMatcher(
                None, physAdd.lower(), physAdd_CMS.lower()
            ).ratio()

            if mailAdd:
                similarityMail = difflib.SequenceMatcher(
                    None, mailAdd.lower(), mailAdd_CMS.lower()
                ).ratio()
            else:
                similarityMail = 0

            if similarity >= 0.9:
                print("Same Office Address")
            else:
                print("Different Office Address – updating via AHK")
                editclick = driver.find_element(By.ID, "ctl00_Body_ClaimAdjuster_ctl00_EditLink")#.click()
                try: driver.execute_script("arguments[0].scrollIntoView(true);", editclick)
                except WebDriverException: pass

                newClick(editclick)
                splitPhyAdd = split_address(physAdd)
                splitMailing = splitMailingAdd(mailAdd)

                all_windows = driver.window_handles
                for handle in all_windows:
                    if handle != main_window:
                        try:
                            driver.switch_to.window(handle)
                            break
                        except (NoSuchWindowException, WebDriverException):
                            pass
                time.sleep(2)
                search_address(splitPhyAdd, splitMailing)
                time.sleep(2)
                driver.switch_to.window(main_window)
                driver.switch_to.default_content
                time.sleep(3)
                WebDriverWait(driver, 20).until(
                    EC.frame_to_be_available_and_switch_to_it((By.XPATH, "/html/body/form/div[5]/iframe"))
                )
        except:
            notify('Manual Process','Unable to determine Address. Kindly Check Adjuster Mailing and  Physical Address\n Click Ok once address is corrected or verified.')

            driver.switch_to.window(main_window)
            driver.switch_to.default_content
            time.sleep(3)
            WebDriverWait(driver, 20).until(
                EC.frame_to_be_available_and_switch_to_it((By.XPATH, "/html/body/form/div[5]/iframe"))
            )
            
            pass
        
        # ================== CLAIMANT TAB ==================
        # notify("Next", "Proceed to Claimant Tab")
        claimantTab = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[3]/div[1]/table/tbody/tr[2]/td[2]/table[1]/tbody/tr/td[2]/a"))
        )#.click()
        element_click(claimantTab)
        time.sleep(5)
        
        # ================== CASE TAB ==================
        # notify("Next", "Proceed to Case Tab")
        elemClick = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[3]/div[1]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td[2]/a"))
        )#.click()
        element_click(elemClick)
        wbSup = openpyxl.load_workbook("Liberty Mutual Medical Dedicated Supervisor Mapping 04-07-2025.xlsx")
        wsSup = wbSup.active 
        time.sleep(10)
        

        ReferralSourceField = driver.find_element(By.ID,"ctl00_Body_ReferralSourceField")
        RefSourceVal = (ReferralSourceField.get_attribute("value") or "").strip()

        if not RefSourceVal:
            ReferralSourceField.clear()
            legacy_safe_type(By.ID, "ctl00_Body_ReferralSourceField", xlReferralSource)

        dropdown = Select(driver.find_element(By.ID, "ctl00_Body_ReferralTypeField"))
        caseReferralType = dropdown.first_selected_option.text

        if "med" in refType.lower():
            ie_select_by_visible_text(By.ID, "ctl00_Body_InventoryTypeField", "Medical - FEE")  #migz
        elif "voc" in refType.lower():
            ie_select_by_visible_text(By.ID, "ctl00_Body_InventoryTypeField", "Vocational - FEE") 
        time.sleep(1)
        caseDateofFirstAppointment = driver.find_element(By.ID, "udf_2_1329_2436")#.get_attribute("value") 
        # if apptDate:
        #     driver.execute_script("arguments[0].scrollIntoView(true);", caseDateofFirstAppointment)
        #     legacy_safe_type(By.ID, "udf_2_1329_2436", apptDate)
        time.sleep(1)
        caseClaimNumber = driver.find_element(By.ID, "udf_1_1329_3081")#.get_attribute("value") 
        # if claimNUm:
        #     driver.execute_script("arguments[0].scrollIntoView(true);", caseClaimNumber)
        #     legacy_safe_type(By.ID, "udf_1_1329_3081", claimNUm)

        time.sleep(5)

        #adjusted logic
        caseNumPrefix = driver.find_element(By.ID, "ctl00_Body_CaseNumberField")
        existingWindow = driver.window_handles
        caseSup = driver.find_element(By.ID, "ctl00_Body_CaseSupervisorNameField").get_attribute("value")
        time.sleep(5)
        if not caseSup:
            caseSupXl = "" 
            print(caseNumPrefix.text.strip()[:2])
            for rowSup in  wsSup.iter_rows(min_row=6, min_col=1, max_col=5):
                if rowSup[0].value == caseNumPrefix.text.strip()[:2]:
                    caseSupXl = rowSup[1].value
                    break
            
            caseSupXlLName = caseSupXl.split()[0] if caseSupXl else None
            caseSupXlFName = caseSupXl.split()[1] if caseSupXl else None
            time.sleep(3)
            SearchSup = driver.find_element(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[9]/td[2]/a")
            element_click(SearchSup)  
            time.sleep(3)
            newWindows = driver.window_handles
            for handle in newWindows:
                if handle not in existingWindow:
                    driver.switch_to.window(handle)
                    break
            
            supTable = driver.find_element(By.ID, "ctl00_Body_Results_Table")
            supTrs = supTable.find_elements(By.XPATH, ".//tbody/tr")

            for rowAttach in supTrs:
                driver.execute_script("arguments[0].scrollIntoView(true);", rowAttach)
                cells = rowAttach.find_elements(By.TAG_NAME, "td")
                assignElem = cells[2].find_element(By.TAG_NAME, "a")
                if not cells:
                    continue
                attachFrom = cells[0].text.strip()
                if caseSupXlLName in attachFrom and caseSupXlFName in attachFrom :
                    WebDriverWait(driver, 30).until(lambda d: d.execute_script("return document.readyState") in ("complete", "interactive"))
                    time.sleep(1)
                    newClickTry(assignElem)
                    time.sleep(5)
                    break
                    
        wbSup.close()
        time.sleep(5)
        driver.switch_to.window(driver.window_handles[0])
        driver.switch_to.default_content
        time.sleep(3)
        WebDriverWait(driver, 20).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, "/html/body/form/div[5]/iframe"))
        )
        #

        SaveSup = driver.find_element(By.ID, "ctl00_Body_SaveButton")
        driver.execute_script("arguments[0].scrollIntoView(true);", SaveSup)
        #SaveSup.click() # migz change to element_click
        element_click(SaveSup)
        time.sleep(5)
        # messagebox.showinfo("Next", "Press ok to Continue")
        # ================== Customer TAB ==================
        # notify("Next", "Proceed to Customer Tab")
        elemClick = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[3]/div[1]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td[2]/a"))
        )#.click()
        element_click(elemClick)

        time.sleep(5)
        # customerName = driver.find_element(By.ID, "ctl00_Body_CustomerSelector_InsuredName").get_attribute("value")
        customerCheckName = driver.find_element(By.ID,"EmployerField").text
        # if customerCheckName == "Not Given / Not Applicable":
        #     # driver.find_element(By.ID, "ctl00_Body_CustomerSelector_InsuredName").send_keys(xlCustomerName)
        #     legacy_safe_type(By.ID, "ctl00_Body_CustomerSelector_InsuredName", xlCustomerName)
        xlAddress1 = ""
        # xlAddress1 = row[54].value
        
        # if not xlAddress1:
        #     legacy_clear(By.ID, "ctl00_Body_CustomerSelector_InsuredStreet1")
        #     legacy_clear(By.ID, "ctl00_Body_CustomerSelector_InsuredStreet2")
        #     legacy_clear(By.ID, "ctl00_Body_CustomerSelector_InsuredCity")
        #     legacy_clear(By.ID, "ctl00_Body_CustomerSelector_InsuredState")
        #     legacy_clear(By.ID, "ctl00_Body_CustomerSelector_InsuredPostalCode")
        customerCheckName = driver.find_element(By.ID,"EmployerField").text
        
        if customerCheckName == "Not Given / Not Applicable": #not customerName: #or customerCheckName == "Not Given / Not Applicable"
            customerName = driver.find_element(By.XPATH, "/html/body/form/div[3]/table[1]/tbody/tr[1]/td[2]/a")
            newClick(customerName)
            xlCustomerName = xlCustomerName.replace(",", "")
            time.sleep(5)
            driver.switch_to.window(driver.window_handles[-1])
            time.sleep(5)
            # empFull = driver.find_element(By.ID,"ctl00_Body_Name")
            clean_customer = re.sub(r"[^A-Za-z0-9 &]+", "", xlCustomerName)
            legacy_safe_type(By.ID, "ctl00_Body_Name", clean_customer)
            time.sleep(3)
            empSearch = driver.find_element(By.ID,"ctl00_Body_Search")
            element_click(empSearch)
            supTables = driver.find_elements(By.ID, "ctl00_Body_SearchResults")

            elements = driver.find_elements(By.TAG_NAME,"a")
            if len(elements) == 1:
                for el in elements:
                    try:
                        el_text = el.text
                        # el_href = get_attribute("id")
                        if el_text == "Assign":
                            element_click(el)
                            break
                        # outerHTML = el.get_attribute("outerHTML")
                        print("Element A:" , el_text)
                    except Exception:
                            continue
            elif len(elements) > 1:
                notify("Notice","Please verify the results. Please manually select correct customer.")
            elif len(elements) == 0:
                notify("Notice","Kindly validate Search Criteria. Please manually select correct customer.")
            driver.switch_to.window(driver.window_handles[0])
            driver.switch_to.default_content()
            WebDriverWait(driver, 30).until(
                EC.frame_to_be_available_and_switch_to_it((By.XPATH, "/html/body/form/div[5]/iframe"))
            )
        time.sleep(5)

        #----------------------------------------------------------- Bill To Start Customer
        # billToName = driver.find_element(By.ID, "ctl00_Body_BillToDetail_Content_CustomerName")
        # billToName = billToName.text.strip()
        # # if billToName != "Liberty Mutual Commercial Market":
        # billToAdd = driver.find_element(By.ID, "ctl00_Body_BillToDetail_Content_Address1")

        # assignBillTo = driver.find_element(By.ID, "ctl00_Body_BillToDetail_Content_AssignPrimaryLocationLink")
        # element_click(assignBillTo)

        # WebDriverWait(driver, 60).until(lambda d: len(d.window_handles) > 1)
        # driver.switch_to.window(driver.window_handles[-1])

        # # match = re.search(r",\s*([A-Za-z\s]+?),\s*[A-Za-z]{2}\s*\d{5}", mailAdd,re.IGNORECASE) 
        # match = re.search(r"([A-Za-z\s]+),\s*[A-Za-z]{2}\s*\d{5}(?:-\d{4})?", mailAdd,re.IGNORECASE) 
        # if match: 
        #     city_name = match.group(1).strip() 
        # else:
        # # Fallback: assume 2nd to last comma-separated part is the city 
        #     parts = [p.strip() for p in mailAdd.split(",")] 
        #     city_name = parts[-2] if len(parts) >= 2 else parts[-1] 

        # time.sleep(10)
        # # citySearch = driver.find_element(By.ID, "ctl00_Body_City")
        # legacy_safe_type(By.ID, "ctl00_Body_City", city_name)
        # btnSearch = driver.find_element(By.ID, "ctl00_Body_Search")
        # element_click(btnSearch)
        # time.sleep(3)
        # print(mailAdd)
        # select_unknown_contact(mailAdd)
        # time.sleep(3)
        # driver.switch_to.window(driver.window_handles[0])
        # driver.switch_to.default_content()
        # WebDriverWait(driver, 30).until(
        #     EC.frame_to_be_available_and_switch_to_it((By.XPATH, "/html/body/form/div[5]/iframe"))
        # )
        #------------------------------------------------------- Bill to End Customer

        #Liberty Bill To---------------------------------------------------------
        billToName = driver.find_element(By.ID, "ctl00_Body_BillToDetail_Content_AssignPrimaryLocationLink")
        billToName = billToName.text.strip()
        # if billToName != "Liberty Mutual Commercial Market":
        billToAdd = driver.find_element(By.ID, "ctl00_Body_BillToDetail_Content_Address1")

        assignBillTo = driver.find_element(By.ID, "ctl00_Body_BillToDetail_Content_AssignPrimaryLocationLink")
        element_click(assignBillTo)

        WebDriverWait(driver, 60).until(lambda d: len(d.window_handles) > 1)
        driver.switch_to.window(driver.window_handles[-1])

        # match = re.search(r",\s*([A-Za-z\s]+?),\s*[A-Za-z]{2}\s*\d{5}", mailAdd,re.IGNORECASE) 
        match = re.search(r"([A-Za-z\s]+),\s*[A-Za-z]{2}\s*\d{5}(?:-\d{4})?", mailAdd,re.IGNORECASE) 
        if match: 
            city_name = match.group(1).strip() 
        else:
        # Fallback: assume 2nd to last comma-separated part is the city 
            parts = [p.strip() for p in mailAdd.split(",")] 
            city_name = parts[-2] if len(parts) >= 2 else parts[-1] 

        time.sleep(10)
        # citySearch = driver.find_element(By.ID, "ctl00_Body_City")
        legacy_safe_type(By.ID, "ctl00_Body_City", city_name)
        btnSearch = driver.find_element(By.ID, "ctl00_Body_Search")
        element_click(btnSearch)
        time.sleep(3)
        print(mailAdd)
        select_unknown_contact(mailAdd)
        time.sleep(3)
        driver.switch_to.window(driver.window_handles[0])
        driver.switch_to.default_content()
        WebDriverWait(driver, 30).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, "/html/body/form/div[5]/iframe"))
        )

        #

        # messagebox.showinfo("Next", "Press ok to Continue")
        #Providers Taab
        # ================== PROVIDER TAB ==================
        # notify("Next", "Proceed to Providers Tab")
        elemClick = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[3]/div[1]/table/tbody/tr[2]/td[2]/table[8]/tbody/tr/td[2]/a"))
        )#.click()
        element_click(elemClick)
        time.sleep(3)
        
        RowProvCounts = driver.find_elements(By.XPATH,"/html/body/form/div[3]/table/tbody/tr[2]/td/table/tbody/tr")
        

        if not elementExist(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/a"):
            elemClick = driver.find_element(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[1]/td/a")#.click()
            element_click(elemClick)
            time.sleep(3)
            if provType == "facility":
                    elemClick = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "ctl00_Body_FacilityName")))
                    legacy_clear(By.ID, "ctl00_Body_FacilityName")
                    legacy_safe_type(By.ID, "ctl00_Body_FacilityName", xlProvider[:4])
            else:
                elemClick = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "ctl00_Body_LastName")))
                legacy_clear(By.ID, "ctl00_Body_LastName")
                legacy_safe_type(By.ID, "ctl00_Body_LastName", xlProvider[:4])
            time.sleep(.5)
            legacy_safe_type(By.ID, "ctl00_Body_Phone_PNumber",xlProviderPhone)

            elemClick = driver.find_element(By.ID, "ctl00_Body_SearchButton")
            try: driver.execute_script("arguments[0].scrollIntoView(true);", elemClick)
            except WebDriverException: pass
            element_click(elemClick)
            time.sleep(3)
            notify("Notice","For Manual Process on Selecting Provider. Click ok Once Provider is Selected")
        # else:
        #     cmsProvider = driver.find_element(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/a")
        #     cmsProvider = cmsProvider.text.strip()
        #     # xlProvider = row[37].value
        #     print(f"{cmsProvider} {xlProvider}")
        #     # if cmsProvider != xlProvider:
        #     if len(RowProvCounts) > 1:
        #         # elemClick = driver.find_element(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[2]/td/table/tbody/tr/td[14]/a")#.click()
        #         # element_click(elemClick)
        #         # time.sleep(3)
        #         elemClick = driver.find_element(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[2]/td/table/tbody/tr/td[15]/a")# remove
        #         element_click(elemClick)
        #         time.sleep(3)

        #         elemClick = driver.find_element(By.ID, "ctl00_Body_Question_YesButton")
        #         element_click(elemClick)
        #         time.sleep(3)
                

        #         elemClick = driver.find_element(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[1]/td/a")
        #         element_click(elemClick)
        #         time.sleep(3)

        #         if provType == "facility":
        #             elemClick = WebDriverWait(driver, 10).until(
        #             EC.presence_of_element_located((By.ID, "ctl00_Body_FacilityName")))
        #             legacy_clear(By.ID, "ctl00_Body_FacilityName")
        #             legacy_safe_type(By.ID, "ctl00_Body_FacilityName", xlProvider[:3])
        #         else:
        #             elemClick = WebDriverWait(driver, 10).until(
        #             EC.presence_of_element_located((By.ID, "ctl00_Body_LastName")))
        #             legacy_clear(By.ID, "ctl00_Body_LastName")
        #             legacy_safe_type(By.ID, "ctl00_Body_LastName", xlProvider[:3])
        #         time.sleep(.5)
        #         legacy_safe_type(By.ID, "ctl00_Body_Phone_PNumber",xlProviderPhone)

        #         elemClick = driver.find_element(By.ID, "ctl00_Body_SearchButton")
        #         try: driver.execute_script("arguments[0].scrollIntoView(true);", elemClick)
        #         except WebDriverException: pass
        #         element_click(elemClick)
        #         time.sleep(3)
        #         notify("Notice","For Manual Process on Selecting Provider. Click ok Once Provider is Selected")



                
                #OLD COde
                # legacy_clear(By.ID, "ctl00_Body_FacilityLastName")
                # # driver.find_element(By.ID, "ctl00_Body_FacilityLastName").send_keys(xlProvider)
                # legacy_safe_type(By.ID, "ctl00_Body_FacilityLastName", xlProvider)
                # elemClick = driver.find_element(By.ID, "ctl00_Body_SaveButton")#.click()
                # scroll = driver.find_element(By.ID, "ctl00_Body_LastName")
                # try: driver.execute_script("arguments[0].scrollIntoView(true);", scroll)
                # except WebDriverException: pass
                # element_click(elemClick)
        # ================== ATTORNEY TAB ==================
        # notify("Next", "Proceed to Attorney Tab")
        elemClick = WebDriverWait(driver, 60).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[3]/div[1]/table/tbody/tr[2]/td[2]/table[9]/tbody/tr/td[2]/a")))#.click()
        element_click(elemClick)
        attorneyTrue = elementExist(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[5]/td/table/tbody/tr/td")
        # attorneyTrue = driver.find_element(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[5]/td/table/tbody/tr/td").text.strip()
        xlAttorneyFullAdd = f"{xlAttorneyAdd1} {xlAttorneyAdd2}, {xlAttorneyState}, {xlAttorneyZip}"
        if attorneyTrue and xlAttorney:
            elemClick = driver.find_element(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[4]/td/a")#.click() 
            element_click(elemClick)
            xlAttorneyFname, xlAttorneyLname  = SplitName(xlAttorney)
            # driver.find_element(By.ID, "ctl00_Body_LastName").send_keys(xlAttorneyLname)
            legacy_safe_type(By.ID, "ctl00_Body_LastName", xlAttorneyLname)
            # driver.find_element(By.ID, "ctl00_Body_FirstName").send_keys(xlAttorneyFname)
            legacy_safe_type(By.ID, "ctl00_Body_FirstName", xlAttorneyFname)   
            elemClick =driver.find_element(By.ID, "ctl00_Body_SearchButton")#.click()
            element_click(elemClick)
            time.sleep(3)

            if not elementExist(By.XPATH, "/html/body/form/div[3]/div/div[2]/span/table/tbody/tr[1]/td[2]/a"):
                elemClick = driver.find_element(By.ID, "ctl00_Body_AddLink")#.click()
                element_click(elemClick)
                time.sleep(3)
                # driver.find_element(By.ID, "ctl00_Body__LastName").send_keys(xlAttorneyLname)
                legacy_safe_type(By.ID, "ctl00_Body__LastName", xlAttorneyLname) 
                # driver.find_element(By.ID, "ctl00_Body__FirstName").send_keys(xlAttorneyFname) 
                legacy_safe_type(By.ID, "ctl00_Body__FirstName", xlAttorneyFname)                                    
                # driver.find_element(By.ID, "ctl00_Body__Address1").send_keys(xlAttorneyAdd1)   
                legacy_safe_type(By.ID, "ctl00_Body__Address1", xlAttorneyAdd1)                                  
                # driver.find_element(By.ID, "ctl00_Body__Address2").send_keys(xlAttorneyAdd2)   
                legacy_safe_type(By.ID, "ctl00_Body__Address2", xlAttorneyAdd2)                                  
                # driver.find_element(By.ID, "ctl00_Body__City").send_keys(xlAttorneyCity)
                legacy_safe_type(By.ID, "ctl00_Body__City", xlAttorneyCity) 
                
                ie_select_by_value(By.ID, "ctl00_Body__State_SingleSelectList", xlAttorneyState)                                   
                # driver.find_element(By.ID, "ctl00_Body__ZipCode").send_keys(xlAttorneyZip)
                legacy_safe_type(By.ID, "ctl00_Body__ZipCode", xlAttorneyZip) 
                # driver.find_element(By.ID, "ctl00_Body__Phone_PNumber").send_keys(xlAttorneyPhone) 
                legacy_safe_type(By.ID, "ctl00_Body__Phone_PNumber", xlAttorneyPhone) 
                # driver.find_element(By.ID, "ctl00_Body__Phone_PExtension").send_keys(xlAttorneyExt)  
                # legacy_safe_type(By.ID, "ctl00_Body__Phone_PExtension", xlAttorneyExt) 
                elemClick = driver.find_element(By.ID, "ctl00_Body__AttorneyType_List_2")#.click()
                element_click(elemClick)
                elemClick = driver.find_element(By.XPATH, "/html/body/form/div[3]/div[4]/input[1]")#.click()
                element_click(elemClick)
            else:
                attorneysTable = driver.find_element(By.ID, "ctl00_Body_SearchResultsTable_Table")
                attorneyTbody = attorneysTable.find_elements(By.XPATH, ".//tbody/tr")
                for rowAtt in attorneyTbody:
                    cells = rowAtt.find_elements(By.TAG_NAME, "td")
                    if not cells:
                        continue

                    cmsAttorneyName = cells[1].text.strip()
                    xlAttorneyFullName = f"{xlAttorneyLname}, {xlAttorneyFname}"

                    if xlAttorneyFullName in cmsAttorneyName:
                        links = rowAtt.find_elements(By.TAG_NAME, "a")
                        element_click(links[-1])
                        # links[-1].click()
        else:
            try:
                attorneyTbl = driver.find_element(By.ID, "ctl00_Body_AttorneyListTable_Table")
                attorneyTr = attorneyTbl.find_elements(By.XPATH, ".//tbody/tr")
                attorneyName = driver.find_element(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/a").text.strip()
                attorneyAddress = driver.find_element(By.XPATH, "/html/body/form/div[3]/table/tbody/tr[3]/td/table/tbody/tr/td[5]").text.strip()
                similarityAttorney = difflib.SequenceMatcher(None, xlAttorneyFullAdd.lower(), attorneyAddress.lower()).ratio()

                if similarityAttorney  >= 0.7:
                    print("Same Attorney's Address")
                else:
                    print("Different Attorney's Address")
            except: pass
        # ================== BILLING TAB ==================
        # notify("Next", "Proceed to Billing Tab")
        elemClick = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[3]/div[1]/table/tbody/tr[1]/td[2]/table[9]/tbody/tr/td[2]/a"))
        )#.click()
        element_click(elemClick)
        time.sleep(5)
        activityCodeText = driver.find_element(By.XPATH,"/html/body/form/div[3]/div[6]/div[2]/div/table/tbody/tr[2]/td[2]/input[3]").get_attribute("value")
        

        if not activityCodeText:
            activity_tab = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.XPATH,
                    "/html/body/form/div[3]/div[1]/table/tbody/tr[2]/td[2]/table[1]/tbody/tr/td[2]/a")
            )
        )
            element_click(activity_tab)
            if  caseReferralType == "Medical - Task Assignment":
                time.sleep(3)
                # driver.find_element(By.ID,  "ctl00_Body__Activity__Code").send_keys("OPN71")
                legacy_safe_type(By.ID, "ctl00_Body__Activity__Code", "OPN71") 
                # driver.find_element(By.ID,  "ctl00_Body__Activity__Note").send_keys("Case Opening - Task")
                legacy_safe_type(By.ID, "ctl00_Body__Activity__Note", "Case Opening - Task") 
            else:
                # driver.find_element(By.ID,  "ctl00_Body__Activity__Code").send_keys("OPN01")
                legacy_safe_type(By.ID, "ctl00_Body__Activity__Code", "OPN01") 
                # driver.find_element(By.ID,  "ctl00_Body__Activity__Note").send_keys("Case Opening - Field")
                legacy_safe_type(By.ID, "ctl00_Body__Activity__Note", "Case Opening - Field") 
            
            # driver.find_element(By.ID,  "ctl00_Body__Activity__Description").send_keys("Case Opening")
            legacy_safe_type(By.ID, "ctl00_Body__Activity__Description", "Case Opening") 
            elemClick = driver.find_element(By.ID,  "ctl00_Body__Activity__NoteCompleted")#.click()
            element_click(elemClick)
            #save
            #driver.find_element(By.ID,  "ctl00_Body__Activity_Save").click()

        


        # driver.switch_to.default_content()
        # WebDriverWait(driver, 20).until(
        # EC.frame_to_be_available_and_switch_to_it((By.XPATH, "/html/body/form/div[5]/iframe"))
        # ) 
        # time.sleep(3) 
        # elemClick = driver.find_element(By.XPATH, "/html/body/form/div/div[2]/div[10]/div[1]/span")#.click()
        # element_click(elemClick)
        


        # ================== ATTACHMENTS TAB (REFERRAL SOURCE) ==================
        # notify("Next", "Proceed to Attachments Tab")
        attach_tab = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable(
                (By.XPATH,
                    "/html/body/form/div[3]/div[1]/table/tbody/tr[2]/td[2]/table[2]/tbody/tr/td[2]/a")
                    # "/html/ body/form/div[1]/table/tbody/tr[2]/td[2]/table[2]/tbody/tr/td[2]/a")
            )
        )
        element_click(attach_tab)
        time.sleep(3)

        if elementExist(By.XPATH, "/html/body/form/div[3]/div[2]/div[2]/table/tbody/tr/td"):
            attachmentsTable = driver.find_element(
                By.ID, "ctl00_Body__AttachmentTable_Section_Table_Table"
            )
            attachmentsTbody = attachmentsTable.find_elements(By.XPATH, ".//tbody/tr")
            for rowAttach in attachmentsTbody:
                cells = rowAttach.find_elements(By.TAG_NAME, "td")
                if not cells:
                    continue
                attachFrom = cells[1].text.strip()
                if "UnityApiUser" in attachFrom:
                    selectAttach = cells[6]
                    elemClick = selectAttach.find_element(By.TAG_NAME, "a")
                    element_click(elemClick)
                    time.sleep(5)
                    AttachSubject = driver.find_element(
                        By.ID, "ctl00_Body__Subject"
                    ).get_attribute("value")
                    AttachSubject = f"{AttachSubject} - {xlReferralSource}"
                    legacy_safe_type(By.ID, "ctl00_Body__Subject", AttachSubject)
                    save_btn = driver.find_element(By.ID, "ctl00_Body__SaveButton")
                    element_click(save_btn)
                    break

        time.sleep(3)

        # #LettersTab
        # notify("Next", "Proceed to Letters Tab")
        elemClick = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[3]/div[1]/table/tbody/tr[2]/td[2]/table[7]/tbody/tr/td[2]/a"))
        )#.click()  
        element_click(elemClick) 
        time.sleep(3)  
        wait = WebDriverWait(driver, 5)
        def switch_to_frame_LettersTab(driver, timeout=5) -> bool:
            driver.switch_to.default_content()
            frames = driver.find_elements(By.TAG_NAME, "iframe")
            print("Found iframes:", len(frames))
        
            for idx, frame in enumerate(frames):
                try:
                    driver.switch_to.default_content()
                    driver.switch_to.frame(frame)
        
                    # is there any element whose text contains CPD?
                    elems = driver.find_elements(By.XPATH, "//*[contains(normalize-space(.), 'CPD')]")
                    if elems:
                        print("Found CPD in iframe index:", idx)
                        return True
                except Exception as e:
                    print(f"Error checking frame {idx}: {e}")
                    continue
        
            driver.switch_to.default_content()
            return False

        if switch_to_frame_LettersTab(driver):
            # Now we are inside the iframe that actually has the CPD link
            cpd_span = wait.until(EC.presence_of_element_located((
                By.XPATH,
                "//span[normalize-space()='CPD']"   # <a> that has a <span>CPD</span> inside
            )))

            cpd_link = driver.find_element(By.XPATH,"//span[normalize-space()='CPD']/preceding::a[contains(@class,'exp')][1]")
            element_click(cpd_link)
            time.sleep(0.5)
            cpderfs_link = driver.find_element(By.XPATH,
                                                "//span[normalize-space()='CPD']"
                                                "/ancestor::div[@id][1]"
                                                "/following-sibling::div[1]"
                                                "//a[normalize-space(.)='CPD ERFS']")
            element_click(cpderfs_link)
            # elemClick = WebDriverWait(driver, 60).until(
            # EC.element_to_be_clickable((By.XPATH, "/html/body/form/div/div[2]/div[10]/div[2]/div[2]/a"))
            # )#.click()
            # element_click(elemClick)  
            # driver.find_element(By.ID, "instructions").send_keys(xxlSpecialInstructions)
            legacy_safe_type(By.ID, "instructions", Special_Instructions)
            elemClick = driver.find_element(By.XPATH, "//input[@type='submit' and @value='Configure']")#.click() 
            element_click(elemClick) 

            # time.sleep(2)
            # elemClick = driver.find_element(By.XPATH, "/html/body/form/table/tbody/tr[2]/td[6]/a")#.click() 
            # element_click(elemClick) 
            # time.sleep(2)
            # elemClick = driver.find_element(By.XPATH, "/html/body/form/table[1]/tbody/tr/td/div/table/tbody/tr[2]/td[2]/table[8]/tbody/tr/td[2]/a")
            
            # # elemClick = driver.find_element(By.XPATH, "/html/body/form/table[3]/tbody/tr[2]/td/input")#.click() 
            # element_click(elemClick)
            # driver.find_element(By.XPATH, "/html/body/form/center/table/tbody/tr/td/input[1]").click() 
        else:
            raise Exception("Could not find CPD in any iframe")
        
        print("Finished main CMS updates for case.")

        main = Desktop(backend="uia").window(title_re=r"Referral Routing System*",control_type="Window")
        main.set_focus()

        #Add checker here

        grid = main.child_window(auto_id="dgWorkWindow",control_type="Table")#.wrapper_object()
        # cell_el = grid.iface_grid.GetItem(0,4)
        # cell = UIAWrapper(UIAElementInfo(cell_el))
        cell_el = grid.child_window(title="Claimant Row 0",control_type="Edit").wrapper_object()
        ClaimantRRS = cell_el.iface_value.CurrentValue
        # ClaimantRRS = cell.iface_value.CurrentValue
        FNameRRS = re.split(r"\s+",ClaimantRRS)[0]
        print(FNameRRS)
        isCEM = False
        isCEM = claim_exists(claimant_name, claim_number)
        if ClaimantRRS.strip().upper() == claimant_name.strip().upper():
            cell = grid.child_window(title=f"Select Row 0", control_type="CheckBox").wrapper_object()
            cell.click_input()
            time.sleep(0.5)
            SubjectLineBuilder = main.child_window(auto_id="btnReg",control_type="Button").wrapper_object()
            SubjectLineBuilder.set_focus()
            SubjectLineBuilder.click_input()

            time.sleep(0.5)
            RegPrompter = main.child_window(auto_id="frmSmallInputBox",control_type="Window")
            RegPrompter.wait('exists',timeout=5)
            
            if isCEM:
                tbInput = RegPrompter.child_window(auto_id="tbInput", control_type="Edit").wrapper_object()
                tbInput.iface_value.SetValue("10")
                time.sleep(0.5)
                btnOk = RegPrompter.child_window(auto_id="btnOk",control_type="Button").wrapper_object()
                btnOk.click_input()
            else:
                tbInput = RegPrompter.child_window(auto_id="tbInput", control_type="Edit").wrapper_object()
                tbInput.iface_value.SetValue("3")
                time.sleep(0.5)
                btnOk = RegPrompter.child_window(auto_id="btnOk",control_type="Button").wrapper_object()
                btnOk.click_input()

                pid = main.element_info.process_id
                try:
                #Zip Code?
                    print(caseID)
                    app = Application(backend="win32").connect(process=pid)
                    prompt = app.window(title_re=r"Case Id")
                    prompt.wait("ready", 10)
                    prompt.Edit.set_edit_text(caseID)
                    time.sleep(2)
                    # prompt.Edit.set_edit_text("90017")
                    prompt.OK.click()
                except Exception:
                    pass
        # asdasd


        notify("Notice","Finished CMS Update for case")
        driver.get(CMS_CASESEARCH_URL)
        time.sleep(5)
        # break
 
    except Exception as ex:
        exc_type, exc_obj, tb = sys.exc_info()
        line_number = tb.tb_lineno if tb else "unknown"
        print(f"error occurred: {ex} (line {line_number})")
 
    time.sleep(3)
 
 
# ===================== PUBLIC ENTRY POINT =====================
EXCEL_FILE_NAME = "FCM CEM Checher.xlsx"

def get_script_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))
 
 
def get_excel_path():
    return os.path.join(get_script_dir(), EXCEL_FILE_NAME)
 
 
def get_today_sheet_name():
    return datetime.now().strftime("%m%d%Y")
 
 
def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip().lower()
 
 
def claim_exists(ClaimantName, ClaimID):
    excel_path = get_excel_path()
 
    if not os.path.exists(excel_path):
        return False
 
    sheet_name = get_today_sheet_name()
 
    wb = load_workbook(excel_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return False
 
        ws = wb[sheet_name]
 
        target_claimant = normalize_text(ClaimantName)
        target_claim_id = normalize_text(ClaimID)

        for row in range(2, ws.max_row + 1):
            claimant_val = normalize_text(ws.cell(row=row, column=1).value)
            claim_id_val = normalize_text(ws.cell(row=row, column=2).value)
 
            if claimant_val == target_claimant and claim_id_val == target_claim_id:
                return True
 
        return False
 
    finally:
        wb.close()

def CheckCaseUnity(claim_number):
    UNITY_URL = "https://uat-cm-portal-eus-wa.azurewebsites.net/Ng/search/caseSearch"
    def loginUnity(driver):
        driver.maximize_window()
        driver.get(UNITY_URL)
        WebDriverWait(driver, 15).until(EC.title_contains("Unity"))
    
        unity_handle = driver.current_window_handle
    
        # Find the account chooser window (sometimes same handle)
        WebDriverWait(driver, 15).until(lambda d: len(d.window_handles) >= 1)
        target_handle = None
        for h in driver.window_handles:
            driver.switch_to.window(h)
            try:
                WebDriverWait(driver, 5).until(EC.title_contains("Choose your account"))
                target_handle = h
                break
            except Exception:
                continue
    
        if not target_handle:
            # Sometimes the AAD picker is embedded; try clicking by id directly
            try:
                driver.switch_to.window(unity_handle)
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "AzureADCommonExchange")))
                btn = driver.find_element(By.ID, "AzureADCommonExchange")
                driver.execute_script("arguments[0].click();", btn)
            except Exception:
                raise RuntimeError('No window with title containing "Choose your account" found.')
    
        else:
            # In chooser window
            WebDriverWait(driver, 10).until(lambda d: d.execute_script("return document.readyState") == "complete")
            btn = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID, "AzureADCommonExchange")))
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                btn.click()
            except Exception:
                driver.execute_script("arguments[0].click();", btn)
    
        # Return to Unity and ensure dashboard is loaded
        driver.switch_to.window(unity_handle)
        driver.get(UNITY_URL)
        WebDriverWait(driver, 15).until(EC.title_contains("Unity"))
        
    def create_driver():
        EDGE_DRIVER_PATH = CONFIG_EDGE_DRIVER_PATH
        opts = EdgeOptions()
    
        # your existing args
        opts.add_argument("--ignore-certificate-errors")
        opts.add_argument("--allow-insecure-localhost")
        opts.add_argument("--disable-popup-blocking")
    
        # ✅ stability / crash prevention
        opts.add_argument("--disable-gpu")
        opts.add_argument("--disable-software-rasterizer")
        opts.add_argument("--no-sandbox")  # harmless on Windows; helps in some environments
        opts.add_argument("--disable-dev-shm-usage")  # mostly Linux, harmless on Windows
    
        # ✅ reduce Chromium noise in console
        opts.add_argument("--log-level=3")
        opts.add_experimental_option("excludeSwitches", ["enable-logging"])
    
        # ✅ optional: stop Edge “on-device model / LLM” feature from initializing
        opts.add_argument("--disable-features=OnDeviceModel,OptimizationHints,OptimizationGuideModelDownloading")
    
        # ✅ write webdriver logs so we can see if it crashed vs hung
        service = EdgeService(executable_path=EDGE_DRIVER_PATH, log_output="edgedriver.log")
        driver = webdriver.Edge(service=service, options=opts)

        # ✅ prevent “hangs forever”
        driver.set_page_load_timeout(60)
        driver.set_script_timeout(60)
        return driver  
    
    def NavigateSearch(ClaimNumber,driver):
        # SearchBtn = driver.find_element(By.XPATH,"/html/body/my-app/div/header/nav/div/ul/li[3]/a")
        # SearchBtn.click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "tasks-claim-number-general-input")))


        ClaimNumberField = driver.find_element(By.ID,"tasks-claim-number-general-input")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", ClaimNumberField)
        ClaimNumberField.send_keys(ClaimNumber)

        StatusDropdown = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "tasks-case-status-select-input")))
        selection = Select(StatusDropdown)
        selection.select_by_value("1")

        SearchCaseBtn = driver.find_element(By.XPATH,"/html/body/my-app/div/div/div/ng-component/div/div/case-search/div/div/form/div[2]/div/button")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", SearchCaseBtn)
        SearchCaseBtn.click()
        time.sleep(5)
        # mainTab = driver.current_window_handle
        # resultRows = driver.find_elements(By.XPATH,"/html/body/my-app/div/div/div/ng-component/div/div/case-search/div/div/div[2]")

        results_container_xpath = "/html/body/my-app/div/div/div/ng-component/div/div/case-search/div/div/div[2]/div[2]"
        all_rows_xpath = results_container_xpath + "/div"
        
        assigned_date_id = "case-tab-case-info-assigned-date-date-input"
        case_number_xpath = "/html/body/my-app/div/div/div/fcm-app/ng-component/div[3]/div/ng-component/div/div/loading-section/div[1]/form-template/div/div/case-info/save-cancel-form/section/div/div[2]/form/div[1]/div[3]/span"
        
        main_tab = driver.current_window_handle
        
        latest_case_id = None
        latest_assigned_date = None
        
        WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located((By.XPATH, all_rows_xpath))
        )
        
        row_count = len(driver.find_elements(By.XPATH, all_rows_xpath))
        print(f"Total rows found: {row_count}")
        
        for i in range(1, row_count + 1):
            try:
                driver.switch_to.window(main_tab)
                if i == 1:
                    current_row_xpath = f"{results_container_xpath}/div"
                else:
                    current_row_xpath = f"{results_container_xpath}/div[{i}]"
                row = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH, current_row_xpath))
                )
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", row)
                # row.click()
                # time.sleep(5)
                old_tabs = driver.window_handles[:]
                row.click()
                time.sleep(5)
                # driver.execute_script("arguments[0].click();", row)
                # WebDriverWait(driver, 15).until(
                #     lambda d: len(d.window_handles) > len(old_tabs)
                # )
                WebDriverWait(driver, 15).until(
                    EC.new_window_is_opened(old_tabs)
                )
                
                new_tabs = [tab for tab in driver.window_handles if tab not in old_tabs][0]
                driver.switch_to.window(new_tabs)
        
                assigned_date_el = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.ID, assigned_date_id))
                )
                assigned_date_str = assigned_date_el.get_attribute("value").strip()
                case_number_el = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.XPATH, case_number_xpath))
                )
                case_id = case_number_el.text.strip()
                print(f"Row {i}: CaseID={case_id}, AssignedDate={assigned_date_str}")
        
                assigned_date_obj = None
                if assigned_date_str:
                    try:
                        assigned_date_obj = datetime.strptime(assigned_date_str, "%m/%d/%Y")
                    except ValueError:
                        print(f"Invalid date format in row {i}: {assigned_date_str}")
        
                if assigned_date_obj:
                    if latest_assigned_date is None or assigned_date_obj > latest_assigned_date:
                        latest_assigned_date = assigned_date_obj
                        latest_case_id = case_id
        
                driver.close()
                driver.switch_to.window(main_tab)
        
            except Exception as e:
                print(f"Error processing row {i}: {e}")
        
                try:
                    if len(driver.window_handles) > 1 and driver.current_window_handle != main_tab:
                        driver.close()
                except:
                    pass
        
                try:
                    driver.switch_to.window(main_tab)
                except:
                    pass
        
        if latest_case_id and latest_assigned_date:
            print(f"Latest CaseID: {latest_case_id}")
            print(f"Latest Assigned Date: {latest_assigned_date.strftime('%m/%d/%Y')}")
        else:
            print("No valid case found.")

        return latest_case_id

    driver = create_driver()
    loginUnity(driver)
    CaseID = NavigateSearch(claim_number,driver)
    driver.quit()
    return CaseID


def run_case(
    connection_string: str,
    claimant_name: str,
    claim_number: str,
    cms_user: str,
    cms_pass: str,
    interactive: bool = False
):
    global driver, INTERACTIVE
    INTERACTIVE = interactive
 
    try:
        ensure_readable_path(LIBERTY_OFFICE_MATRIX, "Liberty Office Matrix Excel")

        caseID = CheckCaseUnity(claim_number)
 
        print("Starting IE driver...")
        driver = create_ie_driver()
 
        print("Logging into CMS...")
        login(cms_user, cms_pass)
 
        print(f"Processing case for {claimant_name} / {claim_number}")
        searchCase(connection_string, claimant_name, claim_number,caseID)
 
    except (PermissionError, FileNotFoundError) as e:
        print(f"CRITICAL FILE/ACCESS PROBLEM: {e}")
    except Exception as ex:
        print("Unhandled error:", ex)
    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass
 
 
# Optional backward-compatible name
def main(connection_string, claimant_name, claim_number, cms_user, cms_pass, interactive=False):
    run_case(connection_string, claimant_name, claim_number, cms_user, cms_pass, interactive)
 
 
if __name__ == "__main__":
    # Example standalone test usage.
    # In real usage, your other Python script will import run_case(...) instead.
    conn_str = (
        "DRIVER={SQL Server};"
        "SERVER=tcp:uat-rrs-core-eus-dbs.database.windows.net;"
        "DATABASE=RRS;"
        "UID=RRSApplication;"
        "PWD=;"
    )
 
    run_case(
        connection_string=conn_str,
        claimant_name="Kirsten Lark",
        claim_number="",
        cms_user="",
        cms_pass="",
        interactive=True  # set False when used by another bot
    )
    # CheckCaseUnity()




