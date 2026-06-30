from pywinauto import Desktop, uia_defines, Application
from fcm_intake.config import CMS_LOGIN_URL as CONFIG_CMS_LOGIN_URL, EDGE_PATH as CONFIG_EDGE_PATH, IE_DRIVER_PATH as CONFIG_IE_DRIVER_PATH
from pywinauto.uia_element_info import UIAElementInfo
from pywinauto.controls.uiawrapper import UIAWrapper
from pywinauto.keyboard import send_keys
from pywinauto.mouse import click as mouse_click
#migz
from pywinauto.findwindows import find_windows
from pywinauto.timings import wait_until_passes
from pywinauto.timings import wait_until
import ctypes

import re
import time
from datetime import datetime 
# import tkinter as tk
# from tkinter import messagebox
from pywinauto.application import Application
import win32clipboard as wcb
from pprint import pprint
from titlecase import titlecase
import logging
from dataclasses import dataclass
import pyodbc
import sys
# import ReOpenCheck
from legacy.legacy_providerresultchecker import *
import ReOpenCheck_shared as reopen_mod
import CustomerCheckerV2_shared as customer_mod
# from legacy.legacy_reopencheck import *
from legacy.legacy_googlesearch import *
# from legacy.legacy_customerchecker import *
import openpyxl
from pathlib import Path
import traceback
import difflib
import win32com.client as win32


import win32api

def get_app_folder() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
 
    return os.path.dirname(os.path.abspath(__file__))
 
 
def save_employer_contact_to_excel(parsed_data: dict) -> None:
    excel = None
    wb = None
 
    try:
        folder_path = get_app_folder()
        file_path = os.path.join(folder_path, "FCM Intake Employer Contact Details.xlsb")
        sheet_name = "Contact Details"
 
        today_date = datetime.today().strftime("%m/%d/%Y")
 
        claimant_name = (
            parsed_data.get("claimantFull", "")
            or parsed_data.get("Claimant Name", "")
            or parsed_data.get("claimant_name", "")
        )
 
        claim_number = (
            parsed_data.get("claimNumber", "")
            or parsed_data.get("Claim Number", "")
            or parsed_data.get("claim_number", "")
        )
 
        employer_name = parsed_data.get("employerContactName", "")
        employer_phone = parsed_data.get("employerContactPhone", "")
        employer_email = parsed_data.get("employerContactEmail", "")
 
        if not employer_name:
            employer_name = "Unknown"
 
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
 
        if os.path.exists(file_path):
            wb = excel.Workbooks.Open(file_path)
        else:
            wb = excel.Workbooks.Add()
 
            while wb.Worksheets.Count > 1:
                wb.Worksheets(wb.Worksheets.Count).Delete()
 
            ws = wb.Worksheets(1)
            ws.Name = sheet_name
 
            wb.SaveAs(file_path, FileFormat=50)
 
        try:
            ws = wb.Worksheets(sheet_name)
        except Exception:
            ws = wb.Worksheets.Add()
            ws.Name = sheet_name
 
        headers = [
            "Date",
            "Claimant Name",
            "Claim Number",
            "Employer Contact Name",
            "Employer Phone",
            "Employer Email",
        ]
 
        header_is_missing = False
 
        for col_index, header in enumerate(headers, start=1):
            current_value = ws.Cells(1, col_index).Value
 
            if current_value != header:
                header_is_missing = True
                break
 
        if header_is_missing:
            for col_index, header in enumerate(headers, start=1):
                ws.Cells(1, col_index).Value = header
 
            ws.Range("A1:F1").Font.Bold = True
 
        last_row = ws.Cells(ws.Rows.Count, 1).End(-4162).Row
 
        if last_row == 1 and not ws.Cells(2, 1).Value:
            next_row = 2
        else:
            next_row = last_row + 1
 
        ws.Cells(next_row, 1).Value = today_date
        ws.Cells(next_row, 2).Value = claimant_name
        ws.Cells(next_row, 3).Value = claim_number
        ws.Cells(next_row, 4).Value = employer_name
        ws.Cells(next_row, 5).Value = employer_phone
        ws.Cells(next_row, 6).Value = employer_email
 
        ws.Columns("A:A").NumberFormat = "mm/dd/yyyy"
        ws.Columns("A:F").AutoFit()
 
        wb.Save()
 
    except Exception as e:
        print(f"Failed to save employer contact details to Excel: {e}")
 
    finally:
        try:
            if wb:
                wb.Close(SaveChanges=True)
        except Exception:
            pass
 
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass



def focus_control(ctrl, timeout=10):
    try:
        ctrl.wait("visible", timeout=timeout)
    except Exception:
        pass
    try:
        ctrl.restore()
    except Exception:
        pass
    try:
        control_type = str(getattr(getattr(ctrl, "element_info", None), "control_type", "") or "")
        if control_type.lower() == "window":
            rect = ctrl.rectangle()
            width = rect.width()
            height = rect.height()
            screen_width = win32api.GetSystemMetrics(0)
            screen_height = win32api.GetSystemMetrics(1)
            new_x = int((screen_width - width) / 2)
            new_y = int((screen_height - height) / 2)
            ctrl.move_window(new_x, new_y, width, height, repaint=True)
    except Exception:
        pass
    try:
        ctrl.set_focus()
    except Exception:
        pass

def safe_click(ctrl, timeout=10):
    try:
        parent = ctrl.top_level_parent()
    except Exception:
        parent = None

    if parent is not None:
        focus_control(parent, timeout=timeout)
    else:
        focus_control(ctrl, timeout=timeout)

    try:
        ctrl.click_input()
        return
    except Exception:
        pass
    try:
        ctrl.click()
        return
    except Exception:
        pass
    try:
        ctrl.invoke()
    except Exception:
        pass

# root = tk.TK()
# root.withdraw()

main = Desktop(backend="uia").window(title_re=r"Referral Routing System*",control_type="Window")
focus_control(main)

connection_string = ("DRIVER={SQL Server};"
            "SERVER=tcp:uat-rrs-core-eus-dbs.database.windows.net;"
            "DATABASE=RRS;"
            "UID=RRSApplication;"
            "PWD=;"
            # add name="SQLProd" connectionString="Server=;Initial Catalog=RRS;User ID=;Password=;MultipleActiveResultSets=False;Encrypt=True;TrustServerCertificate=False;Connection Timeout=500;
            )

correctZipCode = ""
data ={}
# allData={}
isManualEntered = False
botStop = False
isManualEnteredAtty = False
CTMedOnly = False

BASE_DIR = Path(__file__).resolve().parent
LM_MATRIX_PATH = BASE_DIR / "Liberty Mutual Claims Office Matrix_08-29-19.xlsx"


# === your global used in parse_referral_pdf ===
correctZipCode = ""
def formataddrSearchStr (s: str) -> str:
    # your existing formatter – placeholder
    return " ".join(str(s).split())

def parse_address(full_address: str):
    text = full_address.strip()
    text = re.sub(r"\s+", " ", text)           # normalize spaces
    text = re.sub(r",\s*,+", ",", text)        # collapse double commas: ",," -> ","
    text = text.replace(" ,", ",")             # normalize " ,"

    # ---------- 1) ZIP ----------
    zip_match = re.search(r"(\d{5})(?:-\d{4})?\s*,?\s*$", text)
    zipcode = zip_match.group(1) if zip_match else ""
    if zipcode:
        text = text[:zip_match.start()].rstrip(" ,")

    # ---------- 2) STATE ----------
    state = ""
    state_codes = {
        "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA","KS","KY",
        "LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ","NM","NY","NC","ND",
        "OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT","VA","WA","WV","WI","WY","DC"
    }
    full_to_abbr = {
        "ALABAMA":"AL","ALASKA":"AK","ARIZONA":"AZ","ARKANSAS":"AR","CALIFORNIA":"CA",
        "COLORADO":"CO","CONNECTICUT":"CT","DELAWARE":"DE","FLORIDA":"FL","GEORGIA":"GA",
        "HAWAII":"HI","IDAHO":"ID","ILLINOIS":"IL","INDIANA":"IN","IOWA":"IA","KANSAS":"KS",
        "KENTUCKY":"KY","LOUISIANA":"LA","MAINE":"ME","MARYLAND":"MD","MASSACHUSETTS":"MA",
        "MICHIGAN":"MI","MINNESOTA":"MN","MISSISSIPPI":"MS","MISSOURI":"MO","MONTANA":"MT",
        "NEBRASKA":"NE","NEVADA":"NV","NEW HAMPSHIRE":"NH","NEW JERSEY":"NJ",
        "NEW MEXICO":"NM","NEW YORK":"NY","NORTH CAROLINA":"NC","NORTH DAKOTA":"ND",
        "OHIO":"OH","OKLAHOMA":"OK","OREGON":"OR","PENNSYLVANIA":"PA","RHODE ISLAND":"RI",
        "SOUTH CAROLINA":"SC","SOUTH DAKOTA":"SD","TENNESSEE":"TN","TEXAS":"TX",
        "UTAH":"UT","VERMONT":"VT","VIRGINIA":"VA","WASHINGTON":"WA","WEST VIRGINIA":"WV",
        "WISCONSIN":"WI","WYOMING":"WY","DISTRICT OF COLUMBIA":"DC"
    }

    m2 = re.search(r",\s*([A-Za-z]{2,})\s*$", text)
    if m2:
        cand = m2.group(1)
        up = cand.upper()
        if up in state_codes:
            state = up
        elif up in full_to_abbr:
            state = full_to_abbr[up]
        if state:
            text = text[:m2.start()].rstrip(" ,")
    else:
        tokens = text.replace(",", " ").split()
        for tok in reversed(tokens):
            up = tok.upper()
            if up in state_codes or up in full_to_abbr:
                state = up if up in state_codes else full_to_abbr[up]
                idx = text.upper().rfind(tok.upper())
                text = (text[:idx] + text[idx+len(tok):]).rstrip(" ,")
                break

    # ---------- 3) Remaining text: addr1 / addr2 / city ----------
    addr1 = addr2 = city = ""
    addr2_keywords = {
        "suite","ste","apt","apartment","bldg","building",
        "unit","fl","floor","#","po","p.o.","po box","p.o. box","rm","room"
    }

    def split_with_commas(s):
        nonlocal addr1, addr2, city
        parts = [p.strip(" ,") for p in s.split(",") if p.strip(" ,")]
        if not parts:
            return True
        if len(parts) == 1:
            return False

        addr1 = parts[0]
        if len(parts) >= 3:
            addr2 = parts[1]
            city = " ".join(parts[2:])
        else:
            second = parts[1]
            tokens2 = second.split()
            if tokens2 and tokens2[0].lower() in addr2_keywords:
                i = 1
                # allow digits, #, and short suite letters like "B", "2B", "B1"
                while i < len(tokens2) and (
                    any(c.isdigit() for c in tokens2[i]) or
                    tokens2[i].startswith("#") or
                    re.fullmatch(r"[A-Za-z]{1,2}\d{0,2}", tokens2[i]) is not None
                ):
                    i += 1
                addr2 = " ".join(tokens2[:i])
                city = " ".join(tokens2[i:])
            else:
                city = second
        return True

    used_commas = split_with_commas(text)

    if not used_commas:
        tokens = text.replace(",", " ").split()
        if not tokens:
            return "", "", "", state, zipcode

        kw_idx = None
        for i, t in enumerate(tokens):
            if t.replace(".", "").lower() in addr2_keywords:
                kw_idx = i
                break

        if kw_idx is None:
            if len(tokens) == 1:
                addr1 = tokens[0]
            else:
                addr1 = " ".join(tokens[:-1])
                city = tokens[-1]
        else:
            j = kw_idx + 1
            while j < len(tokens) and (
                any(c.isdigit() for c in tokens[j]) or
                tokens[j].startswith("#") or
                re.fullmatch(r"[A-Za-z]{1,2}\d{0,2}", tokens[j]) is not None  # <-- captures "B"
            ):
                j += 1
            addr1 = " ".join(tokens[:kw_idx])
            addr2 = " ".join(tokens[kw_idx:j])
            city = " ".join(tokens[j:])

    return addr1, addr2, city, state, zipcode

import re
from typing import Optional, Tuple, List

AddrParts = Tuple[Optional[str], Optional[str], Optional[str], Optional[str], Optional[str]]

STOP_HEADERS_RE = re.compile(
    r"""(?im)^\s*(?: 
        Appt\.?\s*Date: |
        Appointment\s*Date: |
        Employer\b |
        Employer\s+name\b |
        Contact\s+name\b |
        CURRENT\s+WORK\s+STATUS\b |
        COMPENSABLE\s+BODY |
        WORK\s+STATUS |
        INSTRUCTIONS: |
        REFERRAL\s+INSTRUCTIONS\b |
        Referrer\s+Name: |
        Attorney\b |
        Vendor\b |
        ASSIGNMENT\s+TYPE\b |
        Goal\b |
        Date\s*: |
        Time\s*: |
        Ph\.?\#\s*: |
        Phone\b
    )""",
    re.VERBOSE,
)

ADDRESS_LABEL_RE = re.compile(
    r"""(?im)^\s*(?:
        Address\s*of\s*Appt\.?\s*location |
        Address\s*\(.*?\) |
        Location\s*Address |
        Provider\s*Address |
        Facility\s*Address |
        Organization\s*Address |
        Address\s*\(Street/City/St\.?/zip\) |
        Address
    )\s*[:\-]\s*(.*)$""",
    re.VERBOSE,
)

CITY_STATE_ZIP_RE = re.compile(
    r"""(?ix)
    ^\s*
    (?P<addr>.+?)
    \s+
    (?P<city>[A-Za-z][A-Za-z .'\-/&]+?)
    \s*,?\s*
    (?P<st>[A-Z]{2})
    \s*
    (?P<zip>\d{5})(?:-\d{4})?
    \s*$
    """,
)

CITY_ST_ZIP_ONLY_RE = re.compile(
    r"""(?ix)^\s*([A-Za-z][A-Za-z .'\-/&]+?)\s+([A-Z]{2})\s*(\d{5})(?:-\d{4})?\s*$"""
)
SUITE_LINE_RE = re.compile(r"(?i)^\s*(suite|ste|apt|apartment|unit|#|bldg|building|fl|floor|rm|room)\b")

ST_ZIP_RE = re.compile(r"(?i)\b([A-Z]{2})\s*(\d{5})(?:-\d{4})?\b")

BAD_CONTEXT_RE = re.compile(r"(?i)\b(Injured\s*Worker|Claimant|Attorney|Insurer)\b")
GOOD_CONTEXT_RE = re.compile(r"(?i)\b(PROVIDER|Provider\s*Office|Appt\.?\s*location|Location|Facility)\b")

# ----------------------------
# Appointment address block capture (handles "Address of Appt. location:" on one line)
# ----------------------------
APPT_ADDR_BLOCK_RE = re.compile(
    r"""(?is)
    Address\s*of\s*(?:Appt\.?|App\.?|Appointment)\s*(?:location)?\s*:\s*
    (?P<addr>.*?)
    (?=
        Appt\.?\s*Date\b|
        Appointment\s*Date\b|
        Appt\.?\s*Time\b|
        Employer\b|
        Attorney\b|
        INSTRUCTIONS\b|
        Referrer\b|
        $)
    """,
    re.VERBOSE,
)

def extract_appt_address_block(special: str) -> str:
    """Return the raw appointment address string after the label, trimmed to the next header."""
    if not special:
        return ""
    m = APPT_ADDR_BLOCK_RE.search(special)
    if not m:
        return ""
    block = (m.group("addr") or "").strip()
    # normalize whitespace (keep commas; parse_address will normalize commas to spaces)
    block = re.sub(r"\s+", " ", block).strip()
    return block

BLOCK_LABEL_RE = re.compile(
    r"""(?is)
    (?:^|\n)\s*(?:
        Address\s*of\s*Appt\.?\s*location |
        Address\s*\(.*?\) |
        Location\s*Address |
        Provider\s*Address |
        Facility\s*Address |
        Organization\s*Address |
        Address\s*\(Street/City/St\.?/zip\) |
        Address
    )\s*[:\-]\s*
    """,
    re.VERBOSE,
)

def _norm(s: str) -> str:
    s = (s or "").replace("\u00A0", " ").replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def parse_address_fallback(seg: str) -> AddrParts:
    seg = _norm(seg)
    if not seg:
        return (None, None, None, None, None)

    m = CITY_STATE_ZIP_RE.match(seg)
    if not m:
        return (None, None, None, None, None)

    addr1 = m.group("addr").strip(" ,")
    city = m.group("city").strip(" ,")
    st_ = m.group("st").upper()
    z_ = m.group("zip")[:5]

    addr2 = None
    m2 = re.search(r"(?i)\b(Suite|Ste|Apt|Unit|#)\s*([A-Za-z0-9\-]+)\b", addr1)
    if m2:
        addr2 = m2.group(0).strip()
        addr1 = re.sub(re.escape(m2.group(0)), "", addr1).strip(" ,")

    return (addr1 or None, addr2, city or None, st_ or None, z_ or None)

def score_address_candidate(line: str, prev_line: str = "", next_line: str = "") -> int:
    ctx = " ".join([prev_line, line, next_line])
    s = 0

    if GOOD_CONTEXT_RE.search(ctx):
        s += 5
    if re.search(r"(?i)^\s*(Address|Street\s*Address)\s*[:\-]", line):
        s += 3
    if re.search(r"(?i)^\s*Address\s*\(.*?\)\s*[:\-]", line):
        s += 3
    if BAD_CONTEXT_RE.search(ctx):
        s -= 10

    return s

def extract_provider_address_dynamic(
    special: str,
    parse_address=None,
    provider_first: str = "",
    provider_last: str = "",
    facility: str = "",
    scan_window: int = 25,
) -> Optional[AddrParts]:
    """
    Returns (addr1, addr2, city, state, zip) or None.

    Key upgrades:
    - Anchor-first scan: if provider/facility tokens are found in Special Instructions,
      return the FIRST valid address after that anchor.
    - Handles 2-line (street + "City ST ZIP") and 3-line (street + suite + "City ST ZIP") formats.
    - Trims junk after ZIP (phone, extra text).
    - Falls back to the original global candidate scoring approach.
    """
    if not special:
        return None

    raw_lines = [ln.rstrip() for ln in special.splitlines()]
    lines = [ln.replace("\u00A0", " ").strip() for ln in raw_lines if ln.strip()]

    # -------------------------
    # Helpers (local to avoid impacting the rest of your script)
    # -------------------------
    def _norm_spaces(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").strip())

    def _norm_lower(s: str) -> str:
        return _norm_spaces(s).lower()

    def _first_word(s: str) -> str:
        s = _norm_spaces(s)
        if not s:
            return ""
        return re.split(r"[,\s/]+", s)[0].strip()

    STOP_WORDS = {"the", "a", "an", "&", "and"}

    GENERIC_FACILITY = {
        "health","healthcare","care","medical","medicine","med","clinic","hospital",
        "center","centre","group","practice","physicians","physician","doctor","doctors",
        "specialist","specialists","specialty","specialties","industrial",
        "rehab","rehabilitation","therapy","therapies","therapeutic",
        "pt","ot","st",
        "urgent","urgentcare","emergency","er",
        "imaging","radiology","diagnostic","diagnostics","diagnosis","labs","laboratory","laboratories",
        "pathology","xray","mri","ct","ultrasound",
        "orthopedic","orthopedics","ortho","spine","pain","podiatry","chiropractic","chiropractor",
        "surgery","surgical","surgicenter","ambulatory","asc","outpatient",
        "services","service","solutions","associates","association","company","corp","corporation",
        "inc","llc","ltd","pllc","pc","co","enterprise","enterprises",
        "network","partners","partner","system","systems",
        "department","dept","division","unit","team",
        "north","south","east","west","n","s","e","w",
        "downtown","uptown","central","regional","metro",
        "building","tower","plaza","suite","ste","floor","fl","room","rm",
        "the","a","an","&","and",
    }

    def _first_meaningful_word(s: str) -> str:
        s = _norm_spaces(s)
        if not s:
            return ""
        parts = [p for p in re.split(r"[,\s/]+", s) if p]
        i = 0
        while i < len(parts) and parts[i].lower() in STOP_WORDS:
            i += 1
        return parts[i] if i < len(parts) else ""

    def is_facility_name(name: str) -> bool:
        name_clean = _norm_spaces(name).lower()
        words = [p.lower() for p in re.split(r"[,\s/]+", name_clean) if p]
    
        # If any word is in GENERIC_FACILITY, treat the full name as a facility
        return any(w in GENERIC_FACILITY for w in words)

    def build_anchor_tokens(pf: str, pl: str, fac: str) -> List[str]:
        toks: List[str] = []
    
        pf1 = _first_word(pf)
        pl1 = _first_word(pl)
        fac1 = _first_meaningful_word(fac)
    
        fac_parts = [p for p in re.split(r"[,\s/]+", _norm_spaces(fac)) if p]
        fac_words_lower = [p.lower() for p in fac_parts]
    
        is_facility = any(w in GENERIC_FACILITY for w in fac_words_lower)
    
        # If this looks like a facility, do not use provider first/last as anchors
        if is_facility:
            pf1 = ""
            pl1 = ""
    
        # If first facility word is generic, move to the next useful word
        if fac1 and fac1.lower() in GENERIC_FACILITY:
            i = 0
    
            while i < len(fac_parts) and (
                fac_parts[i].lower() in GENERIC_FACILITY
                or fac_parts[i].lower() in STOP_WORDS
            ):
                i += 1
    
            fac1 = fac_parts[i] if i < len(fac_parts) else ""
    
        for t in (pf1, pl1, fac1):
            t = (t or "").strip()
            if len(t) >= 3 and t.lower() not in {"dr", "md", "mrs", "mr", "of"}:
                toks.append(t.lower())
    
        seen = set()
        out = []
        for t in toks:
            if t not in seen:
                out.append(t)
                seen.add(t)
    
        return out

    def find_anchor_index_by_tokens(ls: List[str], tokens: List[str]) -> Optional[int]:
        if not ls or not tokens:
            return None

        best_i = None
        best_hits = 0

        for i, line in enumerate(ls):
            L = _norm_lower(line)
            hits = 0
            for tok in tokens:
                if re.search(rf"(?<![a-z]){re.escape(tok)}(?![a-z])", L):
                    hits += 1
            if hits > best_hits:
                best_hits = hits
                best_i = i

        if best_i is None:
            return None

        if best_hits >= 2:
            return best_i

        # single-hit case: only accept if the token is longer (reduces false positives)
        L = _norm_lower(ls[best_i])
        for tok in tokens:
            if re.search(rf"(?<![a-z]){re.escape(tok)}(?![a-z])", L):
                return best_i if len(tok) >= 6 else None

        return None

    def _trim_after_zip(seg: str) -> str:
        if not seg:
            return seg
        m = ST_ZIP_RE.search(seg)
        if not m:
            return seg
        return _norm(seg[: m.end()])

    def _try_parse(seg: str) -> Optional[AddrParts]:
        seg = _trim_after_zip(_norm(seg))
        if not seg:
            return None

        # Prefer your existing parse_address()
        if callable(parse_address):
            try:
                addr1, addr2, city_, st_, z_ = parse_address(seg)
                if addr1 or city_ or st_ or z_:
                    return (addr1, addr2, city_, st_, z_)
            except Exception:
                pass

        a1, a2, c_, st_, z_ = parse_address_fallback(seg)
        if a1 or c_ or st_ or z_:
            return (a1, a2, c_, st_, z_)
        return None

    # -------------------------
    # 1) Anchor-first scan (reduces false positives in free-form SI)
    # -------------------------
    anchor_tokens = build_anchor_tokens(provider_first, provider_last, facility)
    anchor_i = find_anchor_index_by_tokens(lines, anchor_tokens)

    if anchor_i is not None:
        start_i = min(len(lines), anchor_i + 1)
        end_i = min(len(lines), start_i + max(5, scan_window))

        # B2) three-line: street + suite/unit + "City ST ZIP"
        for i in range(start_i, max(start_i, end_i - 2)):
            a, b, c = lines[i], lines[i + 1], lines[i + 2]
            is_street = re.match(r"^\s*\d{1,6}\s+\S+", a or "") is not None
            is_suite = SUITE_LINE_RE.match(b or "") is not None
            is_citystzip = CITY_ST_ZIP_ONLY_RE.match(c or "") is not None
            if is_street and is_suite and is_citystzip:
                parsed = _try_parse(f"{a} {b} {c}")
                if parsed and parsed[0] and parsed[2] and parsed[3] and parsed[4]:
                    return parsed

        # B) two-line: street + "City ST ZIP"
        for i in range(start_i, max(start_i, end_i - 1)):
            a, b = lines[i], lines[i + 1]
            is_street = re.match(r"^\s*\d{1,6}\s+\S+", a or "") is not None
            is_citystzip = CITY_ST_ZIP_ONLY_RE.match(b or "") is not None
            if is_street and is_citystzip:
                parsed = _try_parse(f"{a} {b}")
                if parsed and parsed[0] and parsed[2] and parsed[3] and parsed[4]:
                    return parsed

        # Single-line with zip (rare but possible)
        for i in range(start_i, end_i):
            ln = lines[i]
            if re.match(r"^\s*\d{1,6}\s+\S+", ln) and ST_ZIP_RE.search(ln):
                if re.search(r"(?i)\b(ph|phone)\b", ln):
                    continue
                parsed = _try_parse(ln)
                if parsed and parsed[0] and parsed[2] and parsed[3] and parsed[4]:
                    return parsed

    # -------------------------
    # 2) Global candidate scan (your original approach + small tweaks)
    # -------------------------
    candidates: List[Tuple[int, str]] = []

    # A) labeled single-line: ". Address .: <tail>"
    for i, ln in enumerate(lines):
        m = ADDRESS_LABEL_RE.match(ln)
        if m:
            tail = _norm(m.group(1))
            if tail:
                prev_ = lines[i - 1] if i > 0 else ""
                next_ = lines[i + 1] if i + 1 < len(lines) else ""
                sc = score_address_candidate(ln, prev_, next_)
                if sc > -5:
                    candidates.append((sc, tail))

    # B) two-line street then "City ST ZIP"
    for i in range(len(lines) - 1):
        a, b = lines[i], lines[i + 1]
        if re.match(r"^\s*\d{1,6}\s+\S+", a) and CITY_ST_ZIP_ONLY_RE.match(b):
            prev_ = lines[i - 1] if i > 0 else ""
            next_ = lines[i + 2] if i + 2 < len(lines) else ""
            # IMPORTANT: score against next_ (context) and accept unlabeled blocks
            sc = score_address_candidate(a, prev_, next_)
            if sc > -5:
                candidates.append((sc, _norm(f"{a} {b}")))

    # B2) three-line street + suite/unit + "City ST ZIP"
    for i in range(len(lines) - 2):
        a, b, c = lines[i], lines[i + 1], lines[i + 2]
        is_street = re.match(r"^\s*\d{1,6}\s+\S+", a) is not None
        is_suite = SUITE_LINE_RE.match(b or "") is not None
        is_citystzip = CITY_ST_ZIP_ONLY_RE.match(c or "") is not None
        if is_street and is_suite and is_citystzip:
            prev_ = lines[i - 1] if i > 0 else ""
            sc = score_address_candidate(a, prev_, c)
            if sc > -5:
                candidates.append((sc + 1, _norm(f"{a} {b} {c}")))

    # C) unlabeled single-line, only if context looks provider-ish
    for i, ln in enumerate(lines):
        if re.match(r"^\s*\d{1,6}\s+\S+", ln) and ST_ZIP_RE.search(ln):
            if re.search(r"(?i)\b(ph|phone)\b", ln):
                continue
            prev_ = lines[i - 1] if i > 0 else ""
            next_ = lines[i + 1] if i + 1 < len(lines) else ""
            sc = score_address_candidate(ln, prev_, next_)
            if sc > 0:
                candidates.append((sc, _norm(ln)))

    # D) labeled wrapped blocks: capture after label until stop header
    text = "\n".join(raw_lines)
    for bm in BLOCK_LABEL_RE.finditer(text):
        start = bm.end()
        tail = text[start:]
        stop = STOP_HEADERS_RE.search(tail)
        chunk = tail[: stop.start()] if stop else tail
        chunk = _norm(chunk)
        if chunk and ST_ZIP_RE.search(chunk):
            candidates.append((2, chunk))

    if not candidates:
        return None

    candidates.sort(key=lambda x: (x[0], len(x[1])), reverse=True)

    for _, seg in candidates:
        parsed = _try_parse(seg)
        if parsed:
            return parsed

    return None
def fill_provider_address_from_special(
    specialInstructions: str,
    providerAddr: Optional[str],
    providerCity: Optional[str],
    providerState: Optional[str],
    providerZip: Optional[str],
    parse_address=None,
    providerFirst: str = "",
    providerLast: str = "",
    facilityName: str = "",
):
    """
    Fill ONLY missing provider fields, do not overwrite existing.

    IMPORTANT:
    - providerAddr will contain BOTH address lines (addr1 + addr2) if addr2 exists.
    - providerZip is forced to 5 digits.
    - Uses providerFirst/providerLast/facilityName as an anchor (free-form tolerant):
      we take the FIRST valid address block AFTER the anchor line if found.
    """
    if not (providerAddr and providerCity and providerState and providerZip):

        # NEW (Fix A): direct capture from "Address of Appt. location:" block
        # Only accept when City/State/ZIP are present (do NOT apply street-only fallback here).
        raw_appt_addr = extract_appt_address_block(specialInstructions)
        if raw_appt_addr:
            _parser = parse_address if callable(parse_address) else globals().get("parse_address")
            try:
                a1, a2, c_, st_, z_ = _parser(raw_appt_addr)
            except Exception:
                a1 = a2 = c_ = st_ = z_ = ""
            if a1 and c_ and st_ and z_:
                if not providerAddr:
                    providerAddr = a1 + (f", {a2}" if a2 else "")
                if not providerCity:
                    providerCity = c_
                if not providerState:
                    providerState = st_
                if not providerZip:
                    providerZip = str(z_)[:5]  # force 5 digits
            elif a1:
                if not providerAddr:
                    providerAddr = a1 + (f", {a2}" if a2 else "")
        addr_parts = extract_provider_address_dynamic(
            specialInstructions,
            parse_address=parse_address,
            provider_first=providerFirst,
            provider_last=providerLast,
            facility=facilityName,
        )
        if addr_parts:
            a1, a2, c_, st_, z_ = addr_parts

            if a1 and not providerAddr:
                providerAddr = a1 + (f", {a2}" if a2 else "")
            if c_ and not providerCity:
                providerCity = c_
            if st_ and not providerState:
                providerState = st_
            if z_ and not providerZip:
                providerZip = str(z_)[:5]  # force 5 digits

    return providerAddr, providerCity, providerState, providerZip
def extract_physical_address(filename: str, target_id) -> tuple[str | None, str]:
    target_id_up = str(target_id).strip().upper()
    physAddress: str | None = None
    mailAddress: str = ""  # only populated from Sheet 1; otherwise stays blank

    wb = openpyxl.load_workbook(filename, data_only=True)
    try:
        # 1) Sheet: "Liberty and Middle Market"
        try:
            ws1 = wb["Liberty and Middle Market"]
            max_row = ws1.max_row

            collecting = False
            physical_lines: list[str] = []
            mailing_lines: list[str] = []

            for i in range(6, max_row + 1):  # starts row 6
                cell_id = ws1.cell(row=i, column=1).value  # col A

                # New ID starts
                if cell_id not in (None, ""):
                    # if we were collecting and a new ID starts, stop
                    if collecting:
                        break

                    current_id = str(cell_id).strip()
                    collecting = target_id_up in current_id.upper()

                    if collecting:
                        physical_lines = []
                        mailing_lines = []

                if collecting:
                    # Physical address (col G)
                    physical1 = ws1.cell(row=i, column=7).value
                    if physical1 not in (None, ""):
                        physical_lines.append(str(physical1).strip())

                    # Mailing address (col H)
                    mailing1 = ws1.cell(row=i, column=8).value
                    if mailing1 not in (None, ""):
                        mailing_lines.append(str(mailing1).strip())

            if physical_lines:
                physAddress = formataddrSearchStr(" ".join(physical_lines))
            if mailing_lines:
                mailAddress = formataddrSearchStr(" ".join(mailing_lines))

        except KeyError:
            # sheet not present, just skip
            pass

        if physAddress:
            return physAddress, mailAddress  # mailAddress filled ONLY from sheet 1 (or blank)

        # 2) Sheet: "LM Physical Address"  (mailAddress stays blank)
        try:
            ws2 = wb["LM Physical Address"]
            max_row2 = ws2.max_row

            for i in range(2, max_row2 + 1):  # starts row 2
                cell_id = ws2.cell(row=i, column=1).value  # ID in col A
                if cell_id is None:
                    continue

                if target_id_up in str(cell_id).strip().upper():
                    addr_value = ws2.cell(row=i, column=6).value  # address in col F
                    if addr_value:
                        physAddress = formataddrSearchStr(str(addr_value))
                    break
        except KeyError:
            pass

        if physAddress:
            return physAddress, ""  # mailing blank (only from sheet 1)

        # 3) Sheet: "Manual Tracking LM Claim Office"  (mailAddress stays blank)
        try:
            ws3 = wb["Manual Tracking LM Claim Office"]
            max_row3 = ws3.max_row

            for i in range(1, max_row3 + 1):  # starts row 1
                cell_val = ws3.cell(row=i, column=1).value  # ID + address in col A
                if not cell_val:
                    continue

                text = str(cell_val).strip()
                text = text.replace("–", "-").replace("—", "-")

                if "-" not in text:
                    continue

                id_part, addr_part = text.split("-", 1)

                if target_id_up in id_part.strip().upper():
                    physAddress = formataddrSearchStr(addr_part.strip())
                    break
        except KeyError:
            pass

        if not physAddress:
            print(f"No physical address found for ID {target_id}")

        return physAddress, ""  # always blank mailing unless sheet 1 matched

    finally:
        wb.close()

def parse_referral_pdf(PDFData: str) -> dict:
    global correctZipCode

    # ---- defaults to avoid UnboundLocalError when a PDF format is missing sections ----
    providerFirst = ""
    providerLast = ""
    providerName = ""
    facilityName = ""
    providerAddr = ""
    providerCity = ""
    providerState = ""
    providerZip = ""
    providerPhone = ""
    specialInstructions = ""
    nextApptDate = ""
    nextApptTime = ""
    AccidentDesc = ""
    InjuryDesc = ""
    referralType = ""
    employerContactName = ""
    employerContactEmail = ""
    employerContactPhone = ""
    ncmContactName = ""
    referralNumber = ""
    ClaimAttyName = ""
    ClaimAttyaddressLine1 = ""
    ClaimAttyaddressLine2 = ""
    ClaimAttycity = ""
    ClaimAttystate = ""
    ClaimAttyZip = ""
    ClaimAttyPhoneNumber = ""
    injuryType = ""
    injuryCause = ""
    bodyPart = ""
    languageSpoken = ""
    URTServiceType = ""
    URTrefType = ""
    URTCaseObjective = ""
    adjAddrMail = ""
    try:
    # ---------- helpers ----------
        def trim(s): return s.strip() if s is not None else ""
        def split_lines(s): return s.split("\n")
        def after_colon(line):
            parts = line.split(":", 1)
            return trim(parts[1]) if len(parts) >= 2 else ""
    
        # def find_line_value(lines, label, rx=None):
        #     """Find first line containing label; return text after ':' or regex group."""
        #     for ln in lines:
        #         if label.lower() in ln.lower():
        #             if rx:
        #                 m = re.search(rx, ln, re.I)
        #                 if m:
        #                     # Return first capturing group or full match tail
        #                     return trim(m.group(1) if m.lastindex else m.group(0))
        #             return after_colon(ln)
        #     return ""
        
        def find_line_value(lines, label, rx=None, nth=1):
            """Return the nth occurrence (1-based) of label; by default returns the first."""
            n = 0
            label_lc = label.lower()
            for ln in lines:
                if label_lc in ln.lower():
                    if rx:
                        m = re.search(rx, ln, re.I)
                        if not m:
                            continue
                        val = (m.group(1) if m.lastindex else m.group(0)).strip()
                    else:
                        parts = ln.split(":", 1)
                        if "phone number" in label_lc:
                            if "Extension:" in parts[1].strip():
                                val = parts[1].strip().replace("Extension:","")
                            else:
                                val = parts[1].strip() if len(parts) >= 2 else ""
                        else:
                            val = parts[1].strip() if len(parts) >= 2 else ""
                    n += 1
                    if n == nth:
                        return val
            return ""
    
        def normalize_spaces(s: str) -> str:
            return re.sub(r"\s+", " ", s or "").strip()
    
        def pad2(n: str) -> str:
            return ("0" + n)[-2:]
    
        def normalize_date(d: str) -> str:
            d = (d or "").strip()
            # MM/DD/YYYY or M/D/YY
            m = re.match(r"^\s*(\d{1,2})/(\d{1,2})/(\d{2,4})\s*$", d)
            if m:
                mm, dd, yy = m.groups()
                if len(yy) == 2:
                    yy = ("19" if int(yy) >= 70 else "20") + yy
                return f"{pad2(mm)}/{pad2(dd)}/{yy}"
            # M/D (no year) -> use current year
            m = re.match(r"^\s*(\d{1,2})/(\d{1,2})\s*$", d)
            if m:
                mm, dd = m.groups()
                return f"{pad2(mm)}/{pad2(dd)}/{datetime.now().year}"
            return ""


        # def extract_next_appt_date_from_block(special_block: str) -> str:

        #     if not special_block:
        #         return ""

        #     lines = [ln.strip() for ln in special_block.replace("\u00A0", " ").splitlines()]

        #     # Pass 1: look for 'Date:' (case‑insensitive) with date on same line
        #     for ln in lines:
        #         m = re.search(r"(?i)\bdate\b\s*[:\-]\s*(\d{1,2}/\d{1,2}(?:/\d{2,4})?)", ln)
        #         if m:
        #             return normalize_date(m.group(1))

        #     # Pass 2: handle 'Date:' on one line and date on the next line
        #     for i, ln in enumerate(lines):
        #         if re.search(r"(?i)\bdate\b\s*[:\-]\s*$", ln):
        #             if i + 1 < len(lines):
        #                 m = re.search(r"(\d{1,2}/\d{1,2}(?:/\d{2,4})?)", lines[i + 1])
        #                 if m:
        #                     return normalize_date(m.group(1))

        #     # Pass 3 (fallback): any date in the block that looks like MM/DD(/YY)
        #     m = re.search(r"\b(\d{1,2}/\d{1,2}(?:/\d{2,4})?)\b", special_block)
        #     if m:
        #         return normalize_date(m.group(1))

        #     return ""
        def extract_next_appt_date_from_block(special_block: str) -> str:
            if not special_block:
                return ""
        
            special_block = special_block.replace("\u00A0", " ")
            lines = [ln.strip() for ln in special_block.splitlines() if ln.strip()]
        
            date_pattern = r"\b(\d{1,2}/\d{1,2}(?:/\d{2,4})?)\b"
        
            # 1. Highest priority: appointment-specific lines
            # Handles:
            # Appt. Date: 3/19/26
            # Appt. Date: Time: NOV 3/19/26 10:15AM
            # Appointment Date Time NOV 3/19/26
            appt_line_re = re.compile(
                r"(?i)\b(?:appt|apt|appointment)\.?\s*date\b"
            )
        
            for ln in lines:
                if appt_line_re.search(ln):
                    m = re.search(date_pattern, ln)
                    if m:
                        return normalize_date(m.group(1))
        
            # 2. Also support "Date of Appt" / "Next Appt Date"
            date_of_appt_re = re.compile(
                r"(?i)\b(?:date\s+of\s+(?:appt|apt|appointment)\.?|next\s+(?:appt|apt|appointment)\.?\s*date)\b"
            )
        
            for ln in lines:
                if date_of_appt_re.search(ln):
                    m = re.search(date_pattern, ln)
                    if m:
                        return normalize_date(m.group(1))
        
            # 3. Criteria on one line, date on next line
            criteria_only_re = re.compile(
                r"(?i)^\s*(?:appt|apt|appointment)\.?\s*date\s*[:\-]?\s*$"
            )
        
            for i, ln in enumerate(lines):
                if criteria_only_re.search(ln):
                    if i + 1 < len(lines):
                        m = re.search(date_pattern, lines[i + 1])
                        if m:
                            return normalize_date(m.group(1))
        
            # 4. Lower priority: generic Date label only if line starts with Date:
            # This prevents "Effective 3/1/19" from being selected.
            generic_date_re = re.compile(
                r"(?i)^\s*date\s*[:\-]?\s*"
            )
        
            for ln in lines:
                if generic_date_re.search(ln):
                    m = re.search(date_pattern, ln)
                    if m:
                        return normalize_date(m.group(1))
        
            # 5. Last fallback: avoid Effective dates if possible
            non_effective_text = "\n".join(
                ln for ln in lines
                if not re.search(r"(?i)\beffective\b", ln)
            )
        
            m = re.search(date_pattern, non_effective_text)
            if m:
                return normalize_date(m.group(1))
        
            # 6. True final fallback
            m = re.search(date_pattern, special_block)
            if m:
                return normalize_date(m.group(1))
        
            return ""

            
        #V1
        # def normalize_time(t: str) -> str:
        #     t = (t or "").strip()
        #     t = re.sub(r"^\s*@\s*", "", t)
        #     t = re.sub(r"\s+", " ", t)
    
        #     # compact 530pm / 0530pm
        #     m = re.match(r"^\s*(\d{3,4})\s*([ap]m)\s*$", t, re.I)
        #     if m:
        #         raw, ap = m.groups()
        #         if len(raw) == 3:
        #             h, mnt = raw[0], raw[1:]
        #         else:
        #             h, mnt = raw[:2], raw[2:]
        #         return f"{int(h)}:{pad2(mnt)} {ap.upper()}"
    
        #     # 12h with : and optional am/pm
        #     m = re.match(r"^\s*([01]?\d|2[0-3]):([0-5]\d)\s*([ap]m)?\s*$", t, re.I)
        #     if m:
        #         hh, mm, ap = m.groups()
        #         if ap:
        #             return f"{int(hh)}:{mm} {ap.upper()}"
        #         # convert 24h to 12h
        #         h = int(hh)
        #         ap = "PM" if h >= 12 else "AM"
        #         h12 = 12 if h == 0 else (h - 12 if h > 12 else h)
        #         return f"{h12}:{mm} {ap}"
    
        #     # pure 24h
        #     m = re.match(r"^\s*([01]?\d|2[0-3]):([0-5]\d)\s*$", t)
        #     if m:
        #         hh, mm = m.groups()
        #         h = int(hh)
        #         ap = "PM" if h >= 12 else "AM"
        #         h12 = 12 if h == 0 else (h - 12 if h > 12 else h)
        #         return f"{h12}:{mm} {ap}"
        #     return ""

        #V2
        # def normalize_time(t: str) -> str:
        #     t = (t or "").strip()
        #     t = re.sub(r"^\s*@\s*", "", t)
        #     t = re.sub(r"\s+", " ", t)
        
        #     # hour-only: 10am / 10 am / 10 a.m.
        #     m = re.match(r"(?i)^\s*([1-9]|1[0-2])\s*([ap])\.?\s*m\.?\s*$", t)
        #     if m:
        #         h, ap = int(m.group(1)), m.group(2).upper()
        #         return f"{h}:00 {ap}M"
        
        #     # compact: 0530pm / 530pm
        #     m = re.match(r"^\s*([0-2]?\d)([0-5]\d)\s*([ap])m\s*$", t, re.I)
        #     if m:
        #         h, mnt, ap = int(m.group(1)), m.group(2), m.group(3).upper()
        #         return f"{h}:{mnt} {ap}M"
        
        #     # with colon, optional am/pm: 1:30 / 1:30pm / 13:30
        #     m = re.match(r"^\s*([01]?\d|2[0-3]):([0-5]\d)\s*([ap])?m?\s*$", t, re.I)
        #     if m:
        #         hh, mm, ap = m.groups()
        #         if ap:
        #             return f"{int(hh)}:{mm} {ap.upper()}M"
        #         h = int(hh)
        #         ap = "PM" if h >= 12 else "AM"
        #         h12 = 12 if h == 0 else (h - 12 if h > 12 else h)
        #         return f"{h12}:{mm} {ap}"
        
        #     # pure 24h (fallback)
        #     m = re.match(r"^\s*([01]?\d|2[0-3]):([0-5]\d)\s*$", t)
        #     if m:
        #         hh, mm = m.groups()
        #         h = int(hh)
        #         ap = "PM" if h >= 12 else "AM"
        #         h12 = 12 if h == 0 else (h - 12 if h > 12 else h)
        #         return f"{h12}:{mm} {ap}"
        
        #     return ""


        def first_date(s: str) -> str:
            m = re.search(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", s)
            if m: return normalize_date(m.group(0))
            m = re.search(r"\b\d{1,2}/\d{1,2}\b", s)
            if m: return normalize_date(m.group(0))
            return "" 

        #4/29
        # def normalize_time(t: str) -> str:
        #     """
        #     Normalizes:
        #     - 2:00 p.m. / 2:00 pm / 2 pm -> 2:00 PM
        #     - 530pm / 05:30pm           -> 5:30 PM
        #     - 13:30                     -> 1:30 PM
        #     """
        #     t = (t or "").strip()
        #     if not t:
        #         return ""

        #     # Normalize all a.m./p.m. variants to "am"/"pm"
        #     # p.m., p. m., PM, P.M. -> pm
        #     # a.m., a. m., AM, A.M. -> am
        #     t = re.sub(r'(?i)\b([ap])\s*\.?\s*m\.?\b', r'\1m', t)

        #     # Clean up spaces and stray "@"
        #     t = re.sub(r"\s*@\s*", " ", t)
        #     t = re.sub(r"\s+", " ", t)

        #     # 1) h or h:mm with explicit am/pm: "2pm", "2 pm", "2:00pm", "2:00 pm"
        #     m = re.match(r'(?i)^\s*(\d{1,2})(?::(\d{2}))?\s*([ap])m\s*$', t)
        #     if m:
        #         h = int(m.group(1))
        #         mm = m.group(2) or "00"
        #         ap = m.group(3).upper()
        #         return f"{h}:{mm} {ap}M"

        #     # 2) Compact time with am/pm: "530pm", "0530pm"
        #     m = re.match(r'(?i)^\s*([01]?\d|2[0-3])([0-5]\d)\s*([ap])m\s*$', t)
        #     if m:
        #         h = int(m.group(1))
        #         mm = m.group(2)
        #         ap = m.group(3).upper()
        #         return f"{h}:{mm} {ap}M"

        #     # 3) 24h / no-am-pm clock: "13:30", "8:05", "2:00"
        #     m = re.match(r'^\s*([01]?\d|2[0-3]):([0-5]\d)\s*$', t)
        #     if m:
        #         h = int(m.group(1))
        #         mm = m.group(2)
        #         ap = "PM" if h >= 12 else "AM"
        #         h12 = 12 if h == 0 else (h - 12 if h > 12 else h)
        #         return f"{h12}:{mm} {ap}"

        #     return ""
        
        # def first_time(s: str) -> str:
        #     # prefer "at|@" chunks first
        #     m = re.search(r"\b(?:at|@)\s*([0-2]?\d:[0-5]\d\s*[ap]m|[0-2]?\d[0-5]\d\s*[ap]m)\b", s, re.I)
        #     if m: return normalize_time(m.group(1))
        #     m = re.search(r"\b([0-2]?\d[0-5]\d\s*[ap]m)\b", s, re.I)
        #     if m: return normalize_time(m.group(1))
        #     m = re.search(r"\b([0-2]?\d:[0-5]\d\s*[ap]m)\b", s, re.I)
        #     if m: return normalize_time(m.group(1))
        #     m = re.search(r"\b([0-2]?\d:[0-5]\d)\b", s)
        #     if m: return normalize_time(m.group(1))
        #     return ""


        # def first_time(s: str) -> str: #asdasd
        #     s = (s or "").replace("\u00A0", " ")  # normalize NBSP
        
        #     # "at 10am" / "at 10 am"
        #     m = re.search(r"(?i)\b(?:at|@)\s*([1-9]|1[0-2])\s*([ap])\.?\s*m\.?\b", s)
        #     if m:
        #         h, ap = int(m.group(1)), m.group(2).upper()
        #         return f"{h}:00 {ap}M"
        
        #     # bare hour-only "10am" / "10 am"
        #     m = re.search(r"(?i)\b([1-9]|1[0-2])\s*([ap])\.?\s*m\.?\b", s)
        #     if m:
        #         h, ap = int(m.group(1)), m.group(2).upper()
        #         return f"{h}:00 {ap}M"
        
        #     # your existing patterns
        #     m = re.search(r"(?i)\b(?:at|@)\s*([0-2]?\d:[0-5]\d\s*[ap]m|[0-2]?\d[0-5]\d\s*[ap]m)\b", s)
        #     if m: return normalize_time(m.group(1))
        #     m = re.search(r"(?i)\b([0-2]?\d[0-5]\d\s*[ap]m)\b", s)
        #     if m: return normalize_time(m.group(1))
        #     m = re.search(r"(?i)\b([0-2]?\d:[0-5]\d\s*[ap]m)\b", s)
        #     if m: return normalize_time(m.group(1))
        #     m = re.search(r"\b([0-2]?\d:[0-5]\d)\b", s)
        #     if m: return normalize_time(m.group(1))
        #     return ""

        #4/29
        # def first_time(s: str) -> str:
        #     s = (s or "").replace("\u00A0", " ")

        #     # "at 2 pm" / "at 2 p.m."
        #     m = re.search(r"(?i)\b(?:at|@)\s*([1-9]|1[0-2])\s*([ap])\.?\s*m\.?\b", s)
        #     if m:
        #         h = int(m.group(1))
        #         ap = m.group(2).upper()
        #         return f"{h}:00 {ap}M"

        #     # bare hour-only "2 pm" / "2 p.m."
        #     m = re.search(r"(?i)\b([1-9]|1[0-2])\s*([ap])\.?\s*m\.?\b", s)
        #     if m:
        #         h = int(m.group(1))
        #         ap = m.group(2).upper()
        #         return f"{h}:00 {ap}M"

        #     # explicit clock with am/pm: "2:00 p.m.", "2:00pm"
        #     m = re.search(r"(?i)\b([0-2]?\d:[0-5]\d\s*[ap]\s*\.?\s*m\.?)\b", s)
        #     if m:
        #         return normalize_time(m.group(1))

        #     # plain clock (no am/pm) – fallback, treated as 24h
        #     m = re.search(r"\b([0-2]?\d:[0-5]\d)\b", s)
        #     if m:
        #         return normalize_time(m.group(1))

        #     return ""

        def normalize_time(t: str) -> str:
            """
            Normalizes:
            - 8:0 AM / 8:00 AM       -> 8:00 AM
            - 2:00 p.m. / 2 pm       -> 2:00 PM
            - 530pm / 0530pm         -> 5:30 PM
            - 13:30                  -> 1:30 PM
            """
            t = (t or "").strip()
            if not t:
                return ""
        
            # Normalize a.m./p.m. variants to am/pm
            t = re.sub(r'(?i)\b([ap])\s*\.?\s*m\.?\b', r'\1m', t)
        
            # Clean spaces and stray @
            t = re.sub(r"\s*@\s*", " ", t)
            t = re.sub(r"\s+", " ", t).strip()
        
            # 1) h or h:mm with explicit am/pm
            # accepts 8 AM, 8:0 AM, 8:00 AM, 2:30pm
            m = re.match(r'(?i)^\s*(\d{1,2})(?::([0-5]?\d))?\s*([ap])m\s*$', t)
            if m:
                h = int(m.group(1))
                mm = m.group(2) or "00"
                mm = mm.zfill(2)
                ap = m.group(3).upper()
                return f"{h}:{mm} {ap}M"
        
            # 2) Compact time with am/pm: 530pm / 0530pm
            m = re.match(r'(?i)^\s*([01]?\d|2[0-3])([0-5]\d)\s*([ap])m\s*$', t)
            if m:
                h = int(m.group(1))
                mm = m.group(2)
                ap = m.group(3).upper()
                return f"{h}:{mm} {ap}M"
        
            # 3) 24h or no-am-pm clock: 13:30, 8:0, 8:00
            m = re.match(r'^\s*([01]?\d|2[0-3]):([0-5]?\d)\s*$', t)
            if m:
                h = int(m.group(1))
                mm = m.group(2).zfill(2)
                ap = "PM" if h >= 12 else "AM"
                h12 = 12 if h == 0 else (h - 12 if h > 12 else h)
                return f"{h12}:{mm} {ap}"
        
            return ""
        
        
        def first_time(s: str) -> str:
            s = (s or "").replace("\u00A0", " ")
        
            # Normalize OCR spacing
            s = re.sub(r"\s+", " ", s).strip()
        
            # 1) "at 8:0 AM", "@ 8:00 AM", "at 2 pm"
            m = re.search(
                r"(?i)\b(?:at|@)\s*(\d{1,2})(?::([0-5]?\d))?\s*([ap])\.?\s*m\.?\b",
                s
            )
            if m:
                h = m.group(1)
                mm = m.group(2)
                ap = m.group(3)
                raw = f"{h}:{mm or '00'} {ap}m"
                return normalize_time(raw)
        
            # 2) Bare time with AM/PM: "8:0 AM", "8:00 AM", "2 pm"
            m = re.search(
                r"(?i)\b(\d{1,2})(?::([0-5]?\d))?\s*([ap])\.?\s*m\.?\b",
                s
            )
            if m:
                h = m.group(1)
                mm = m.group(2)
                ap = m.group(3)
                raw = f"{h}:{mm or '00'} {ap}m"
                return normalize_time(raw)
        
            # 3) Plain clock fallback: "8:0", "8:00", "13:30"
            m = re.search(r"\b([0-2]?\d):([0-5]?\d)\b", s)
            if m:
                return normalize_time(f"{m.group(1)}:{m.group(2)}")
        
            return ""

        # ---------- normalize input ----------
        text = PDFData.replace("\r", "")
        text = text.replace("\u00A0", " ")  # NBSP -> space
        lines = split_lines(text)
    
        # ---------- Customer Contact Name (label says "Customer Name:" in your sample) ----------
        customer = find_line_value(lines, "Customer Name:")

        if customer.lower().strip() =="ups":
            customer= "United Parcel Service"
        elif customer.lower().strip() =="comcast":
            customer= "Comcast Cablevision"
        #dxCode
        dxCode = find_line_value(lines, "Diagnosis Code:")
    
        # ---------- Claimant block & names ----------
        claimantFull = ""
        claimantFirst = ""
        claimantLast = ""
    
        # capture Claimant Information block
        block = []
        capturing = False
        end_headers = re.compile(
            r"^(Claim Information|Provider Information|Referral Instructions|Special Instructions|"
            r"Case Manager Information|Vendor Information|Attorney Information)\b", re.I)
        for ln in lines:
            if not capturing and "Claimant Information" in ln:
                capturing = True
                continue
            if capturing:
                if end_headers.search(ln):
                    break
                block.append(ln)
    
        # find "Name:" in block
        if not claimantFull:
            for ln in block:
                if "Name:" in ln:
                    m = re.search(r"(?i)Name:\s*(.+)$", ln)
                    if m:
                        claimantFull = trim(m.group(1))
                        break
    
        # fallback: elsewhere (e.g., "Claimant Name:")
        if not claimantFull:
            for ln in lines:
                if "Name:" in ln and "Claimant" in ln:
                    m = re.search(r"(?i)Name:\s*(.+)$", ln)
                    if m:
                        claimantFull = trim(m.group(1))
                        break
    
        # split first/last
        if claimantFull:
            if "," in claimantFull:  # LAST, FIRST [M...]
                left, right = [trim(x) for x in claimantFull.split(",", 1)]
                first_token = right.split()[0] if right else right
                claimantFirst, claimantLast = first_token, left
            else:
                toks = claimantFull.split()
                if len(toks) >= 2:
                    claimantFirst, claimantLast = toks[0], toks[-1]
                else:
                    claimantLast = claimantFull
    
        # ---------- Claim Number ----------
        claimNumber = ""
        for ln in lines:
            if "Claim Number:" in ln:
                claimNumber = after_colon(ln) or re.search(r"Claim Number:\s*([^\r\n]+)", ln).group(1).strip()
                break
        
        # ---------- Claim ID ----------
        claimID = ""
        for ln in lines:
            if "Claim ID:" in ln:
                try:
                    claimID = after_colon(ln) or re.search(r"Claim ID:\s*([^\r\n]+)", ln).group(1).strip()
                except:
                    pass
                break

        # ---------- Claim Number ----------
        claimTypeLOI = ""
        for ln in lines:
            if "Claim Type:" in ln:
                claimTypeLOI = after_colon(ln) or re.search(r"Claim Type:\s*([^\r\n]+)", ln).group(1).strip()
                break
        if "wc" in claimTypeLOI.lower():
            claimTypeLOI = "Workers Compensation"
        elif "auto" in claimTypeLOI.lower():
            claimTypeLOI = "Auto No-Fault"
        else:
            claimTypeLOI = "Workers Compensation"
    
        # ---------- Referral Type ----------
        referralType = find_line_value(lines, "Referral Type:", r"Referral Type:\s*(.+)")
    
        # Lists (you can use as needed)
        list1 = {"Field Case Management", "One-Time RN Visit - Provider", "One-Time RN Visit"}
        list2 = {"Vocational Rehabilitation", "Independent Medical Exam"}
        # InList equivalent:
        def in_list(value, s): return value in s
    
        # ---------- Claim Jurisdiction (State/Jurisdiction of Claim) ----------
        state_map = {
            "AL":"Alabama","AK":"Alaska","AZ":"Arizona","AR":"Arkansas","CA":"California","CO":"Colorado","CT":"Connecticut",
            "DE":"Delaware","FL":"Florida","GA":"Georgia","HI":"Hawaii","ID":"Idaho","IL":"Illinois","IN":"Indiana","IA":"Iowa",
            "KS":"Kansas","KY":"Kentucky","LA":"Louisiana","ME":"Maine","MD":"Maryland","MA":"Massachusetts","MI":"Michigan",
            "MN":"Minnesota","MS":"Mississippi","MO":"Missouri","MT":"Montana","NE":"Nebraska","NV":"Nevada","NH":"New Hampshire",
            "NJ":"New Jersey","NM":"New Mexico","NY":"New York","NC":"North Carolina","ND":"North Dakota","OH":"Ohio","OK":"Oklahoma",
            "OR":"Oregon","PA":"Pennsylvania","RI":"Rhode Island","SC":"South Carolina","SD":"South Dakota","TN":"Tennessee",
            "TX":"Texas","UT":"Utah","VT":"Vermont","VA":"Virginia","WA":"Washington","WV":"West Virginia","WI":"Wisconsin","WY":"Wyoming"
        }
        claimStateAbbr = ""
        for ln in lines:
            if "State/Jurisdiction of Claim:" in ln:
                val = after_colon(ln)
                claimStateAbbr = re.sub(r"[^\w]", "", val or "").upper()
                break
        claimStateFull = state_map.get(claimStateAbbr, "")
    
        # ---------- Address lines / City / State / Zip / Phone ----------
        addressLine1 = find_line_value(lines, "Address-line-1:")
        addressLine2 = find_line_value(lines, "Address-line-2:")
        city         = find_line_value(lines, "City:")
        state        = find_line_value(lines, "State:", r"State:\s*(\w+)")
        zip_ = ""
        for ln in lines:
            if "Zip:" in ln:
                m = re.search(r"Zip:\s*([0-9\-]+)", ln)
                if m:
                    raw = m.group(1)
                    m2 = re.match(r"(\d{5})", raw)
                    zip_ = m2.group(1) if m2 else raw.strip()
                break
        phoneNumber  = find_line_value(lines, "Phone Number:")
        if phoneNumber == "999-9999-999":
            phoneNumber = "989-9999-999"
    
        # ---------- Gender ----------
        gender = ""
        for ln in lines:
            if "Gender:" in ln:
                m = re.search(r"(?i)Gender:\s*([A-Za-z]+)", ln)
                if m:
                    gender = trim(m.group(1))
                    break
        if not gender:
            if re.search(r"\bShe\b|\bHer\b", text): gender = "Female"
            elif re.search(r"\bHe\b|\bHis\b", text): gender = "Male"
    
        # ---------- DOB / DOI ----------
        dob = find_line_value(lines, "Date of Birth:")
        doi = find_line_value(lines, "Date of Injury/Accident/Illness:")

        # ------- SSN ---------
        claimantSSN = find_line_value(lines, "Social Security Number:")

        # ---------- Referral Source & Adjuster ----------
        refSource = find_line_value(lines, "Nurse Case Manager E-mail Address:")
        adjuster  = find_line_value(lines, "Claims Case Manager Name:")
        officeNumber = find_line_value(lines, "Claims Office Number:")
        adjEmail = find_line_value(lines, "Claims Case Manager E-mail Address:")
        adjPhone = find_line_value(lines,"Office Phone Number:")
        if officeNumber.startswith("0"):
            officeNumber = officeNumber[1:]
        adjAddrFull, adjAddrMail = extract_physical_address(str(LM_MATRIX_PATH), officeNumber)
        adjAddr1,adjAddr2,adjCity,adjState,adjZip = parse_address(adjAddrFull)






        # ---------- Attorney Info ----------
        ClaimAttyName = find_line_value(lines,"Attorney Name:")
        ClaimAttyaddressLine1 = find_line_value(lines, "Address-line-1:",nth=2)
        ClaimAttyaddressLine2 = find_line_value(lines, "Address-line-2:",nth=2)
        ClaimAttycity         = find_line_value(lines, "City:",nth=2)
        ClaimAttystate        = find_line_value(lines, "State:", r"State:\s*(\w+)",nth=2)
        ClaimAttyZip          = find_line_value(lines, "Zip:", r"(?i)\bZip(?:\s*Code)?\s*:\s*(\d{5})(?:-\d{4})?",nth=2)
        ClaimAttyPhoneNumber  = find_line_value(lines, "Phone Number:",nth=5)
        # ---------- Special Instructions (block until 'Referrer Name:' etc.) ----------
        specialInstructions = ""
        nextApptDate = ""
        collecting = False
        for ln in lines:
            if not collecting and re.search(r"(?i)\bSpecial\s*Instructions\s*:", ln):
                # include tail on the same line after colon
                m = re.search(r"(?i)\bSpecial\s*Instructions\s*:\s*(.*)", ln)
                if m and trim(m.group(1)):
                    specialInstructions += trim(m.group(1)) + "\n"
                collecting = True
                continue
            if collecting:
                if re.search(r"(?i)^\s*(Referrer\s*Name:|©|Page\s+\d+\s+of\s+\d+|Vendor Name:)\b", ln):
                    break
                if ln.strip():
                    specialInstructions += ln.strip() + "\n"
        specialInstructions = specialInstructions.strip()

        nextApptDate = extract_next_appt_date_from_block(specialInstructions)

        # if you still want a global fallback:
        # if not nextApptDate:
        #     m = re.search(r"\b(\d{1,2}/\d{1,2}(?:/\d{2,4})?)\b", specialInstructions)
        #     if m:
        #         nextApptDate = normalize_date(m.group(1))

    
        # ---------- Employer contact name from Special Instructions ----------
        def _clean_contact_fragment(fragment: str):
            """Clean the raw value after an employer/contact label."""
            if not fragment:
                return None

            frag = fragment.strip(" -;:,").strip()
            lower = frag.lower()

            # Ignore obvious non-name phrases
            non_name_prefixes = (
                "account instructions",
                "approval is not required",
                "no communication between employer and ncm",
                "n/a",
                "na",
                "none",
            )
            if any(lower.startswith(p) for p in non_name_prefixes):
                return None

            # If there is an email, drop everything after the first " - "
            if "@" in frag:
                parts = re.split(r"\s+-\s+", frag, 1)
                if len(parts) > 1:
                    frag = parts[0].strip()

            # Strip emails and <angle brackets>
            frag = re.sub(r"<[^>]+>", "", frag)  # remove <...>
            frag = re.sub(r"\b[\w\.-]+@[\w\.-]+\.\w+\b", "", frag)

            # Remove obvious "Phone" tails if present on same line
            frag = re.split(r"\bphone\b|ph #|ph#|tel\.", frag, flags=re.IGNORECASE)[0]

            # Normalize whitespace
            frag = " ".join(frag.split())

            return frag or None

        # ---------- Employer contact details from Special Instructions ----------
 
        def _extract_emails(value: str) -> list[str]:
            if not value:
                return []
        
            value = value.replace("<", " ").replace(">", " ")
        
            emails = re.findall(
                r"(?<![\w.-])[\w\.-]+@[\w\.-]+\.[A-Za-z]{2,}(?![\w.-])",
                value
            )
        
            cleaned = []
            for e in emails:
                e = e.strip(" .;,)")
                if e and e.lower() not in [x.lower() for x in cleaned]:
                    cleaned.append(e)
        
            return cleaned
        
        
        def _extract_phones(value: str) -> list[str]:
            if not value:
                return []
        
            matches = re.findall(
                r"(?<!\d)(?:\(?\d{3}\)?|\d{3}\))[\s\.-]*\d{3}[\s\.-]*\d{4}(?!\d)(?:\s*(?:ext\.?|extension)\s*\d+)?",
                value,
                flags=re.I
            )
        
            phones = []
        
            for phone in matches:
                ext_match = re.search(r"(?i)\b(?:ext\.?|extension)\s*(\d+)", phone)
                ext = ext_match.group(1) if ext_match else ""
        
                digits = re.sub(r"\D", "", phone)
        
                if len(digits) >= 10:
                    normalized = f"{digits[0:3]}-{digits[3:6]}-{digits[6:10]}"
                    if ext:
                        normalized = f"{normalized} EXT {ext}"
                else:
                    normalized = phone.strip()
        
                if normalized and normalized not in phones:
                    phones.append(normalized)
        
            return phones
        
        
        def _is_blank_employer_value(value: str) -> bool:
            if not value:
                return True
        
            value_check = value.lower().strip()
            value_check = re.sub(r"\s+", " ", value_check)
        
            value_check_no_punct = re.sub(r"[^a-z0-9\s]", "", value_check)
            value_check_no_punct = re.sub(r"\s+", " ", value_check_no_punct).strip()
        
            bad_values = {
                "n/a",
                "n.a.",
                "n.a",
                "na",
                "none",
                "unknown",
                "not applicable",
        
                "no contact required",
                "no contact needed",
                "no contact",
                "no employer contact",
                "no employer contact required",
                "no customer contact needed",
                "customer contact not needed",
        
                "no communication between employer and ncm",
                "adjuster will make contacts",
                "we will address all contacts with customer per protocol",
        
                "multiple per ssi",
                "multiple per ssis",
        
                "i can update employer",
                "i can update employer no atty on file",
                "tncm/cs will update",
                "tncm cs will update",
                "will update",
                "customer will update",
                "claims will update",
                "claim specialist will update",
        
                "does not require ncm contact",
                "do not require ncm contact",
                "does not require contact",
                "thd does not require ncm contact",
        
                "do not contact the store",
                "do not contact customer",
                "do not contact employer",
                "do not contact",
        
                "defer",
            }
        
            bad_values_no_punct = {
                re.sub(r"[^a-z0-9\s]", "", x.lower()).strip()
                for x in bad_values
            }
        
            if value_check in bad_values or value_check_no_punct in bad_values_no_punct:
                return True
        
            instruction_phrases = [
                "no contact required",
                "no contact needed",
                "no customer contact",
                "no employer contact",
                "adjuster will make contacts",
                "no communication between employer and ncm",
                "approval is not required",
                "per protocol",
                "i can update employer",
                "tncm/cs will update",
                "tncm cs will update",
                "will update",
                "do not contact",
                "do not contact the store",
                "does not require ncm contact",
                "do not require ncm contact",
                "does not require contact",
                "thd does not require ncm contact",
                "job descriptions are no provided",
                "all jobs are in the heavy category",
            ]
        
            return any(p in value_check for p in instruction_phrases)
        
        
        def _title_name_keep_apostrophe(name: str) -> str:
            if not name:
                return ""
        
            parts = []
        
            for token in name.split():
                if "'" in token:
                    sub = token.split("'")
                    token = "'".join(x[:1].upper() + x[1:].lower() for x in sub if x)
                else:
                    token = token[:1].upper() + token[1:].lower()
        
                parts.append(token)
        
            return " ".join(parts)
        
        
        def _name_from_email(email: str) -> str:
            if not email or "@" not in email:
                return ""
        
            local = email.split("@", 1)[0].strip().lower()
        
            generic_locals = {
                "info", "inbox", "claims", "claim", "workcomp", "wc", "hr",
                "support", "admin", "contact", "customerservice", "customer.service",
                "service", "noreply", "no.reply", "donotreply", "do.not.reply",
                "risk", "safety", "benefits", "payroll", "team", "office",
                "referrals", "intake", "medical", "provider", "providers",
                "customers", "customer", "mail", "email"
            }
        
            if local in generic_locals:
                return ""
        
            if re.search(r"\d", local):
                return ""
        
            # Only infer when email has separator like first.last or first_last.
            # Do not guess from edillard@ups.com or marciarodriguez@ups.com.
            if not re.search(r"[._-]", local):
                return ""
        
            parts = re.split(r"[._-]+", local)
            parts = [p for p in parts if len(p) >= 2]
        
            if len(parts) < 2:
                return ""
        
            if any(p in generic_locals for p in parts):
                return ""
        
            return _title_name_keep_apostrophe(" ".join(parts))
        
        
        def _remove_emails_and_phones(value: str) -> str:
            if not value:
                return ""
        
            value = re.sub(
                r"(?<![\w.-])[\w\.-]+@[\w\.-]+\.[A-Za-z]{2,}(?![\w.-])",
                " ",
                value
            )
        
            value = re.sub(
                r"(?<!\d)(?:\(?\d{3}\)?|\d{3}\))[\s\.-]*\d{3}[\s\.-]*\d{4}(?!\d)(?:\s*(?:ext\.?|extension)\s*\d+)?",
                " ",
                value,
                flags=re.I
            )
        
            return " ".join(value.split())
        
        
        def _has_email_glued_by_dash(value: str) -> bool:
            """
            Detect risky pattern:
                GT Lomas-georgelomas@ups.com
        
            In this format, the text before the dash may look like a name,
            but the dash is actually glued to the email username.
            To avoid wrong contact name, return Unknown later if email exists.
            """
            if not value:
                return False
        
            return re.search(
                r"[A-Za-z]\s*-\s*[\w\.-]+@[\w\.-]+\.[A-Za-z]{2,}",
                value
            ) is not None
        
        
        def _looks_like_company_only(value: str) -> bool:
            if not value:
                return True
        
            original = value.strip()
            v = original.lower().strip()
        
            v = re.sub(r"[^a-z0-9\s&./-]", " ", v)
            v = re.sub(r"\s+", " ", v).strip()
        
            if not v:
                return True
        
            instruction_phrases = [
                "does not require ncm contact",
                "do not require ncm contact",
                "does not require contact",
                "do not contact",
                "do not contact the store",
                "no contact",
                "no employer contact",
                "no customer contact",
                "will update",
                "can update employer",
                "defer",
                "per protocol",
                "adjuster will make contacts",
                "claims will coordinate",
                "claim specialist will update",
                "ncm should contact",
            ]
        
            if any(p in v for p in instruction_phrases):
                return True
        
            tokens = [t for t in re.split(r"\s+", v) if t]
        
            if not tokens:
                return True
        
            # Real person pattern: First Last / First Middle Last / O'Keefe Denise etc.
            if re.fullmatch(r"[A-Za-z][A-Za-z'\-]+(?:\s+[A-Za-z][A-Za-z'\-]+){1,3}", original):
                return False
        
            # Allow first-name-only values from employer contact line.
            # Example: Angelique, Cindy, James
            if re.fullmatch(r"[A-Za-z][A-Za-z'\-]+", original):
                return False
        
            company_words = {
                "inc", "llc", "ltd", "corp", "corporation", "company", "co",
                "holdings", "management", "services", "service", "solutions",
                "electric", "construction", "transportation", "logistics",
                "foods", "food", "health", "healthcare", "insurance",
                "home", "depot", "amazon", "comcast", "ups", "store",
                "customer", "employer", "transport", "rubber", "tire",
                "university", "sports", "entertainment", "group", "systems",
                "materials", "manufacturing", "distribution", "restaurant",
                "restaurants", "mining", "gas", "authority", "furniture",
                "dms",
            }
        
            if any(t.strip(".") in company_words for t in tokens):
                # Allow company prefix + person suffix:
                # Ruan Melissa Dougherty
                # DMS Mona Garfias
                if re.search(r"\b[A-Z][a-zA-Z'\-]+\s+[A-Z][a-zA-Z'\-]+\b$", original):
                    return False
                return True
        
            if original.isupper() and len(tokens) <= 5:
                return True
        
            if "." in original and len(tokens) <= 2:
                return True
        
            return False
        
        
        def _clean_employer_name(value: str) -> str:
            if not value:
                return ""
        
            value = value.replace("\u00A0", " ")
            value = value.replace("<", " <").replace(">", "> ")
            value = value.strip(" -;:,").strip()
        
            # Outlook-style value:
            # Jennifer Wells - Wells, Jennifer (CONTR) WellsJL@nv.doe.gov
            # Keep: Jennifer Wells
            if "<" in value and ">" in value:
                before_email = value.split("<", 1)[0].strip(" -;:,")
        
                if " - " in before_email:
                    before_email = before_email.split(" - ", 1)[0].strip(" -;:,")
        
                before_email = re.sub(r"\([^)]*\)", "", before_email).strip()
        
                if before_email and not _is_blank_employer_value(before_email):
                    value = before_email
        
            # Handle missing space before labels:
            # Shawn RayPhone #: 903-...
            value = re.sub(
                r"(?i)([a-z])(?=Phone\s*#?\s*:|Phone\s*#?\s*or\s*Email\s*Address\s*:|Ph\.?\s*#?\s*:|Email\s*:)",
                r"\1 ",
                value
            )
        
            # Handle compressed label:
            # Shawn RayPhone # or Email Address: 903-561-2900
            value = re.sub(
                r"(?i)\bPhone\s*#?\s*or\s*Email\s*Address\s*[:\-].*$",
                " ",
                value
            )
        
            # Remove labels and everything after them when embedded in same line
            value = re.sub(r"(?i)\bEmail\s*[:\-].*$", " ", value)
            value = re.sub(r"(?i)\bPhone\s*#?\s*[:\-].*$", " ", value)
            value = re.sub(r"(?i)\bPh\.?\s*#?\s*[:\-].*$", " ", value)
        
            value = _remove_emails_and_phones(value)
        
            value = re.split(
                r"(?i)\s+-\s+ok\s+to\b|"
                r"\s+-\s+okay\s+to\b|"
                r"\s+-\s+cc\b|"
                r"\s+-\s+phone\b|"
                r"\s+-\s+email\b|"
                r"\s+-\s+preferred\b",
                value
            )[0]
        
            value = " ".join(value.split())
            value = value.strip(" -;:,.")
        
            if _is_blank_employer_value(value):
                return ""
        
            # UPS - Juel Gumbs -> Juel Gumbs
            # Comcast - no contact required -> blank
            if " - " in value:
                left, right = [x.strip() for x in value.split(" - ", 1)]
        
                if _is_blank_employer_value(right):
                    return ""
        
                if right:
                    value = right
        
            value = " ".join(value.split())
            value = value.strip(" -;:,.")
        
            if _is_blank_employer_value(value):
                return ""
        
            return value
        
        
        def _split_employer_names(value: str) -> list[str]:
            value = _clean_employer_name(value)
        
            if not value:
                return []
        
            known_company_prefixes = [
                "ruan",
                "ups",
                "amazon",
                "amazon.com",
                "comcast",
                "the home depot",
                "home depot",
                "general mills",
                "petco",
                "costco",
                "mars",
                "goodyear",
                "dms",
            ]
        
            for prefix in known_company_prefixes:
                m = re.match(
                    rf"(?i)^\s*{re.escape(prefix)}\s+([A-Z][a-zA-Z'\-]+(?:\s+[A-Z][a-zA-Z'\-]+){{1,3}})\s*$",
                    value
                )
                if m:
                    value = m.group(1).strip()
                    break
        
            # Builders First Source, Inc. David Love -> David Love
            m_company_person = re.search(
                r"(?i)\b(?:inc|llc|ltd|corp|corporation|company|co)\.?\s+([A-Z][a-zA-Z'\-]+(?:\s+[A-Z][a-zA-Z'\-]+){1,3})\s*$",
                value
            )
        
            if m_company_person:
                value = m_company_person.group(1).strip()
        
            # Joey Jones. UPS -> Joey Jones
            value = re.sub(
                r"(?i)\s*[\.\-]\s*(UPS|Amazon|Comcast|General Mills|Petco|Mars|Costco|Goodyear|DMS)\s*$",
                "",
                value
            ).strip()
        
            names = []
        
            # Handle comma-only value
            if "," in value and not re.search(r"\s*&\s*|\s+and\s+|;", value, re.I):
                comma_parts = [p.strip(" ,;-.") for p in value.split(",") if p.strip(" ,;-.")]
        
                # Ewing, Christal -> Christal Ewing
                if len(comma_parts) == 2:
                    left, right = comma_parts
                    company_suffixes = {"inc", "llc", "ltd", "corp", "co", "corporation", "company"}
        
                    if right.lower().strip(".") not in company_suffixes:
                        if len(left.split()) == 1 and len(right.split()) == 1:
                            candidate = f"{right} {left}"
                            if not _looks_like_company_only(candidate):
                                names.append(candidate)
                                return names
        
                # Angelique, Cindy, James -> Angelique | Cindy | James
                if len(comma_parts) >= 2:
                    for part in comma_parts:
                        if part and not _is_blank_employer_value(part) and not _looks_like_company_only(part):
                            if part.lower() not in [x.lower() for x in names]:
                                names.append(part)
                    return names
        
            # Split multiple contacts by &, and, semicolon
            parts = re.split(r"\s*&\s*|\s+and\s+|;", value, flags=re.I)
        
            for part in parts:
                part = part.strip(" ,;-.")
        
                if not part or _is_blank_employer_value(part):
                    continue
        
                # Last, First -> First Last
                if "," in part:
                    left, right = [x.strip() for x in part.split(",", 1)]
                    company_suffixes = {"inc", "llc", "ltd", "corp", "co", "corporation", "company"}
        
                    if right.lower().strip(".") not in company_suffixes:
                        if left and right:
                            part = f"{right} {left}"
        
                part = " ".join(part.split())
                part = part.strip(" ,;-.")
        
                if not part or _is_blank_employer_value(part):
                    continue
        
                if _looks_like_company_only(part):
                    continue
        
                if part.lower() not in [x.lower() for x in names]:
                    names.append(part)
        
            return names
        
        
        def extract_employer_contact_details(
            special_text: str,
            use_unknown_for_email_only: bool = True
        ) -> tuple[str, str, str]:
            """
            Returns:
                employerContactName, employerContactEmail, employerContactPhone
            """
        
            if not special_text:
                return "", "", ""
        
            raw_text = special_text.replace("\u00A0", " ")
        
            # Add line breaks before labels to handle compressed PDF text.
            raw_text = re.sub(
                r"(?i)(?<!\n)(Employer\s+(?:name\s+)?(?:or\s+)?contact\s*:)",
                r"\n\1",
                raw_text
            )
            raw_text = re.sub(
                r"(?i)(?<!\n)(Employer\s+name\s+and\s+contact\s*:)",
                r"\n\1",
                raw_text
            )
            raw_text = re.sub(
                r"(?i)(?<!\n)(Employer\s+Contact\s*&\s*Phone\s*:)",
                r"\n\1",
                raw_text
            )
            raw_text = re.sub(r"(?i)(?<!\n)(Employer\s*:)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(Contact\s+name\s*:)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(Phone\s*#?\s*or\s*Email\s*Address\s*:)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(Phone\s*#?\s*:)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(Ph\.?\s*#?\s*:)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(Email\s*:)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(SSIs\s*:)", r"\n\1", raw_text)
        
            raw_text = re.sub(r"(?i)(?<!\n)(Atty\s+Permission\s*:)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(Atty\s+represented\s*:)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(Attorney\s+Permission\s*:)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(Attorney\s*:)", r"\n\1", raw_text)
        
            raw_text = re.sub(r"(?i)(?<!\n)(Insurer\s+Information\s*:?)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(COMPENSABLE\s+BODY)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(WORK\s+STATUS\s*:)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(INSTRUCTIONS\s*:)", r"\n\1", raw_text)
            raw_text = re.sub(r"(?i)(?<!\n)(Referrer\s+Name\s*:)", r"\n\1", raw_text)
        
            lines_si = [ln.strip() for ln in raw_text.splitlines() if ln.strip()]
        
            employer_start_re = re.compile(
                r"^\s*(?:"
                r"Employer\s+(?:name\s+)?(?:or\s+)?contact\b|"
                r"Employer\s+name\s+and\s+contact\b|"
                r"Employer\s+Contact\s*&\s*Phone\b|"
                r"Employer\s*:"
                r")",
                re.I
            )
        
            employer_exclude_re = re.compile(
                r"^\s*Employer\s+FEIN\b|"
                r"^\s*Employer\s+preferred\s+contact\s+method\b",
                re.I
            )
        
            stop_re = re.compile(
                r"^\s*(?:"
                r"Atty\b|"
                r"Attorney\b|"
                r"Attorney\s*:|"
                r"Atty\s+represented\b|"
                r"Atty\s+Permission\b|"
                r"Attorney\s+Permission\b|"
                r"CM\s+obtained\s+attorney\s+permission\b|"
                r"Compensable\b|"
                r"COMPENSABLE\b|"
                r"COMPENSABLE\s+BODY\b|"
                r"Current\s+Work\s+Status\b|"
                r"CURRENT\s+WORK\s+STATUS\b|"
                r"Work\s+Status\b|"
                r"WORK\s+STATUS\b|"
                r"Referral\s+Goal\b|"
                r"Referral\s+Instructions\b|"
                r"REFERRAL\s+INSTRUCTIONS\b|"
                r"Outline\b|"
                r"INSTRUCTIONS\b|"
                r"Additional\s+Instructions\b|"
                r"Additional\s+expectations\b|"
                r"Referrer\s+Name\b|"
                r"SSIs\b|"
                r"Insurer\s+Information\b|"
                r"Employee\s+SS#\b|"
                r"State\s+File\s+#\b|"
                r"Employer\s+FEIN\b|"
                r"Vendor\s+Nurse\b|"
                r"TN\s+Onsite\b|"
                r"State\s+File\b"
                r")",
                re.I
            )
        
            employer_block = []
            started = False
        
            for ln in lines_si:
                if not started:
                    if employer_exclude_re.search(ln):
                        continue
        
                    if employer_start_re.search(ln):
                        started = True
                        employer_block.append(ln)
                        continue
                else:
                    if stop_re.search(ln):
                        break
        
                    employer_block.append(ln)
        
            if not employer_block:
                return "", "", ""
        
            block_text = "\n".join(employer_block)
        
            names = []
            emails = []
            phones = []
        
            # Get emails/phones from the employer block only.
            emails.extend(_extract_emails(block_text))
            phones.extend(_extract_phones(block_text))
        
            # Extract names from employer line.
            for i, ln in enumerate(employer_block):
                raw_name = ""
        
                patterns = [
                    r"^\s*Employer\s+(?:name\s+)?(?:or\s+)?contact(?:\s*\(.*?\))?\s*:+\s*(.*)$",
                    r"^\s*Employer\s+name\s+and\s+contact\s*:+\s*(.*)$",
                    r"^\s*Employer\s+Contact\s*&\s*Phone\s*:+\s*(.*)$",
                    r"^\s*Employer\s*:+\s*(.*)$",
                ]
        
                for pat in patterns:
                    m = re.match(pat, ln, re.I)
                    if m:
                        raw_name = m.group(1).strip()
                        break
        
                if raw_name:
                    # Blum-type case:
                    # GT Lomas-georgelomas@ups.com
                    # Do not capture GT Lomas as a name.
                    if not _has_email_glued_by_dash(raw_name):
                        names.extend(_split_employer_names(raw_name))
                else:
                    # If employer label is blank, check the next line only if it is not another label/header.
                    if employer_start_re.search(ln) and i + 1 < len(employer_block):
                        next_line = employer_block[i + 1].strip()
        
                        if not re.match(
                            r"^\s*(Employer preferred|Phone|Ph\s*#|Email|Atty|Attorney|SSIs|Compensable|Work Status|Insurer Information|Employee SS#|State File #|Employer FEIN)\b",
                            next_line,
                            re.I
                        ):
                            names.extend(_split_employer_names(next_line))
        
            # Contact Name is stronger than Employer company.
            # Example:
            # Employer: UPS
            # Contact name: Sylvia Keller
            contact_names = []
        
            for ln in employer_block:
                m = re.match(r"^\s*Contact\s+name\s*:+\s*(.*)$", ln, re.I)
                if m:
                    contact_names.extend(_split_employer_names(m.group(1).strip()))
        
            if contact_names:
                names = contact_names
        
            # Infer from email only when safe.
            if not names and emails:
                inferred_names = []
        
                for email in emails:
                    inferred = _name_from_email(email)
        
                    if inferred and inferred.lower() not in [x.lower() for x in inferred_names]:
                        inferred_names.append(inferred)
        
                names = inferred_names
        
            final_names = []
        
            for n in names:
                n = n.strip(" ,;-.")
        
                if not n or _is_blank_employer_value(n):
                    continue
        
                n = re.sub(r"(?i)\bEmail\b\s*[:\-]?\s*$", "", n).strip()
                n = re.sub(r"(?i)\bPhone\b\s*#?\s*[:\-]?\s*$", "", n).strip()
                n = re.sub(r"(?i)\bPh\.?\b\s*#?\s*[:\-]?\s*$", "", n).strip()
        
                if not n or _is_blank_employer_value(n):
                    continue
        
                if _looks_like_company_only(n):
                    continue
        
                if n.lower() not in [x.lower() for x in final_names]:
                    final_names.append(n)
        
            final_emails = []
        
            for e in emails:
                e = e.strip(" .;,)")
                if e and e.lower() not in [x.lower() for x in final_emails]:
                    final_emails.append(e)
        
            final_phones = []
        
            for p in phones:
                p = p.strip()
                if p and p not in final_phones:
                    final_phones.append(p)
        
            # Fallback:
            # If email exists but no reliable name, return Unknown.
            if use_unknown_for_email_only and not final_names and final_emails:
                final_names = ["Unknown"]
        
            employerContactName = " | ".join(final_names)
            employerContactEmail = " | ".join(final_emails)
            employerContactPhone = " | ".join(final_phones)
        
            return employerContactName, employerContactEmail, employerContactPhone


        def extract_employer_contact_name(special_text: str):
            """
            Return employer contact name from Special Instructions (if any).
            Looks at several label patterns and falls back to next-line values.
            """
            if not special_text:
                return None

            lines_si = [ln.strip() for ln in special_text.splitlines() if ln.strip()]
            employer_contact = None
            contact_name = None

            for i, ln_si in enumerate(lines_si):
                # Employer name and contact:
                m = re.match(r"(?i)^Employer\s+name\s+and\s+contact\s*:\s*(.*)$", ln_si)
                if m:
                    employer_contact = _clean_contact_fragment(m.group(1))
                    if employer_contact:
                        return employer_contact
                    continue

                # Employer Contact & Phone:
                m = re.match(r"(?i)^Employer\s+Contact\s*&\s*Phone\s*:\s*(.*)$", ln_si)
                if m:
                    employer_contact = _clean_contact_fragment(m.group(1))
                    if employer_contact:
                        return employer_contact
                    continue

                # Employer Contact & Phone: (value on next line)
                m = re.match(r"(?i)^Employer\s+Contact\s*&\s*Phone\s*:$", ln_si)
                if m and i + 1 < len(lines_si):
                    employer_contact = _clean_contact_fragment(lines_si[i + 1])
                    if employer_contact:
                        return employer_contact
                    continue

                # Contact name:
                m = re.match(r"(?i)^Contact\s+name\s*:\s*(.*)$", ln_si)
                if m:
                    contact_name = _clean_contact_fragment(m.group(1))
                    continue

            if contact_name:
                return contact_name

            # Last-chance: label with value on next line
            for i, ln_si in enumerate(lines_si):
                if re.match(r"(?i)^Employer\s+name\s+and\s+contact\s*:\s*$", ln_si) and i + 1 < len(lines_si):
                    cand = _clean_contact_fragment(lines_si[i + 1])
                    if cand:
                        return cand

            return None

        # employerContactName = extract_employer_contact_name(specialInstructions) or ""
        employerContactName, employerContactEmail, employerContactPhone = extract_employer_contact_details(specialInstructions,use_unknown_for_email_only=True)

        #Search Keyword Medical Only
        if "medical only" in specialInstructions.lower():
            CTMedOnly = True

        #Search Keyword speaking/Language Spiken
        LANGUAGE_SYNONYMS = {
            "spanish": "Spanish",
            "espanol": "Spanish",
            "english": "English",
            "mandarin": "Mandarin",
            "cantonese": "Cantonese",
            "vietnamese": "Vietnamese",
            "portuguese": "Portuguese",
            "french": "French",
            "arabic": "Arabic",
            "russian": "Russian",
            "korean": "Korean",
            "tagalog": "Tagalog",
            "filipino": "Tagalog",
            "hindi": "Hindi",
            "urdu": "Urdu",
            "punjabi": "Punjabi",
            "polish": "Polish",
            "german": "German",
            "italian": "Italian",
            # add more as needed
        }
        
        
        # def extract_language_from_special_instructions(text: str) -> str:
            
        #     if not text:
        #         return "English"  # default
        
        #     lower = text.lower()
        #     languages_found = []
        
        #     patterns = [
        #         r"\bspeaks?\s+([A-Za-z\- ]+)",           # "IW speaks Spanish only"
        #         r"\b([A-Za-z\- ]+)\s+speaking\b",       # "Spanish speaking"
        #         r"\b([A-Za-z\- ]+)\s+speaker\b",        # "Spanish speaker"
        #         r"interpreter\s*[:\-]?\s*([A-Za-z\- ]+)",  # "interpreter: Spanish"
        #         r"\b([A-Za-z\- ]+)\s+interpreter\b",    # "Spanish interpreter"
        #     ]
        
        #     def add_lang_candidate(raw_phrase: str):
        #         phrase = raw_phrase.strip().lower()
        #         phrase = phrase.replace("only", "").replace("fluently", "").replace("well", "")
        #         phrase = phrase.replace("preferred", "").replace("required", "")
        #         phrase = phrase.replace("language", "")
        #         phrase = phrase.strip(" ,;.")
        
        #         if not phrase:
        #             return
        
        #         tokens = [t for t in re.split(r"[\/,&]| and ", phrase) if t.strip()]
        #         for t in tokens:
        #             w = t.strip().split()[0]  # first word is usually the language name
        #             if not w:
        #                 continue
        #             canon = LANGUAGE_SYNONYMS.get(w.lower(), w.capitalize())
        #             if canon not in languages_found:
        #                 languages_found.append(canon)
        
        #     # Pattern-based extraction
        #     for pat in patterns:
        #         for m in re.finditer(pat, text, re.IGNORECASE):
        #             add_lang_candidate(m.group(1))
        
        #     # Keyword-based backup (plain mentions)
        #     for key, canon in LANGUAGE_SYNONYMS.items():
        #         if key in lower and canon not in languages_found:
        #             if f"not {key}" in lower or f"no {key}" in lower:
        #                 continue
        #             languages_found.append(canon)
        
        #     # Nothing explicit found
        #     if not languages_found:
        #         if "bilingual" in lower:
        #             # still honor the request, but default base language to English
        #             return "English"
        #         return "English"
        
        #     return ", ".join(languages_found)
        def extract_language_from_special_instructions(text: str) -> str:
            if not text or not text.strip():
                return "English"

            lower = text.lower()
            languages_found = []

            patterns = [
            r"\bspeaks?\s+([A-Za-z\-\/,& ]+?)(?=\s+(?:only|fluently|well|required|preferred)\b|[.;,\n]|$)",
            r"\b([A-Za-z\-]+)\s+speaking\b",
            r"\b([A-Za-z\-]+)\s+speaker\b",
            r"interpreter\s*[:\-]?\s*([A-Za-z\-\/,& ]+?)(?=[.;,\n]|$)",
            r"\b([A-Za-z\-]+)\s+interpreter\b",
            r"\bassign\s+([A-Za-z\-]+)\s+speaking\b",
            r"\bneed\s+([A-Za-z\-]+)\s+speaking\b",
            r"\brequest(?:ed)?\s+([A-Za-z\-]+)\s+speaking\b",
            r"\b([A-Za-z\-]+)\s+only\b",
            r"\blanguage\s*[:\-]?\s*([A-Za-z\-\/,& ]+?)(?=[.;,\n]|$)",
            ]

            def add_lang_candidate(raw_phrase: str):
                phrase = raw_phrase.strip().lower()

            # remove helper words
                for junk in ["only", "fluently", "well", "preferred", "required", "language"]:
                    phrase = phrase.replace(junk, " ")

                    phrase = re.sub(r"\s+", " ", phrase).strip(" ,;:.-")

                    if not phrase:
                        return

                    tokens = [t.strip() for t in re.split(r"/|,|&|\band\b", phrase, flags=re.IGNORECASE) if t.strip()]

                    for token in tokens:
                        words = re.findall(r"[A-Za-z\-]+", token.lower())
                        for word in words:
                            if word in LANGUAGE_SYNONYMS:
                                canon = LANGUAGE_SYNONYMS[word]
                                if canon not in languages_found:
                                    languages_found.append(canon)

            # pattern-based extraction
            for pat in patterns:
                for m in re.finditer(pat, text, re.IGNORECASE):
                    add_lang_candidate(m.group(1))

            # keyword backup
            for key, canon in LANGUAGE_SYNONYMS.items():
                if re.search(rf"\b{re.escape(key)}\b", lower):
                    if re.search(rf"\b(?:not|no)\s+{re.escape(key)}\b", lower):
                        continue
                    if canon not in languages_found:
                        languages_found.append(canon)

            # fallback
            if not languages_found:
                if "bilingual" in lower:
                    return "English"
                return "English"

            return ", ".join(languages_found)

        
        languageSpoken = extract_language_from_special_instructions(specialInstructions)

        # ---------- NCM contact name from Special Instructions ----------
        def _clean_ncm_name(raw: str) -> str | None:
            if not raw:
                return None
            s = raw.strip(" ,;:-")

            # drop common credentials / titles
            s = re.sub(
                r"\b(RN|BSN|MSN|LPN|LVN|COHN/CM|COHN|CM|CCM|CRRN|FNP|NP|PA-?C)\b\.?,?",
                "",
                s,
                flags=re.I,
            )
            # drop explicit NCM tokens
            s = re.sub(r"\b(NCM|tele\s*NCM)\b", "", s, flags=re.I)

            # remove leftover parentheses text like "(tele NCM)" if any
            s = re.sub(r"\([^)]*\)", "", s)

            # collapse whitespace
            s = " ".join(s.split())
            return s or None

        def extract_ncm_name_from_special_instructions(text: str) -> str:
            if not text:
                return ""

            name = ""

            pat_contact = re.compile(
                r"(?is)contact\s+NCM[:,]?\s*([A-Za-z .,'/-]+?)\s*"
                r"(?:at\b|@|\bemail\b|\bfor\b|\bbefore\b|\bafter\b|$)"
            )
            m = pat_contact.search(text)
            if m:
                cleaned = _clean_ncm_name(m.group(1))
                if cleaned:
                    name = cleaned

            if not name:
                pat_ncm = re.compile(
                    r"(?is)\bNCM[:,]?\s*([A-Za-z .,'/-]+?)\s*"
                    r"(?:at\b|@|\bemail\b|\bfor\b|\bbefore\b|\bafter\b|$)"
                )
                m2 = pat_ncm.search(text)
                if m2:
                    cleaned = _clean_ncm_name(m2.group(1))
                    if cleaned:
                        name = cleaned
            if not name:
                m3 = re.search(
                    r"(?im)^.*NCM.*?([\w\.-]+@[\w\.-]+\.\w+).*$",
                    text,
                )
                if m3:
                    email = m3.group(1)
                    local = email.split("@", 1)[0]
                    parts = [p for p in re.split(r"[._]", local) if p]
                    if len(parts) >= 2:
                        name = f"{parts[0].capitalize()} {parts[1].capitalize()}"
                    elif parts:
                        name = parts[0].capitalize()

            return name or ""

        ncmContactName = extract_ncm_name_from_special_instructions(specialInstructions)


        #Search Keyword MHAA GAA
        if "-MHAA" in specialInstructions.lower():
            LanguageSpeak = True
            print("MHAA!")
    
        # ---------- Referral Number ----------
        referralNumber = ""
        for ln in lines:
            if "Referral Number:" in ln:
                m = re.search(r"Referral Number:\s*([0-9]+)", ln)
                if m:
                    referralNumber = m.group(1).strip()
                    break
    
        # ---------- Provider extraction ----------
               # ---------- Provider extraction ----------
        providerName = ""
        providerAddr = ""
        providerCity = ""
        providerState = ""
        providerZip = ""
        providerPhone = ""

        # 1) RN Visit address block sequence after "Street Address of One-Time RN Visit:"
        seenRNStreet = False
        for ln in lines:
            if "Street Address of One-Time RN Visit:" in ln:
                providerAddr = after_colon(ln)
                seenRNStreet = True
                continue
            if seenRNStreet and "City:" in ln and not providerCity:
                providerCity = after_colon(ln)
                continue
            if seenRNStreet:
                m = re.match(r"^\s*State:\s*([A-Z]{2})", ln, re.I)
                if m and not providerState:
                    providerState = m.group(1).strip()
                    continue
            if seenRNStreet and "Zip Code:" in ln and not providerZip:
                rawZip = after_colon(ln)
                m = re.match(r"(\d{5})", rawZip)
                providerZip = m.group(1) if m else rawZip.strip()
                seenRNStreet = False
                continue

        # 2) Provider / Facility Name
        # for ln in lines:
        #     if "Provider / Facility Name:" in ln:
        #         tmp = after_colon(ln)
        #         if tmp:
        #             providerName = tmp
        #         break
        for i, ln in enumerate(lines):
            if "Provider / Facility Name:" in ln:
                parts = []

                first = after_colon(ln)
                if first:
                    first = first.strip()
                    if first:
                        parts.append(first)

                # Keep reading the next lines until we hit "Extension"
                j = i + 1
                while j < len(lines):
                    nxt = lines[j].strip()
                    if not nxt:
                        j += 1
                        continue

                    # STOP condition (only)
                    if "Extension" in nxt:
                        break

                    parts.append(nxt)
                    j += 1

                if parts:
                    providerName = " / ".join(parts).strip()

                break


        # 3) Capture Special Instructions again as list + flat text
        si_lines = []
        si_text = ""
        capture = False
        for ln in lines:
            if "Special Instructions" in ln:
                m0 = re.search(r"(?i)\bSpecial\s*Instructions\s*:\s*(.*)", ln)
                if m0 and trim(m0.group(1)):
                    si_lines.append(trim(m0.group(1)))
                    si_text += " " + trim(m0.group(1))
                capture = True
                continue
            if capture:
                if re.search(
                    r"(?i)^\s*(Referrer Name:|©|Page \d+ of \d+|Vendor Name:)\s*$",
                    ln,
                ):
                    break
                l = ln.strip()
                if l:
                    si_lines.append(l)
                    si_text += " " + l
        ##
        # nextApptDate = ""
        nextApptTime = ""

        # 3b) NEW: try to fill provider address from Special Instructions dynamically
        providerAddr, providerCity, providerState, providerZip = fill_provider_address_from_special(
            specialInstructions,
            providerAddr,
            providerCity,
            providerState,
            providerZip,
            parse_address=parse_address,
            providerFirst=providerFirst,
            providerLast=providerLast,
            facilityName=providerName,
        )

        # 4) Fallbacks from Special Instructions: provider name / phone
        if not providerName:
            m = re.search(
                r"(?i)\bProvider\s*(?:Name)?\s*:\s*([^|,\r\n]+)", si_text
            )
            if m:
                providerName = trim(m.group(1))

        if not providerPhone:
            empStopper = re.compile(
                r"(?i)\bemployer\s*(?:[-_/ ]*\s*name)?"
                r"(?:\s*(?:and\s*contact)?)?"
                r"(?:\s*[-_/ ]*\s*address)?\s*[:\-]"
            )

            def SearchPrvPhoneBeforeEmployer(si_text: str) -> str:
                si_text = si_text or ""
                m = empStopper.search(si_text)
                return si_text[: m.start()] if m else si_text

            scopeSearch = SearchPrvPhoneBeforeEmployer(si_text)
            # print(scopeSearch)
            m = re.search(
                r"(?i)\bPhone\s*(?:Number|#)?\s*:\s*([()+\-\s\d]{7,})",
                scopeSearch,
            )
            if m:
                providerPhone = trim(m.group(1))
            else:
                # m = re.search(
                #     r"(?:\+?1[\s\-\.]?)?\(?\d{3}\)?[\s\-\.]?\d{3}[\s\-\.]?\d{4}",
                #     scopeSearch,
                # )
                m = re.search(r"(?i)\b(?:Organization|Provider|Facility)?\s*Phone\s*(?:Number|#)?\s*:\s*((?:\+?\d[\s\-\.\)]?){7,})",scopeSearch)
                if m:
                    providerPhone = trim(m.group(0))

        # sanitize providerState (remove stray "Zip")
        if providerState:
            providerState = trim(providerState.replace("Zip", ""))

        # 5) Address patterns over si_lines if still missing
        if not (providerAddr and providerCity and providerState and providerZip):
            # adjacent two-line pattern
            for a, b in zip(si_lines, si_lines[1:]):
                okA = re.match(r"^\s*\d{1,6}\s+.+$", a)
                mC = re.match(
                    r"(?i)^\s*([A-Za-z .'\-/&]+?)\s*,\s*([A-Z]{2})\s*(\d{5})(?:-\d{4})?\s*$",
                    b,
                )
                if okA and mC:
                    city_, st_, z_ = mC.group(1), mC.group(2), mC.group(3)
                    if not providerAddr:
                        providerAddr = a.strip()
                    if not providerCity:
                        providerCity = city_.strip()
                    if not providerState:
                        providerState = st_.strip()
                    if not providerZip:
                        providerZip = z_.strip()
                    break



        # ---------- Accident / Injury Descriptions ----------
        AccidentDesc = ""
        InjuryDesc   = ""
    
        def grab_wrapped(lines, anchor, stop_rx):
            out = ""
            for i, ln in enumerate(lines):
                if anchor in ln:
                    m = re.search(re.escape(anchor) + r"\s*(.+)", ln)
                    if m:
                        out = trim(m.group(1))
                    if not out:
                        k = i + 1
                        while k < len(lines):
                            nxt = lines[k].strip()
                            if not nxt:
                                k += 1
                                continue
                            if re.search(stop_rx, nxt, re.I):
                                break
                            out = nxt
                            break
                    break
            return out

        def grab_wrapped_criteria(lines, SearchCri, EndCri):
            SearchCri_lc = SearchCri.lower()
            textPhrase = []
            newLine = '\n'
            contGet = False
            for ln in lines:
                if SearchCri_lc in ln.lower() or (contGet and EndCri.lower() not in ln.lower()):
                    parts = ln.split(":", 1)
                    if len(parts) >=2:
                        textPhrase.append(parts[1].strip()) 
                    else:  
                        textPhrase.append(parts[0].strip())
                    contGet = True 
                elif EndCri.lower() in ln.lower():
                    return newLine.join(textPhrase).strip()
                         
            return ""

        # AccidentDesc = grab_wrapped(lines, "Accident Description:",
        #     r"^(Injury Description:|Diagnosis Code:|Case Manager Information|Provider Information|Referral Instructions|Vendor Information|Attorney Information|Claim Information|Page \d+ of \d+)")
        # InjuryDesc = grab_wrapped(lines, "Injury Description:",
        #     r"^(Diagnosis Code:|Case Manager Information|Provider Information|Referral Instructions|Vendor Information|Attorney Information|Claim Information|Page \d+ of \d+)")

        AccidentDesc = grab_wrapped_criteria(lines, "Accident Description:","Injury Description:")
        InjuryDesc   = grab_wrapped_criteria(lines, "Injury Description:","Diagnosis Code:")
        # ---------- Next Appt Date & Time ----------
        # nextApptDate = ""
        # # migz
        nextApptTime = ""
    
        # labeled fields
        for ln in lines:
            if any(lbl in ln for lbl in [
                "Next Provider Appointment Date:",
                "Next Provider Appt. Date:",
                "Next Appointment Date:",
                "Next Appt Date:",
                "Appt. Date:"
            ]):
                tail = re.search(r":\s*(.*)", ln)
                s = tail.group(1) if tail else ""
                if not nextApptDate:
                    nextApptDate = first_date(s)
                if not nextApptTime:
                    nextApptTime = first_time(s)
                continue
    
            if any(lbl in ln for lbl in [
                "Next Provider Appointment Time:",
                "Next Provider Appt. Time:",
                "Next Appointment Time:",
                "Next Appt Time:",
                "Appt. Time:"
            ]):
                tail = re.search(r":\s*(.*)", ln)
                s = tail.group(1) if tail else ""
                if not nextApptTime:
                    nextApptTime = first_time(s)
                continue
    
            if ("Appt. Date" in ln and "Time" in ln) and (not nextApptDate or not nextApptTime):
                mBoth = re.search(r".*:\s*(.*)$", ln)
                both = mBoth.group(1) if mBoth else ""
                nextApptDate = nextApptDate or first_date(both)
                nextApptTime = nextApptTime or first_time(both)
    
        # special instructions fallback
        # if not nextApptDate:
        #     nextApptDate = first_date(specialInstructions)
        if not nextApptTime:
            nextApptTime = first_time(specialInstructions)
    
        if nextApptDate.strip().lower() == "01/01/2099":
            nextApptDate = ""
        nextApptTime = normalize_spaces(nextApptTime)

          
        # ---------- Provider first/last from providerName ----------
        providerFirst = ""
        providerLast = ""
 
        if not providerName:
            providerName = ""
 
        raw = providerName.strip()
        low = raw.lower()
 
        # Values that really mean "no provider"
        NONE_PATTERNS = {
            "na",
            "n/a",
            "n.a",
            "none",
            "no provider",
            "no provider at this time",
            "tbd",
            "unknown",
        }
 
        if raw and low.strip() not in NONE_PATTERNS:
            FACILITY_KEYWORDS = {
                "clinic", "center", "medical", "medicine", "occupational", "urgent",
                "care", "care now", "ortho", "orthopedic", "orthopaedic", "spine",
                "wound", "hospital", "rehab", "rehabilitation", "therapy",
                "concentra", "express care", "rothman", "cmc", "avalon", "imaging","umc",
                "radiology", "healthcare", "health", "group", "associates","orthoxpress",
                "hand", "failure", "foot", "ankle","industrial"
            }

            PERSON_CREDENTIALS = r"(Dr\.?|Doctor|M\.?D\.?|D\.?O\.?|PA-?C|NP|FNP|APRN|RN|DPM|DC|Surgeon)"
 
            #   e.g. "Donna Randolph - Concentra"  -> "Donna Randolph"
            #        "Craig Wiseman/Golden State" -> "Craig Wiseman"
            core = re.split(r"\s-\s|/|@", raw)[0]
 
            # Drop anything in parentheses: "(Hand Surgeon)" etc.
            core = re.sub(r"\([^)]*\)", "", core)
 
            core_clean = re.sub(
                r"\b(Dr\.?|Doctor|M\.?D\.?|D\.?O\.?|PA-?C|PA\b|DPM|NP|FNP|APRN|RN|DC)\b",
                "",
                core,
                flags=re.I,
            ).strip()
 
            core_tokens = re.findall(r"[A-Za-z][A-Za-z'\-]+", core_clean)

            # --- "Facility, First Last, MD" pattern, use the part after the first comma ---
            if "," in raw and any(k in low for k in FACILITY_KEYWORDS):
                # take text after first comma, e.g. "THOMAS DELIZIO, MD"
                person_part = raw.split(",", 1)[1]
                # remove credentials/punctuation
                person_part = re.sub(PERSON_CREDENTIALS, "", person_part, flags=re.I)
                person_part = re.sub(r"[^\w'\- ]+", " ", person_part)
                person_part = re.sub(r"\s+", " ", person_part).strip()

                ptoks = re.findall(r"[A-Za-z][A-Za-z'\-]+", person_part)
                if len(ptoks) >= 2:
                    providerFirst = " ".join(ptoks[:-1])
                    providerLast = ptoks[-1]
 
            looks_like_person = False
 
            # Rule 1: any personal credential ("Dr", MD, PA-C, etc.)
            if re.search(PERSON_CREDENTIALS, raw, re.I):
                looks_like_person = True
 
            # Rule 2: "Last, First" pattern
            if re.match(r"^[A-Z][A-Za-z']+,\s*[A-Z][A-Za-z']+", raw):
                looks_like_person = True
 
            # Rule 3: 2–3 "name-like" tokens in the core (and not obviously facility words)
            if len(core_tokens) in (2, 3) and not any(
                t.lower() in FACILITY_KEYWORDS for t in core_tokens
            ):
                looks_like_person = True
 
            # Facility detection (any facility keyword anywhere in the raw string)
            looks_like_facility = any(k in low for k in FACILITY_KEYWORDS)
 
            # Final classification
            if looks_like_person and not looks_like_facility:
                kind = "person"
            elif looks_like_person and looks_like_facility:
                # If we have credentials or a decent name core, treat as a person
                if re.search(PERSON_CREDENTIALS, raw, re.I) or len(core_tokens) >= 2:
                    kind = "person"
                else:
                    kind = "facility"
            else:
                kind = "facility"
 
            # ----- Extract first/last if we decided it's a PERSON -----
            if kind == "person":
                name_tokens = re.findall(r"[A-Za-z][A-Za-z'\-]+", core_clean)
 
                # Tokens we never want as "names"
                STOPWORDS = {"dr", "doctor", "surgeon"}
 
                # Remove facility-ish words and titles from the final name tokens
                name_tokens = [
                    t
                    for t in name_tokens
                    if t.lower() not in FACILITY_KEYWORDS
                    and t.lower() not in STOPWORDS
                ]
                if not providerFirst and not providerLast:
                    if len(name_tokens) == 1:
                        # Single-word: treat as last name
                        providerFirst = ""
                        providerLast = name_tokens[0]
                    elif len(name_tokens) >= 2:
                        # Multi-word: everything except last = first; last token = last name
                        providerFirst = " ".join(name_tokens[:-1])
                        providerLast = name_tokens[-1]



                # --------- Inury Cause, Type and Body Part ---------- #
        def _blend(accident_desc: str, injury_desc: str, special_instruction: str) -> str:
            """Join all sources safely and lower-case once."""
            return f"{accident_desc or ''} {injury_desc or ''} {special_instruction or ''}".lower()
        
        # ----------------------------
        # Body Part
        # ----------------------------
        def resolve_body_part(accident_desc: str, injury_desc: str, special_instruction: str) -> str:
            text = _blend(accident_desc, injury_desc, special_instruction)
        
            rules = [
                (r"\bskull\b|head\s*injury|concussion",                           "Head"),
                (r"\bhead\b",                                                     "Multiple Head Injury"),
                (r"\bsoft\s*tissue\b.*head|\bhead\b.*soft\s*tissue",              "Soft Tissue - Head"),
                (r"\bface|facial\s*bones",                                        "Facial Bones"),
                (r"\beye\b|eyes\b|ocular",                                        "Eye(s)"),
                (r"\bear\b|ears\b|otitis|tympanic",                               "Ear(s)"),
                (r"\btooth|teeth|dental",                                         "Teeth"),
                (r"\bmouth|oral\b",                                               "Mouth"),
                (r"\bnose|nasal\b",                                               "Nose"),
                (r"\bneck\b|cervical",                                            "Neck"),
                (r"\bupper\s*back\b|thoracic\b",                                  "Upper Back Area (Thoracic Area)"),
                (r"\blow\s*back\b|lumbar\b",                                      "Low Back Area (Lumbar and Lumbo-Sacral)"),
                (r"\bsacrum\b|coccyx",                                            "Sacrum and Coccyx"),
                (r"\bvertebrae\b.*(lumbar|sacral)|\blumbar\b.*vertebrae|\bsacral\b.*vertebrae",
                                                                                "Lumbar &/or Sacral Vertebrae (Vertebra NOC Trunk)"),
                (r"\bspinal\s*cord\b.*(neck|cervical)",                           "Spinal Cord - Neck"),
                (r"\bspinal\s*cord\b",                                            "Spinal Cord - Trunk"),
                (r"\bchest\b|ribs\b|sternum|soft\s*tissue.*chest",                "Chest  (Ribs, Sternum, Soft Tissue)"),
                (r"\babdomen\b|groin",                                            "Abdomen (including groin)"),
                (r"\bpelvis\b",                                                   "Pelvis"),
                (r"\binternal\s*organs\b",                                        "Internal Organs"),
                (r"\bbody\s*systems\b|multiple\s*body\s*systems",                 "Body Systems and Multiple Body Systems"),
                (r"\bmultiple\s*(upper|lower|head|trunk|extremities)\b",          "Multiple Body Parts"),
                
                (r"\bshoulder\b|rotator\s*cuff",                                  "Shoulder(s)"),
                (r"\bupper\s*arm\b",                                              "Upper Arm"),
                (r"\blower\s*arm\b|forearm",                                      "Lower Arm"),
                (r"\belbow\b",                                                    "Elbow"),
                (r"\bwrist\b",                                                    "Wrist"),
                (r"\bhand\b",                                                     "Hand"),
                (r"\bfinger",                                                     "Finger(s)"),
                (r"\bthumb\b",                                                    "Thumb"),
                (r"\bhip\b",                                                      "Hip"),
                (r"\bupper\s*leg\b|thigh",                                        "Upper Leg"),
                (r"\blower\s*leg\b|shin|calf",                                    "Lower Leg"),
                (r"\bknee\b|meniscus|patella",                                    "Knee"),
                (r"\bankle\b|achilles",                                           "Ankle"),
                (r"\bfoot\b|plantar|metatars",                                    "Foot"),
                (r"\btoe",                                                        "Toe(s)"),
                (r"\bbuttock(s)?\b",                                              "Buttocks"),
                (r"\bdisc\b.*neck",                                               "Disc - Neck"),
                (r"\bdisc\b.*(trunk|back)",                                       "Disc - Trunk"),
                (r"\blarynx\b|trachea\b",                                         "Larynx"),
                (r"\bno\s*physical\s*injury\b",                                   "No Physical Injury  (Mental Disorder)"),
            ]
        
            found = {val for pat, val in rules if re.search(pat, text, re.I)}
            if len(found) > 1:
                return "Multiple Body Parts"
            return next(iter(found)) if found else "Insufficient Info to Properly Identify - Unclassified"

        def resolve_injury_type(accident_desc: str,
                        injury_desc: str,
                        special_instruction: str) -> str:
            """
            Determine Injury Type using:
            1) AccidentDesc + InjuryDesc
            2) SpecialInstructions (only if #1 finds nothing)
            """

            core_text = f"{accident_desc or ''} {injury_desc or ''}".lower()
            si_text   = (special_instruction or "").lower()

            # Mapping to Injury Type dropdown values
            rules = [
                (r"\bamputat",                                   "Amputation"),
                (r"\bfractur",                                   "Fracture"),
                (r"\bdislocat",                                  "Dislocation"),
                (r"\blacerat|\bcut\b|\bincis",                   "Laceration"),
                (r"\bcontusion|\bbruise",                        "Contusion"),
                (r"\bsprain\b|\btear\b|\brotator\s*cuff\s*tear", "Sprain or Tear"),
                (r"\bstrain\b|\bpulled\s+muscle",                "Strain"),
                (r"\bconcuss|\btbi\b",                           "Concussion"),
                (r"\bburn\b|thermal|scald",                      "Burn"),
                (r"\bpuncture|\bstab\b",                         "Puncture"),
                (r"\bcrush|\bcrushing",                          "Crushing"),
                (r"\bforeign\s*body|\bsplinter|\bshard",         "Foreign Body"),
                (r"\binfect|\bcellulitis|\bsepsis",              "Infection"),
                (r"\binflamm|\btendinitis|\btenosynov|\bbursitis","Inflammation"),
                (r"\bhernia\b",                                  "Hernia"),
                (r"\bdermatitis|\brash\b|\beczema",              "Dermatitis"),
                (r"\bpoison|\bintoxicat|\btoxic\b",              "Poisoning - General (Not OD or Cumulative Injury)"),
                (r"\bchemical\s+exposure|\bfume|\bgas|\bvapor",  "Respiratory Disorders (Gas, Fumes, Chems)"),
                (r"\bradiation",                                 "Radiation"),
                (r"\bheat\s+prostration|\bheat\s+exhaustion|\bheat\s+stroke", "Heat Prostration"),
                (r"\bfreez|\bfrostbite",                         "Freezing"),
                (r"\belectric\s+shock|\belectrocution",          "Electric Shock"),
                (r"\bsyncope|\bfaint",                           "Syncope (Fainting)"),
                (r"\bcarpal\s+tunnel",                           "Carpal Tunnel Syndrome"),
                (r"\bhearing\s+loss|\bimpairment\s*\(traumatic only\)", "Hearing Loss or Impairment (Traumatic Only)"),
                (r"\bmyocardial\s+infarction|\bheart\s+attack",  "Myocardial Infarction (Heart Attack)"),
                (r"\bblack\s*lung",                              "Black Lung"),
                (r"\basbest",                                    "Asbestosis"),
                (r"\bpandemic|\bcovid",                          "Pandemic"),
                (r"\bmultiple\s+injur|\bpolytrauma",             "Multiple Physical Injuries Only"),
                (r"\bno\s*physical\s*injury\b",                  "No Physical Injury"),
                (r"\bvdt|\brepetitive\s+motion|\bcumulative\s+trauma",
                                                                "All Other Cumulative Injury, NOC"),
            ]

            def classify(text: str) -> tuple[bool, set]:
                # Any clear mental / psych flags?
                psych_found = bool(
                    re.search(
                        r"\b(mental\s+stress|mental\s+disorder|ptsd|anxiety|"
                        r"depress(ion|ed)?|panic|psych(ological)?|psychiatr(ic|y)?)\b",
                        text,
                        re.I,
                    )
                )

                phys: set[str] = set()
                for pat, val in rules:
                    if re.search(pat, text, re.I):
                        # If 'No Physical Injury' is present, it overrides others
                        if val == "No Physical Injury":
                            phys = {val}
                            break
                        phys.add(val)

                return psych_found, phys

            def finalize(psych_found: bool, phys: set) -> str:
                if psych_found and phys:
                    return "Multiple Injuries (Including Both Phys and Psych)"
                if psych_found and not phys:
                    return "Mental Stress"
                if len(phys) > 1:
                    return "Multiple Physical Injuries Only"
                return next(iter(phys)) if phys else "All Other Specific Injuries, NOC"

            # 1) First pass: Accident + Injury description only
            psych_core, phys_core = classify(core_text)
            if psych_core or phys_core:
                return finalize(psych_core, phys_core)

            # 2) Second pass: Special Instructions only
            psych_si, phys_si = classify(si_text)
            return finalize(psych_si, phys_si)
        

        # ----------------------------
        # Cause (optional, now also includes SI)
        # ---------------------------- migz changes 11pm
        # def resolve_cause(accident_desc: str, injury_desc: str, special_instruction: str) -> str:
        #     text = _blend(accident_desc, injury_desc, special_instruction)
        
        #     rules = [
        #         (r"\bslip|trip|fell|fall\b",                     "Slip/Trip/Fall"),
        #         (r"\bstruck\s+by|\bstruck\s+against|\bhit by|\bimpact", "Struck By/Against"),
        #         (r"\bcaught\s+in|\bcaught\s+between|\bpinch|entrap",    "Caught In/Between"),
        #         (r"\bmotor\s*vehicle|car |truck|forklift|auto\b",        "Motor Vehicle"),
        #         (r"\blifting|lifted",                                   "Lifting"),
        #         (r"\bpushing|pulling",                                   "Pushing or Pulling"),
        #         # (r"\b",                                           ""),
        #         # (r"\bcarrying",                                          "Carrying"),
        #         (r"\brepetitive|cumulative|vdt",                         "Repetitive Motion"),
        #         (r"\bchemical|fume|gas|vapor|inhalation|exposure",       "Exposure to Chemicals"),
        #         (r"\bburn|thermal|hot|scald",                            "Heat/Hot Object"),
        #         (r"\belectric|shock|electrocute",                        "Electrical"),
        #         (r"\banimal|dog bite|insect|bee |wasp",                   "Animal/Insect"),
        #         (r"\bneedle|sharps|puncture",                            "Needle/Sharps"),
        #         (r"\bassault|fight|violence",                            "Assault/Violence"),
        #     ]
        #     for pat, val in rules:
        #         if re.search(pat, text, re.I):
        #             return val
        #     return "'Other - Miscellaneous, NOC"

        def resolve_cause(accident_desc: str,
                        injury_desc: str,
                        special_instruction: str) -> str:
            """
            Determine Cause of Injury using:
            1) AccidentDesc + InjuryDesc
            2) SpecialInstructions (only if #1 finds nothing)
            """

            def _norm(s: str) -> str:
                return (s or "").lower()

            core_text = _norm(f"{accident_desc or ''} {injury_desc or ''}")
            si_text   = _norm(special_instruction or "")

            # ---------- FALL / SLIP / TRIP SPECIAL HANDLING ----------
            def classify_fall(text: str) -> str | None:
                if not re.search(r"\b(slip(ped)?|trip(ped)?|fell|fall|falling)\b", text):
                    return None

                # Did not fall
                if re.search(r"\b(no\s+fall|did\s+not\s+fall)\b", text):
                    return "Slip or Trip, Did Not Fall"

                # Ice or snow
                if re.search(r"\b(ice|icy|snow|snowy)\b", text):
                    return "Fall/Slip/Trip Injury - On Ice or Snow"

                # Stairs
                if re.search(r"\bstair(s)?\b|\bstep(s)?\b", text):
                    return "Fall/Slip/Trip Injury - On Stairs"

                # Ladder / scaffold
                if re.search(r"\bladder\b|\bscaffold(ing)?\b", text):
                    return "Fall/Slip/Trip Injury - From Ladder or Scaffolding"

                # Liquid / grease spills
                if re.search(r"\b(spill(ed|ing)?|spillage|grease|oil|liquid|wet floor)\b", text):
                    return "Fall/Slip/Trip Injury - From Liquid/Grease Spills"

                # Into openings
                if re.search(r"\bhole\b|\bpit\b|\btrench\b|\bmanhole\b|\bopening\b|\bshaft\b", text):
                    return "Fall/Slip/Trip Injury - Into Openings"

                # From different level (generic height)
                if re.search(
                    r"\b(from|off)\b.*\b(level|roof|platform|dock|loading\s+dock|second floor|balcony)\b",
                    text,
                ):
                    return "Fall/Slip/Trip Injury - From Different Level"

                # Otherwise assume same level
                return "Fall/Slip/Trip Injury - On Same Level"

            # ---------- MAIN CLASSIFIER ----------
            def classify_cause(text: str) -> str | None:
                if not text.strip():
                    return None

                # 1) Falls
                fall_code = classify_fall(text)
                if fall_code:
                    return fall_code

                # 2) Specific causes (aligned to dropdown values)
                patterns: list[tuple[str, str]] = [
                    # Abnormal air pressure
                    (r"\babnormal\s+air\s+pressure\b|\bdecompression\b|\bbarotrauma\b",
                        "Abnormal Air Pressure"),

                    # Animal / insect  (word-boundaries to avoid 'been' → 'bee')
                    (r"\banimal(s)?\b|\bdog\b|\b(dog\s*bite)\b|"
                        r"\b(cat|insect|bee|wasp|hornet|spider|mosquito\w*|tick|ant|rodent|rat|mouse)\b",
                        "Animal or Insect"),

                    # Broken glass
                    (r"\bbroken\s+glass\b|\bglass\s+cut\b|\bcut\b.*\bglass\b|\bglass\b.*\bcut\b",
                        "Broken Glass"),

                    # Caught in/under/between
                    (r"\bcaught\s+in\b|\bcaught\s+under\b|\bcaught\s+between\b|\bpinch(ing)?\b|entrap",
                        "Caught In, Under or Between, NOC"),

                    # Chemicals / dust / fumes
                    (r"\bchemical(s)?\b|\bsolvent(s)?\b|\bbleach\b|\bacid\b|\bcaustic\b",
                        "Chemicals"),
                    (r"\bdust\b|\bfume(s)?\b|\bgas(es)?\b|\bvapor(s)?\b|\binhalation\b",
                        "Dust, Gases, Fumes or Vapors"),

                    # Cold objects / substances
                    (r"\bcold\b.*\bobject\b|\bobject\b.*\bcold\b|\bfrozen\b|\bice\b(?!\s*or snow)",
                        "Cold Objects or Substances"),

                    # Collapsing materials / cave-in
                    (r"\bcollapse(d|s|)\b.*\b(trench|dirt|earth|wall|soil)\b|\bcave-?in\b|\blandslide\b",
                        "Collapsing Materials (Sides of Earth)"),

                    # Contact with / NOC
                    (r"\bcontact\b.*\b(with|by)\b",
                        "Contact With, NOC"),

                    # COVID-19 / pandemic
                    (r"\bcovid\b|\bcoronavirus\b|\bsars\-?cov\-?2\b",
                        "COVID-19"),

                    # Vehicles (crashes, collisions, upset)
                    (r"\bcrash\b.*\b(plane|aircraft|airplane)\b",
                        "Crash of Airplane"),
                    (r"\bcrash\b.*\b(train|rail|locomotive)\b",
                        "Crash of Rail Vehicle"),
                    (r"\bcrash\b.*\b(boat|ship|vessel|barge)\b",
                        "Crash of Water Vehicle"),
                    (r"\b(upset\b|\boverturned\b|\brollover\b)\b.*\b(vehicle|truck|car)\b",
                        "Vehicle Upset"),
                    (r"\b(collid(ed|e|ing)?|sideswipe(d)?)\b.*\b(vehicle|car|truck|forklift)\b",
                        "Collision or Sideswipe with Another Vehicle"),
                    (r"\b(collid(ed|e|ing)?|hit\b|struck\b)\b.*\b(pole|guardrail|barrier|wall|tree|building|door|fixed object)\b",
                        "Collision with a Fixed Object"),
                    (r"\bmotor\s+vehicle\b|\bauto\b|\bcar\b|\btruck\b|\bvan\b",
                        "Motor Vehicle, NOC"),

                    # Cut / puncture / scrape
                    (r"\b(cut|puncture|scrape|laceration)\b",
                        "Cut, Puncture, Scrape, NOC"),
                    (r"\bobject\b.*\b(cut|puncture|scrape)\b|\b(cut|puncture|scrape)\b.*\bobject\b",
                        "Object Handled - Cut/Puncture/Scrape"),

                    # Foreign matter in eye
                    (r"\beye\b.*\b(foreign\s+body|debris|dust|dirt|metal|object|particle)\b",
                        "Foreign Matter (Body) in Eyes(s)"),

                    # Electrical
                    (r"\belectric(al)?\s+shock\b|\belectrocute(d|ion)?\b|\bcontact\b.*\bpower\s+line\b",
                        "Electrical Current"),

                    # Explosion / fire / steam / hot objects
                    (r"\bexplosion\b|\bexploded\b|\bflare\-?back\b|\bblast\b",
                        "Explosion or Flare Back"),
                    (r"\bfire\b|\bflame(s)?\b|\bburned\b|\bburning\b",
                        "Fire or Flame"),
                    (r"\bsteam\b|\bhot\s+fluid(s)?\b|\bboiling\b",
                        "Steam or Hot Fluids"),
                    (r"\bhot\b.*\b(object|surface|metal|pipe)\b",
                        "Hot Object or Substances"),

                    # Hand tools, powered / non-powered / machinery
                    (r"\b(hand\s+tool|utensil)\b(?!\s*powered)",
                        "Hand Tool, Utensil, Not Powered"),
                    (r"\b(powered\s+hand\s+tool|power\s+tool|saw\b|drill\b|grinder\b)\b",
                        "Powered Hand Tool, Appliance"),
                    (r"\bmachine\b|\bmachinery\b|\bconveyor\b|\bpress\b",
                        "Machine or Machinery"),
                    (r"\busing\b.*\b(tool|machinery|machine)\b",
                        "Using Tool or Machinery"),

                    # Manual handling / strain-type causes
                    (r"\blift(ing|ed)?\b",
                        "Lifting"),
                    (r"\b(pushing|pulling|push(ed|ing)?|pull(ed|ing)?)\b",
                        "Pushing or Pulling"),
                    (r"\bholding\b|\bcarrying\b|\bcarried\b",
                        "Strain/Injury By - From Holding or Carrying"),
                    (r"\brepetitive\b|\brepetition\b",
                        "Strain/Injury By - Repetitive Motion"),
                    (r"\breach(ing)?\b|\bstretched\b",
                        "Reaching"),
                    (r"\btwist(ed|ing)?\b|\btwisting\b",
                        "Twisting"),
                    (r"\bstrain\b|\boverexert(ion|ed|ing)\b|\bmuscle\b.*\b(pulled|strain)\b",
                        "Strain or Injury by, NOC"),

                    # Rubbing / abrasion
                    (r"\brubb(ed|ing)?\b|\babrasion\b|\bchafe(d|ing)?\b",
                        "Rubbed or Abraded, NOC"),
                    (r"\brubb(ed|ing)?\b.*\brepetitive\b|\babrasion\b.*\brepetitive\b",
                        "Rubbed/Abraded By - Repetitive Motion"),

                    # Stationary object / stepping on / sharp object
                    (r"\bstruck\b.*\bagainst\b.*\b(stationary|fixed)\s+object\b|\bran\s+into\b",
                        "Stationary Object"),
                    (r"\bstepp(ed|ing)?\b.*\b(nail|screw|sharp|tack|glass)\b",
                        "Stepping on Sharp Object"),
                    (r"\bstepp(ed|ing)?\b|\bstruck\b.*\bagainst\b",
                        "Striking Against or Stepping On, NOC"),

                    # Struck / injured by variations
                    (r"\bstruck\b.*\b(falling|flying)\b|\bfalling\b.*\bobject\b|\bflying\b.*\bobject\b",
                        "Struck/Injured By - Falling or Flying Object"),
                    (r"\bstruck\b|\bhit\b|\binjured\b.*\bby\b.*\b(co-worker|coworker|fellow worker|patient)\b",
                        "Struck/Injured By - Fellow Worker, Patient"),
                    (r"\bstruck\b|\bhit\b|\binjured\b.*\bby\b.*\b(hand\s+tool|machine\s+in\s+use|power\s+tool)\b",
                        "Struck/Injured By - Hand Tool or Machine in Use"),
                    (r"\bstruck\b|\bhit\b|\binjured\b.*\bby\b.*\b(vehicle|car|truck|forklift)\b",
                        "Struck/Injured By - Motor Vehicle"),
                    (r"\bstruck\b|\bhit\b|\binjured\b.*\bby\b.*\b(moving\s+part(s)?\b.*\bmachine)\b",
                        "Struck/Injured By - Moving Parts of Machine"),
                    (r"\bstruck\b|\bhit\b|\binjured\b.*\bby\b.*\b(object\b.*handled\b.*other(s)?|object\b.*another\b.*worker)\b",
                        "Struck/Injured By - Object Handled by Others"),
                    (r"\bstruck\b|\bhit\b|\binjured\b.*\bby\b",
                        "Struck/Injured, NOC"),

                    # Temperature extremes
                    (r"\bheat\s+exhaustion\b|\bheat\s+stroke\b|\bheat\s+stress\b|\bextreme\s+heat\b|\bextreme\s+cold\b|\btemperature\s+extreme(s)?\b",
                        "Temperature Extremes"),

                    # Radiation
                    (r"\bradiation\b|\bx-ray(s)?\b|\bxray(s)?\b",
                        "Radiation"),

                    # Mold
                    (r"\bmold\b|\bmould\b",
                        "Mold"),

                    # Natural disasters / terrorism / crime
                    (r"\bearthquake\b|\btornado\b|\bhurricane\b|\bflood\b|\btsunami\b",
                        "Natural Disasters"),
                    (r"\bterroris(m|t)\b|\bactive\s+shooter\b|\bbomb\b",
                        "Terrorism"),
                    (r"\bassault\b|\brobbery\b|\bcrime\b|\barmed\b\s+robbery\b",
                        "Person in Act of a Crime"),

                    # Gunshot
                    (r"\bgunshot\b|\bshot\b.*\bgun\b|\bfirearm\b",
                        "Gunshot"),

                    # Sanding / scraping / cleaning
                    (r"\bsanding\b|\bscrap(ing|e)\b|\bcleaning\b.*\boperation\b",
                        "Sanding, Scraping, Cleaning Operation"),

                    # Cumulative NOC
                    (r"\bcumulative\b",
                        "Cumulative, NOC"),

                    # Other than physical cause (stress only, etc.)
                    (r"\bmental\b|\bstress\b|\bpsycho(logical)?\b",
                        "Other Than Physical Cause of Injury"),
                ]

                for pat, value in patterns:
                    if re.search(pat, text):
                        return value

                return None

            # ---------- PRIORITY: CORE TEXT THEN SPECIAL INSTRUCTIONS ----------
            cause = classify_cause(core_text)
            if cause:
                return cause

            cause = classify_cause(si_text)
            if cause:
                return cause

            # Final default
            return "Other - Miscellaneous, NOC"
        
        bodyPart     = resolve_body_part(AccidentDesc,InjuryDesc,specialInstructions)
        injuryType   = resolve_injury_type(AccidentDesc,InjuryDesc,specialInstructions)
        injuryCause  = resolve_cause(AccidentDesc,InjuryDesc,specialInstructions)

        URTServiceType=""
        URTrefType=""
        URTCaseObjective=""


        #------------------------
        # if claimNumber == "WC390M11643":
        #     nextApptDate = "01/22/2026"
            # nextApptTime = "1:50 PM"
        #    providerAddr ="Hoover/Medplex 4517 Southlake Pkwy"
            # providerCity = "Hoover"
            # providerState = "AL"
            # providerZip = "35244"
           # refSource = "Juan Pedro"
        #------------------------- staticcc

        

        # =========================
        # ADD-ONLY FALLBACK (do not change existing logic)
        # Runs ONLY if provider address was not captured
        # =========================
        if (not providerAddr) and specialInstructions:
            # New format:
            # PROVIDER: ...
            # Addr: 1207 S BAILEY, ELECTRA, TX 763603221
            m_addr = re.search(r"(?im)^\s*Addr\.?\s*:\s*(.+?)\s*$", specialInstructions)
            if m_addr:
                addr_line = m_addr.group(1).strip()

                # If ZIP is written as 9 digits (763603221), normalize to ZIP+4
                # WITHOUT changing your main parse logic
                addr_line = re.sub(r"\b(\d{5})(\d{4})\b", r"\1-\2", addr_line)

                # If ZIP+4 has spaces around dash: 92056 - 3694
                addr_line = re.sub(r"\b(\d{5})\s*[-–—]\s*(\d{4})\b", r"\1-\2", addr_line)

                a1_, a2_, city_, st_, z_ = parse_address(addr_line)

                if a1_ and not providerAddr:
                    providerAddr = a1_
                if a2_ and not providerAddr2:
                    providerAddr2 = a2_
                if city_ and not providerCity:
                    providerCity = city_
                if st_ and not providerState:
                    providerState = st_
                if z_ and not providerZip:
                    providerZip = z_[:5]   # ✅ force 5 digits only

        # Optional: also capture provider name/phone only if missing (still add-only)
        if (not providerName) and specialInstructions:
            m_prov = re.search(r"(?im)^\s*PROVIDER\s*:\s*(.+?)\s*$", specialInstructions)
            if m_prov:
                providerName = m_prov.group(1).strip()

        if (not providerPhone) and specialInstructions:
            # m_ph = re.search(r"(?im)^\s*Ph\s*#\s*:\s*(.+?)\s*$", specialInstructions)
            m_ph = re.search(r"(?im)^\s*(?:Organization\s+Phone|Provider\s+Phone|Facility\s+Phone|Ph\s*#|Phone\s*(?:Number|#)?)\s*:\s*(.+?)\s*$",specialInstructions)
            if m_ph:
                providerPhone = m_ph.group(1).strip()

        if "full case management" in referralType.lower() or "one-time rn visit" in referralType.lower():
            correctZipCode = providerZip

            URTServiceType ="Medical Full"
            URTrefType="Medical"
            URTCaseObjective="Coordination of Care/Services"
        else:
            URTServiceType ="Medical Full"
            URTrefType="Medical"
            URTCaseObjective="Coordination of Care/Services"
            correctZipCode = zip_

        # ---------- package ---------- 2xA33A2 migz
        return {
            "customer": customer,
            "claimantFull": claimantFull, "claimantFirst": claimantFirst, "claimantLast": claimantLast,
            "claimantSSN": claimantSSN,
            "claimNumber": claimNumber,
            "claimID": claimID,
            "claimType":claimTypeLOI,
            "dxCode":dxCode,
            "referralType": referralType,
            "claimStateAbbr": claimStateAbbr, "claimStateFull": claimStateFull,
            "addressLine1": addressLine1, "addressLine2": addressLine2,
            "city": city, "state": state, "zip": zip_,
            "phoneNumber": phoneNumber, "gender": gender,
            "dob": dob, "doi": doi,
            "refSource": refSource, "adjuster": adjuster,
            "adjEmail":adjEmail,
            "adjAddrFull":adjAddrFull,"adjPhone":adjPhone,
            "adjAddr1":adjAddr1,"adjAddr2":adjAddr2,"adjCity":adjCity,"adjState":adjState,"adjZip":adjZip,
            "specialInstructions": specialInstructions,
            "employerContactName":employerContactName,
            # "employerContactName": employerContactName,
            "employerContactEmail": employerContactEmail,
            "employerContactPhone": employerContactPhone,
            "ncmContactName":ncmContactName,
            "referralNumber": referralNumber,
            "providerName": providerName, "providerAddr": providerAddr,
            "providerCity": providerCity, "providerState": providerState, "providerZip": providerZip,
            "providerPhone": providerPhone,
            "AccidentDesc": AccidentDesc, "InjuryDesc": InjuryDesc,
            "nextApptDate": nextApptDate, "nextApptTime": nextApptTime,
            "providerFirst": providerFirst, "providerLast": providerLast,
            # convenience group flags (optional, based on your AHK lists)
            "refTypeInList1": referralType in list1,
            "refTypeInList2": referralType in list2,
            "claimantAttyName": ClaimAttyName,
            "claimantAttyAddr1": ClaimAttyaddressLine1,
            "claimantAttyAddr2": ClaimAttyaddressLine2,
            "claimantAttyCity": ClaimAttycity,
            "claimantAttyState": ClaimAttystate,
            "claimantAttyZip": ClaimAttyZip,
            "claimantAttyPhone": ClaimAttyPhoneNumber,
            "injuryType": injuryType,
            "injuryCause": injuryCause,
            "bodyPart": bodyPart,
            "languageSpoken":languageSpoken,
            "ServiceType":URTServiceType,
            "refType":URTrefType,
            "CaseObj":URTCaseObjective,
            "adjAddrMail": adjAddrMail,
        }
    except Exception as ex:
        exc_type, exc_obj, tb = sys.exc_info()
        line_number = tb.tb_lineno
        print(f"error occurred: {ex} (line {line_number})")
        traceback.print_exc()
        return {}
        # print("error occurred: ", ex)

UPPER_KEEP = {
    "PO","BOX","APT","STE","UNIT","BLDG","FL","RM",
    "NW","NE","SW","SE","N","S","E","W",
    "USA",
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
    "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
    "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
    "VA","WA","WV","WI","WY","DC"
}


def address_callback(word, **kwargs):
    # word includes original punctuation; strip for checks
    core = re.sub(r"[^\w]", "", word).upper()
    if core in UPPER_KEEP:
        return core
    # keep apartment numbers etc. as-is
    if re.fullmatch(r"#?\d[\w\-]*", word):
        return word
    return None  # let titlecase handle normally

def OpenPDFAndCaptureData(PDFName):
    global data
    PDFBrowser = Desktop(backend="uia").window(title_re=f"{PDFName}")#,control_type="Window"
    PDFBrowser.wait("visible ready",timeout=60)
    # PDFBrowser.print_control_identifiers()
    focus_control(PDFBrowser)
    try:
        PDFDoc = PDFBrowser.child_window(title=PDFName ,control_type="Document").wrapper_object()
        PDFData = PDFDoc.iface_text
        # print(PDFData.DocumentRange.GetText(-1))
        # PDFDoc = PDFBrowser.child_window(auto_id="RootWebArea",control_type="Document")
        # PDFDoc.wait("visible ready",timeout=60)
        # PDFBrowser.set_focus()
    except Exception:
        # docs = PDFBrowser.descendants(control_type="Document")
        # match = [d for d in docs if "PDF document containing" in d.window_text()]
        # if match:
        #     PDFDoc = match[0].wrapper_object()
        # else:
        #     PDFDoc = max(docs,key=lambda d: d.rectangle().height()).wrapper_object()
        pass

    data = parse_referral_pdf(PDFData.DocumentRange.GetText(-1))
    save_employer_contact_to_excel(data)
    # pprint(data)
    # print(data["claimantFull"], data["claimNumber"], data["providerAddr"])
    # print("-----------------------------------------------")
    # print(PDFData.DocumentRange.GetText(-1))
    return data

# def ValidateInfo(allData :dict):
#     if allData['referralType'] == "Full Case Management" or allData['referralType'] == "One-Time RN Visit - Provider":
#         if allData['providerName'] or (allData['providerFirst'] or allData['providerLast']):
#             if allData['providerAddr'] and allData['providerCity'] and allData['providerState'] and allData['providerZip']:
#                 print ((not allData['nextApptDate'] and not allData['nextApptTime']) or (allData['nextApptDate'] and allData['nextApptTime']))
#                 if allData['providerPhone'] and ((not allData['nextApptDate'] and not allData['nextApptTime']) or (allData['nextApptDate'] and allData['nextApptTime'])):
#                     return True 
#migz

def notify(title: str, message: str) -> None:
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost',True)
    messagebox.showinfo(title, message)
    root.destroy()
    # else:
    #     print(f"[{title}] {message}")

def get_validation_missing(allData: dict) -> list[str]:

    referral_type = (allData.get("referralType") or "").strip()
    if referral_type not in ("Full Case Management", "One-Time RN Visit - Provider"):
        return []

    missing = []

    provider_name_ok = bool(allData.get("providerName")) or bool(allData.get("providerFirst")) or bool(allData.get("providerLast"))
    if not provider_name_ok:
        missing.append("Provider name (providerName OR providerFirst/providerLast)")

    addr_fields = [
        ("providerAddr",  "Provider address"),
        ("providerCity",  "Provider city"),
        ("providerState", "Provider state"),
        ("providerZip",   "Provider ZIP"),
    ]
    for key, label in addr_fields:
        if not (allData.get(key) or "").strip():
            missing.append(label)

    # if not (allData.get("providerPhone") or "").strip():
    #     missing.append("Provider phone")

    # next_date = (allData.get("nextApptDate") or "").strip()
    # next_time = (allData.get("nextApptTime") or "").strip()
    # if bool(next_date) ^ bool(next_time):
    #     missing.append("Next appointment must have BOTH date and time (or leave BOTH blank)")

    return missing


def build_provider_search_text(allData: dict) -> str:
    """
    Build PROVIDERSRCH (your requested input name).
    You can tweak the order anytime.
    """
    parts = [
        allData.get("providerName", ""),
        allData.get("providerAddr", ""),     # sometimes partial street exists
        allData.get("providerCity", ""),
        allData.get("providerState", ""),
        allData.get("providerZip", ""),
        allData.get("providerPhone", ""),
    ]
    s = " ".join(p for p in parts if p and str(p).strip())
    s = re.sub(r"\s+", " ", s).strip()
    # pprint(s)
    return s


def apply_google_provider_result(allData: dict, picked: dict) -> None:
    """
    Apply Google-picked provider info into allData.
    Fills ONLY when non-empty values are returned.
    """
    if not isinstance(picked, dict):
        return

    # the picker returns these keys by your design
    addr = (picked.get("providerAddr") or "").strip()
    city = (picked.get("providerCity") or "").strip()
    st   = (picked.get("providerState") or "").strip()
    zp   = (picked.get("providerZip") or "").strip()
    ph   = (picked.get("providerPhone") or "").strip()
    ph = format_phone_us(ph)

    if addr: allData["providerAddr"] = addr
    if city: allData["providerCity"] = city
    if st:   allData["providerState"] = st
    if zp:   allData["providerZip"] = zp
    if ph:   allData["providerPhone"] = ph

def ValidateInfo(allData: dict) -> bool:
    referral_type = (allData.get("referralType") or "").strip()
    if referral_type not in ("Full Case Management", "One-Time RN Visit - Provider"):
        return True

    missing = []

    # Provider name (either providerName OR providerFirst/providerLast)
    provider_name_ok = bool(allData.get("providerName")) or bool(allData.get("providerFirst")) or bool(allData.get("providerLast"))
    if not provider_name_ok:
        missing.append("Provider name (providerName OR providerFirst/providerLast)")

    # Provider address pieces
    addr_fields = [
        ("providerAddr",  "Provider address"),
        ("providerCity",  "Provider city"),
        ("providerState", "Provider state"),
        ("providerZip",   "Provider ZIP"),
    ]
    for key, label in addr_fields:
        if not allData.get(key):
            missing.append(label)

    # Provider phone
    # if not allData.get("providerPhone"):
    #     missing.append("Provider phone")

    # Next appt rule: either BOTH blank, or BOTH filled (date + time)
    # next_date = allData.get("nextApptDate")
    # next_time = allData.get("nextApptTime")
    # if bool(next_date) ^ bool(next_time):
    #     missing.append("Next appointment must have BOTH date and time (or leave BOTH blank)")

    if missing:
        msg = "Missing / invalid information:\n\n" + "\n".join(f"• {m}" for m in missing)
        notify("Validation Failed", msg + "\n\n The automation will stop after you click OK")
        return False

    return True

    #Test1
    # else:
    #     notify("Notice","For Manual Process. Other Referral Type aside from Full Case Management is out of scope yet")            
    #     sys.exit()

    return False

GOODYEAR_FACILITIES = [
    "PL0153 Lawton",
    "RE Retail Stores",
    "PL0133 Topeka",
    "CT Commercial Tire Centers",
    "CP0001 Texarkana",
    "PL0126 Danville",
    "PL0161 Fayetteville",
    "PL0155 Asheboro",
    "CP0016 Tupelo",
    "PL0130 Social Circle",
    "PL0131 Gadsden",
    "PL0933 Akron",
    "CP0018 Findlay",
    "CH0148 Bayport",
    "PL0129 Union City",
    "MS9999 Mileage Sales",
    "MS9999 Lake Buena Vista",
    "MS9999 Sun Valley",
    "MS9999 Chatsworth",
    "MS9999 El Monte",
    "PL0170 Kingman",
    "CH0141 Houston",
    "CH0137 Niagara Falls",
    "EP0112 Saint Marys",
    "ML1286 Eglin Air Force Base",
]
 
 
def GoodyearSearchFacilityName(special_instructions: str) -> str:
    if not special_instructions:
        return ""
 
    text = special_instructions.lower()
 
    for item in GOODYEAR_FACILITIES:
        if item.lower() in text:
            return item
 
    return ""


def to_mmddyyyy(date_str: str) -> str:
    if not date_str:
        return ""
    date_str = date_str.strip()
    dt = datetime.strptime(date_str, "%d-%b-%Y")   # 27-Feb-1994
    return dt.strftime("%m/%d/%Y")                 # 02/27/1994

def CreateSubjectLineBuilder(app = None):
    try:
        global isManualEntered
        global botStop
        global data
        global correctZipCode
        botStop = True
        main = Desktop(backend="uia").window(title_re=r"Referral Routing System*",control_type="Window")
        focus_control(main)
        # pid = main.element_info.process_id
        allData = {}
        grid = main.child_window(auto_id="dgWorkWindow",control_type="Table")#.wrapper_object()
        # Selecting from DatagridView
        # try:
        #     rowtbCount = grid.iface_grid.CurrentRowCount
        #     coltbCount = grid.iface_grid.CurrentColumnCount
        #     print(f"Rpws: {rowtbCount} Col:{coltbCount}")
        # except Exception as e:
        #     print(e)
        
        # rowsSelection = grid.descendants(control_type="DataItem")
        # if not rowsSelection:
        #     raise Exception("No Rows found in Grid")
        # rowsSelection = rowsSelection[0]
        # cells = rowsSelection.children()
        # if len(cells) <= 4:
        #     cells = rowsSelection.descendants()

        # cell = cells[4]
        # print(repr(cell.window_text()))

        cell_el = grid.child_window(title="Claimant Row 0",control_type="Edit").wrapper_object()
        ClaimantRRS = cell_el.iface_value.CurrentValue

        # cell_el = grid.iface_grid.GetItem(0,4)
        # cell = UIAWrapper(UIAElementInfo(cell_el))
        # ClaimantRRS = cell.iface_value.CurrentValue
        FNameRRS = re.split(r"\s+",ClaimantRRS)[0]
        # print(FNameRRS)
        cell_el.double_click_input()

        emaiDisplay = Desktop(backend="uia").window(title_re=r"Email Display*",control_type="Window")
        focus_control(emaiDisplay)

        grid = emaiDisplay.child_window(auto_id="dgAttach",control_type="Table")
        AttachRow = grid.iface_grid.CurrentRowCount

        for r in range(AttachRow):
            # cell_el = grid.iface_grid.GetItem(r,2)
            # cell = UIAWrapper(UIAElementInfo(cell_el))

            cell = grid.child_window(title=f"Filename Row {r}, Not sorted.", control_type="Edit").wrapper_object()


            AttachmentName = cell.iface_value.CurrentValue
            Attachment = AttachmentName.lower()
            has_11_letter_w = bool(re.search(r"\bw\w{10}\b", Attachment))


            # if FNameRRS.lower() in AttachmentName.lower():
            if not "icasemanager" in Attachment and ("," in Attachment or has_11_letter_w):
                # print(AttachmentName)
                focus_control(emaiDisplay)
                cell.double_click_input()
                allData = OpenPDFAndCaptureData(AttachmentName)
                print(f'Fetching Information if Referral PDF.')
                pprint(allData)
                #INSERT RETRIEVAL OF DATA

                break
        # def LMContinentalTire(allData:dict):
        if "auto" in allData['claimType'].lower() and "continental tire" in allData['customer'].lower() and "mi" in allData['claimStateAbbr'].lower() :
            allData['ServiceType'] ="Medical Task"
            allData['refType'] ="Task"
            allData['CaseObj'] = "MI MCCA Attendant Care"

        #Goodyear Tire with Facility
        GYFacility = GoodyearSearchFacilityName(allData['specialInstructions'])

        #Comcast
        # if "comcast" in allData['customer'].lower():
        #     notify("Notice","For Manual Process. Refer to Standard Intake Rules Guidelines for Comcast.")
        #     botStop=True
        #     sys.exit()
        
        #Genetech
        # if "genetech" in allData['customer'].lower():
        #     notify("Notice","For Manual Process. Refer to Standard Intake Rules Guidelines for Genetech.")
        #     botStop=True
        #     sys.exit()
        


        #Test1
        # allData['providerZip']=""
        missing = get_validation_missing(allData)
        def _click_first_row_if_any(prvType, FacilityName, ProvFirst, ProvLast):
                global isManualEntered
                # FacilityName = "Roanoke Orthopedic"
                try:
                    grid = ups.child_window(auto_id="dgResults", control_type="Table")
                    rows = int(getattr(grid.iface_grid, "CurrentRowCount", 0) or 0)
            
                    if rows <= 0:
                        isManualEntered = False
                        return False
            
                    RealAddr = f"{allData['providerAddr']} {FacCity} {FacState} {FacZip}"
                    RealAddress = expand_suffix_long(RealAddr)
                    realPhone = format_phone_us(FacPhoneNumber)
            
                    # Read all row data first
                    row_data = []
                    for r in range(rows):
                        print(f'Fetching Result Row {r}. Total Row {rows}')
                        try:
                            name_cell = grid.child_window(title=f"Name Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                            ResultName = name_cell.iface_value.CurrentValue
                            print(ResultName)
                            addr_cell = grid.child_window(title=f"Address Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                            ResultAddress = addr_cell.iface_value.CurrentValue
            
                            phone_cell = grid.child_window(title=f"Phone Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                            ResultPhone = phone_cell.iface_value.CurrentValue
            
                            TempAddress = expand_suffix_long(ResultAddress)
                            ComparisonResult = addressMatch(TempAddress, RealAddress)
            
                            tempPhone = format_phone_us(ResultPhone)
            
                            phone_match = False
                            if realPhone and tempPhone:
                                phone_match = (realPhone == tempPhone)

                            similarity = difflib.SequenceMatcher(
                            None, TempAddress.lower(), RealAddress.lower()
                            ).ratio()
            
                            row_data.append({
                                "row": r,
                                "name": ResultName or "",
                                "address": ResultAddress or "",
                                "phone": ResultPhone or "",
                                "address_match": ComparisonResult,
                                "phone_match": phone_match,
                                "similarity":similarity
                            })
                        except Exception:
                            continue
            
                    # If only one row exists, you can still apply your logic first
                    # instead of blindly selecting it.
                    # ------------------------------
                    # PASS 1: Strong name match + address + phone
                    # ------------------------------
                    print(f'Validating fetched infomation from Provider Search')
                    for item in row_data:
                        # print(item["similarity"])
                        # print(item["similarity"] >= 0.76)
                        if not (item["address_match"] or item["similarity"] >= 0.76):
                            continue
                        if not item["phone_match"]:
                            continue

                        if prvType == "fac":
                            if _facility_name_strong_match(FacilityName, item["name"]):
                                _click_select_row(grid, item["row"])
                                time.sleep(2)
                                btnSelect = ups.child_window(auto_id="btnSelect", control_type="Button").wrapper_object()
                                safe_click(btnSelect)
                                isManualEntered = True
                                return True
            
                        elif prvType == "prof":
                            if _provider_name_strong_match(ProvFirst, ProvLast, item["name"]):
                                _click_select_row(grid, item["row"])
                                time.sleep(2)
                                btnSelect = ups.child_window(auto_id="btnSelect", control_type="Button").wrapper_object()
                                safe_click(btnSelect)
                                isManualEntered = True
                                return True
            
                    # ------------------------------
                    # PASS 2: Loose name match + address + phone
                    # ------------------------------
                    for item in row_data:
                        # print(f'{item["similarity"]}')
                        # print(item["similarity"] >= 0.76)
                        if not (item["address_match"] or item["similarity"] >= 0.76):
                            continue
                        if not item["phone_match"]:
                            continue
            
                        if prvType == "fac":
                            
                            # if _facility_name_loose_match(FacilityName, item["name"]):
                                googlePrvName = prvGoogleSearch(item["name"])
                                if googlePrvName:
                                    # if _facility_name_loose_match(googlePrvName,item["name"]):
                                    _click_select_row(grid, item["row"])
                                    time.sleep(2)
                                    btnSelect = ups.child_window(auto_id="btnSelect", control_type="Button").wrapper_object()
                                    safe_click(btnSelect)
                                    isManualEntered = True
                                    return True
            
                        elif prvType == "prof":
                            looseCheck = _provider_name_loose_match(ProvFirst, ProvLast, item["name"])
                            if not looseCheck:
                                googlePrvName = prvGoogleSearch(item["name"])
                                if googlePrvName:
                                    # if _provider_name_loose_match(ProvFirst, ProvLast,item["name"]):
                                # prvGoogleSearch(item["name"])
                                    _click_select_row(grid, item["row"])
                                    time.sleep(2)
                                    btnSelect = ups.child_window(auto_id="btnSelect", control_type="Button").wrapper_object()
                                    safe_click(btnSelect)
                                    isManualEntered = True
                                return True
                            else:
                                _click_select_row(grid, item["row"])
                                time.sleep(2)
                                btnSelect = ups.child_window(auto_id="btnSelect", control_type="Button").wrapper_object()
                                safe_click(btnSelect)
                                isManualEntered = True
                                return True
            
                    isManualEntered = False
                    return False
            
                except Exception as e:
                    print(f"_click_first_row_if_any error: {e}")
                    isManualEntered = False
                    return False
                
        def run_provider_search():
            global isManualEntered
            FacilityLName = allData['providerLast']
            FacilityFName = allData['providerFirst']
            FacilityName  = allData['providerName']
            FacPhoneNumber = allData['providerPhone']
            FacCity = allData['providerCity']
            FacState = allData['providerState']
            FacZip = allData['providerZip']
            first = (FacilityFName or "").strip()
            last  = (FacilityLName or "").strip()
            fac   = (FacilityName  or "").strip()
            phone = (FacPhoneNumber or "").strip()
            city  = (FacCity or "").strip()
            state = (FacState or "").strip()
            zip_  = (FacZip or "").strip()

            isManualEntered = False
            firstAttempt = False
            facilityFirst = ""
            facility = " ".join(fac.strip().split())
            facilityWords = re.findall(r"[A-Za-z0-9&]+",facility)

            if not facilityWords:
                facilityFirst = ""
            elif facilityWords[0].upper() == "THE":
                facilityFirst = facilityWords[1] if len(facilityWords) > 1 else ""
            else:
                facilityFirst = facilityWords[0]

            if (tbZipVal or cbStateVal or tbCityVal or tbPhoneVal or tbFacilityVal or tbFirstVal or tbLastVal):
                _click_search()
                _wait_search_idle()
                count = _results_count()
                if count > 0:
                    if tbFacilityVal:
                        prvType = 'fac'
                        # if tbFacilityVal:
                            # fac = tbFacilityVal
                        SelectedPrvResult = firstAttempt = _click_first_row_if_any(prvType,fac,None,None)
                        if firstAttempt:
                            return count,SelectedPrvResult
                    else:
                        prvType = 'prof'
                        SelectedPrvResult = firstAttempt = _click_first_row_if_any(prvType,None,first,last)
                        if SelectedPrvResult:
                            return count,SelectedPrvResult

            # else:
                # 1) If Last+First both present, use them (with phone/city/state/zip)
            if fac and (not first and not last):
                count = _search_with(first=None, last=None, facility=facilityFirst,
                                    phone=phone, city=None, state=None, zip_=None)
                prvType = 'fac'
                if count <= 10:
                    SelectedPrvResult = firstAttempt = _click_first_row_if_any(prvType,facilityFirst,None,None)
                    if firstAttempt:
                        return count,SelectedPrvResult
            else:
                count = _search_with(first=first, last=last, facility=None,
                                    phone=phone, city=None, state=None, zip_=None)
                prvType = 'prof'
                if count <= 10:
                    SelectedPrvResult = firstAttempt = _click_first_row_if_any(prvType,None,first,last)
                    if firstAttempt:
                        return count,SelectedPrvResult
            # 2) Else if FacilityName present, use it (with phone/city/state/zip)
            if not firstAttempt:
                if fac and (not first and not last):
                    count = _search_with(first=None, last=None, facility=facilityFirst,
                                        phone=phone, city=city, state=state, zip_=zip_)
                    prvType = 'fac'
                    if count > 0:
                        SelectedPrvResult = _click_first_row_if_any(prvType,facilityFirst,None,None)
                        return count,SelectedPrvResult
                else:
                    count = _search_with(first=first, last=last, facility=None,
                                        phone=phone, city=city, state=state, zip_=zip_)
                    prvType = 'prof'
                    if count > 0:
                        SelectedPrvResult = _click_first_row_if_any(prvType,None,first,last)
                        return count,SelectedPrvResult
            # return 0,False
            return 0,False

        def run_provider_search_part():
                global isManualEntered

                # tbProviderLastExtract = ups.child_window(auto_id="tbProviderLastExtract", control_type="Edit").wrapper_object()
                # tbProviderLastExtractVal = (tbProviderLastExtract.iface_value.CurrentValue or "").strip()

                tbLast = ups.child_window(auto_id="tbLast", control_type="Edit").wrapper_object()
                tbLastVal = (tbLast.iface_value.CurrentValue or "").strip()
                # if tbLastVal:
                #     allData['providerLast'] = tbLastVal
                # elif tbProviderLastExtractVal:
                #     allData['providerLast'] = tbProviderLastExtractVal
                
                # tbProviderFirstExtract = ups.child_window(auto_id="tbProviderFirstExtract", control_type="Edit").wrapper_object()
                # tbProviderFirstExtractVal = (tbProviderFirstExtract.iface_value.CurrentValue or "").strip()
                tbFirst = ups.child_window(auto_id="tbFirst", control_type="Edit").wrapper_object()
                tbFirstVal = (tbFirst.iface_value.CurrentValue or "").strip()
                # if tbFirstVal:
                #     allData['providerFirst'] = tbFirstVal
                # elif tbProviderFirstExtractVal:
                #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
                #     allData['providerFirst'] = tbProviderFirstExtractVal
                
                # tbProviderNameExtract = ups.child_window(auto_id="tbProviderNameExtract", control_type="Edit").wrapper_object()
                # tbProviderNameExtractVal = (tbProviderNameExtract.iface_value.CurrentValue or "").strip()
                tbFacility = ups.child_window(auto_id="tbFacility", control_type="Edit").wrapper_object()
                tbFacilityVal = (tbFacility.iface_value.CurrentValue or "").strip()
                # if tbFacilityVal:
                #     allData['providerName'] = tbFacilityVal
                # elif tbProviderNameExtractVal:
                #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
                #     allData['providerName'] = tbProviderNameExtractVal
                
                # tbAddressExtract = ups.child_window(auto_id="tbAddressExtract", control_type="Edit").wrapper_object()
                # tbAddressExtractVal = (tbAddressExtract.iface_value.CurrentValue or "").strip()
                # if tbAddressExtractVal:
                #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
                #     allData['providerAddr'] = tbAddressExtractVal
                
                # tbPhoneExtract = ups.child_window(auto_id="tbPhoneExtract", control_type="Edit").wrapper_object()
                # tbPhoneExtractVal = (tbPhoneExtract.iface_value.CurrentValue or "").strip()
                tbPhone = ups.child_window(auto_id="tbPhone", control_type="Edit").wrapper_object()
                tbPhoneVal = (tbPhone.iface_value.CurrentValue or "").strip()
                # if tbPhoneVal:
                #     allData['providerPhone'] = tbPhoneVal
                # elif tbPhoneExtractVal:
                #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
                #     allData['providerPhone'] = tbPhoneExtractVal
                
                # tbCityExtract = ups.child_window(auto_id="tbCityExtract", control_type="Edit").wrapper_object()
                # tbCityExtractVal = (tbCityExtract.iface_value.CurrentValue or "").strip()
                tbCity = ups.child_window(auto_id="tbCity", control_type="Edit").wrapper_object()
                tbCityVal = (tbCity.iface_value.CurrentValue or "").strip()
                # if tbCityVal:
                #     allData['providerCity'] = tbCityVal
                # elif tbCityExtractVal:
                #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
                #     allData['providerCity'] = tbCityExtractVal
                
                # tbStateExtract = ups.child_window(auto_id="tbStateExtract", control_type="Edit").wrapper_object()
                # tbStateExtractVal = (tbStateExtract.iface_value.CurrentValue or "").strip()
                cbState = ups.child_window(auto_id="cbState", control_type="ComboBox").wrapper_object()
                cbStateVal = (cbState.iface_value.CurrentValue or "").strip()
                # if cbStateVal:
                #     allData['providerState'] = cbStateVal
                # elif tbStateExtractVal:
                #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
                #     allData['providerState'] = tbStateExtractVal
                
                # tbZipExtract = ups.child_window(auto_id="tbZipExtract", control_type="Edit").wrapper_object()
                # tbZipExtractVal = (tbZipExtract.iface_value.CurrentValue or "").strip()
                tbZip = ups.child_window(auto_id="tbZip", control_type="Edit").wrapper_object()
                tbZipVal = (tbZip.iface_value.CurrentValue or "").strip()
                # if tbZipVal:
                #     allData['providerZip'] = tbZipVal
                # elif tbZipExtractVal:
                #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
                #     allData['providerZip'] = tbZipExtractVal

                FacilityLName = allData['providerLast']
                FacilityFName = allData['providerFirst']
                FacilityName  = allData['providerName']
                FacPhoneNumber = allData['providerPhone']
                FacCity = allData['providerCity']
                FacState = allData['providerState']
                FacZip = allData['providerZip']
                first = (FacilityFName or "").strip()
                last  = (FacilityLName or "").strip()
                fac   = (FacilityName  or "").strip()
                phone = (FacPhoneNumber or "").strip()
                city  = (FacCity or "").strip()
                state = (FacState or "").strip()
                zip_  = (FacZip or "").strip()

                isManualEntered = False
                firstAttempt = False
                facilityFirst = ""
                facility = " ".join(fac.strip().split())
                facility = re.sub(r"[^A-Za-z0-9\s,\.\-]", "" ,facility)
                facility = re.sub(r"\s+", " " ,facility).strip()
                facility = re.sub(r"[-\s]+$", " " ,facility).strip()
                # print(facility)
                facilityWords = re.findall(r"[A-Za-z0-9&]+",facility)


                if not facilityWords:
                    facilityFirst = ""
                elif facilityWords[0].upper() == "THE":
                    facilityFirst = facilityWords[1] if len(facilityWords) > 1 else ""
                else:
                    facilityFirst = facilityWords[0]

                if (tbZipVal or cbStateVal or tbCityVal or tbPhoneVal or tbFacilityVal or tbFirstVal or tbLastVal):
                    _click_search()
                    _wait_search_idle()
                    count = _results_count()
                    if count > 0:
                        if tbFacilityVal:
                            prvType = 'fac'
                            SelectedPrvResult = firstAttempt = _click_first_row_if_any(prvType,fac,None,None)
                            if firstAttempt:
                                return count,SelectedPrvResult
                        else:
                            prvType = 'prof'
                            SelectedPrvResult = firstAttempt = _click_first_row_if_any(prvType,None,first,last)
                            if SelectedPrvResult:
                                return count,SelectedPrvResult

                # 1) If Last+First both present, use them (with phone/city/state/zip)
                if fac and (not first and not last):
                    count = _search_with(first=None, last=None, facility=facility,
                                        phone=phone, city=None, state=None, zip_=None)
                    prvType = 'fac'
                    if count <= 10:
                        firstAttempt,NewResultName,NewResultAddr,NewResultPhone = CheckPrvFromList(prvType,facility,None,None)
                        if firstAttempt:
                            return count,NewResultName,NewResultAddr,NewResultPhone
                else:
                    count = _search_with(first=first, last=last, facility=None,
                                        phone=phone, city=None, state=None, zip_=None)
                    prvType = 'prof'
                    if count <= 10:
                        firstAttempt,NewResultName,NewResultAddr,NewResultPhone = CheckPrvFromList(prvType,None,first,last)
                        if firstAttempt:
                            return count,NewResultName,NewResultAddr,NewResultPhone
                # 2) Else if FacilityName present, use it (with phone/city/state/zip)
                if not firstAttempt:
                    if fac and (not first and not last):
                        count = _search_with(first=None, last=None, facility=facility,
                                            phone=phone, city=city, state=state, zip_=zip_)
                        prvType = 'fac'
                        if count > 0:
                            firstAttempt,NewResultName,NewResultAddr,NewResultPhone = CheckPrvFromList(prvType,facility,None,None)
                            return count,NewResultName,NewResultAddr,NewResultPhone
                    else:
                        count = _search_with(first=first, last=last, facility=None,
                                            phone=phone, city=city, state=state, zip_=zip_)
                        prvType = 'prof'
                        if count > 0:
                            firstAttempt,NewResultName,NewResultAddr,NewResultPhone =CheckPrvFromList(prvType,None,first,last)
                            return count,NewResultName,NewResultAddr,NewResultPhone
                return 0,NewResultName,NewResultAddr,NewResultPhone

        #Add Search RRS
        if missing:
            print(f'Missing Provider Info. Searching thru RRS Provider Search.')
            focus_control(emaiDisplay)
            SubjectLineBuilder = emaiDisplay.child_window(auto_id="btnSubjectSearch",control_type="Button").wrapper_object()
            focus_control(SubjectLineBuilder)
            safe_click(SubjectLineBuilder)

            dlg = Desktop(backend="uia").window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
            dlg.wait("exists enabled visible ready",timeout=120,retry_interval=1)
            focus_control(dlg)

            tbCustomer = dlg.child_window(auto_id="tbCustomer", control_type="Edit").wrapper_object()
            tbCustomer.iface_value.SetValue("Liberty Mutual Commercial Market")
            time.sleep(3)

            rbProvider = dlg.child_window(auto_id="rbProvider", control_type="RadioButton").wrapper_object()
            rbProvider.select()
            time.sleep(3)
            # pbProvider (Pane Magnifying Glass)
            pbProvider = dlg.child_window(auto_id="pbProvider",control_type="Pane").wrapper_object()
            safe_click(pbProvider)
            time.sleep(3)
            #ADD Searching here!
            desk = Desktop(backend="uia")
            owner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
            owner.wait('exists',timeout=5.0)
            ups = owner.child_window(auto_id="frmUnityProviderLM",control_type="Window")
            ups.wait('exists',timeout=5.0)
            
            
            #Add Search Prv
            def CheckPrvFromList(prvType,FacilityName,ProvFirst,ProvLast):
                try:
                    grid = ups.child_window(auto_id="dgResults", control_type="Table")#.wrapper_object()
                    rows = int(getattr(grid.iface_grid, "CurrentRowCount", 0) or 0)
                    if rows >= 1:
                        for r in range(rows):
                            # cell_el = grid.iface_grid.GetItem(r,2)
                            # cell = UIAWrapper(UIAElementInfo(cell_el))
                            cell = grid.child_window(title=f"Name Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                            ResultName = cell.iface_value.CurrentValue
                            
                            if ProvFirst or ProvLast:
                                RealName = f"{ProvFirst} {ProvLast}"
                            else:
                                RealName = f"{FacilityName}"
                            CompareNameResult = RealName in ResultName


                            # cell_el = grid.iface_grid.GetItem(r,3)
                            # cell = UIAWrapper(UIAElementInfo(cell_el))
                            cell = grid.child_window(title=f"Address Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                            ResultAddress = cell.iface_value.CurrentValue
                            RealAddr = f"{allData['providerAddr']} {allData['providerCity']} {allData['providerState']} {allData['providerZip']}"
                            # print(RealAddr)
                            TempAddress = expand_suffix_long(ResultAddress)
                            RealAddress = expand_suffix_long(RealAddr)
                            ComparisonResult = addressMatch(TempAddress,RealAddress)

                            if allData['providerPhone']:
                                # cell_el = grid.iface_grid.GetItem(r,4)
                                # cell = UIAWrapper(UIAElementInfo(cell_el))
                                cell = grid.child_window(title=f"Phone Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                                ResultPhone = cell.iface_value.CurrentValue
                                realPhone = format_phone_us(allData['providerPhone'])
                                tempPhone = format_phone_us(ResultPhone)
                                ComparePhoneResult = realPhone in tempPhone
                            else:
                                ResultPhone = None
                                ComparePhoneResult = False
                                
                            if (CompareNameResult and ComparisonResult):
                                return True  ,ResultName, ResultAddress,ResultPhone
                        return False,None,None,None
                except Exception:
                    pass
                return False,None,None,None
            
            def _get_edit(auto_id):
                return ups.child_window(auto_id=auto_id, control_type="Edit").wrapper_object()

            def _set_value(ctrl, value):
                try:
                    ctrl.iface_value.SetValue("")   # clear
                    if value:
                        ctrl.iface_value.SetValue(value)

                except Exception:
                    # Fallback to type_keys if needed
                    try:
                        focus_control(ctrl)
                        ctrl.type_keys("^a{BACKSPACE}")
                        if value:
                            ctrl.type_keys(value, with_spaces=True)
                    except Exception:
                        pass

            def _select_state(value):
                try:
                    cb = ups.child_window(auto_id="cbState", control_type="ComboBox").wrapper_object()
                    if value:
                        cb.select(value)
                    else:
                        try:
                            cb.select(0)
                        except Exception:
                            pass
                except Exception:
                    pass

            def _click_search():
                btn = ups.child_window(auto_id="btnSearch", control_type="Button").wrapper_object()
                safe_click(btn)

            def _wait_search_idle(timeout=300.0):
                lbl = ups.child_window(auto_id="lblStatus", control_type="Text")
                end_t = time.monotonic() + timeout
                while time.monotonic() < end_t:
                    try:
                        w = lbl.wrapper_object()
                        if (not w.is_visible()) or ((w.window_text() or "").strip() == ""):
                            return True
                    except Exception:
                        return True
                    time.sleep(0.2)
                return False

            def _results_count():
                try:
                    grid = ups.child_window(auto_id="dgResults", control_type="Table").wrapper_object()
                    return int(getattr(grid.iface_grid, "CurrentRowCount", 0) or 0)
                except Exception:
                    return 0
                
            def _fill_common_filters(phone, city, state, zip_):
                # Phone
                try: _set_value(_get_edit("tbPhone"), phone)
                except Exception: pass
                # City
                try: _set_value(_get_edit("tbCity"), city)
                except Exception: pass
                # State
                _select_state(state)
                # Zip
                try: _set_value(_get_edit("tbZip"), zip_)
                except Exception: pass

            def _clear_all_inputs():
                for eid in ("tbFirst","tbLast","tbFacility","tbPhone","tbCity","tbZip"):
                    try:
                        _set_value(_get_edit(eid), "")
                    except Exception:
                        pass
                _select_state(None)

            def _search_with(first=None, last=None, facility=None,
                            phone=None, city=None, state=None, zip_=None):
                """Populate fields, click Search, wait, return count."""
                
                # Names
                if first is not None:
                    try: _set_value(_get_edit("tbFirst"), first)
                    except Exception: pass
                if last is not None:
                    try: 
                        _set_value(_get_edit("tbLast"), last)
                        try:
                            _set_value(_get_edit("tbFacility"), "")
                        except Exception:
                            pass
                    except Exception: pass

                # Facility
                if facility is not None:
                    try: _set_value(_get_edit("tbFacility"), facility)
                    except Exception: pass
                # Common filters
                _fill_common_filters(phone, city, state, zip_)
                # Search
                _click_search()
                _wait_search_idle()

                return _results_count()
            
            
            
            
            rows_found,NewResultName,NewResultAddr,NewResultPhone = run_provider_search_part()
            btn = ups.child_window(auto_id="btnClose", control_type="Button").wrapper_object()
            safe_click(btn)

            tempAddr1 = ""
            tempAddr2 = ""
            tempCity = ""
            tempState = ""
            tempZip = ""
            # providerPhone = ""
            if NewResultAddr:
                tempAddr1,tempAddr2,tempCity,tempState,tempZip = parse_address(NewResultAddr)

            # if not allData['providerAddr'] and (tempAddr1 or tempAddr2):
            #     allData['providerAddr'] = f"{tempAddr1} {tempAddr2}"
            # if not allData['providerCity'] and tempCity:
            #     allData['providerCity'] = tempCity
            # if not allData['providerState'] and tempState:
            #     allData['providerState'] = tempState
            # if not allData['providerZip'] and tempZip:
            #     allData['providerZip'] = tempZip
            
            # if not allData['providerPhone'] and NewResultPhone:
            #     allData['providerPhone'] = NewResultPhone
            focus_control(dlg)
            btnCloseSL = dlg.child_window(title="Close", control_type = "Button")
            btnCloseSL.wait("exists enabled visible ready",timeout=120,retry_interval=1)
            btnCloseSL.invoke()
            # btnCloseSL.click_input()

                
            if missing:
                # Only attempt Google if the missing set is something Google can help with
                google_fixable = {
                    "Provider address",
                    "Provider city",
                    "Provider state",
                    "Provider ZIP",
                    # "Provider phone",
                }
                needs_google = any(m in google_fixable for m in missing)
            
                if needs_google:
                    try:
                        # from GoogleSearchV2OLD import find_provider_address  # <-- filename from step 1
                        from legacy.legacy_googlesearch import find_provider_address
                    except Exception as e:
                        notify("Google Import Error", f"Cannot import google_provider_finder.find_provider_address\n\n{e}")
                        botStop = True
                        sys.exit()

                    # run google picker (user selects a result)
                    #picked = find_provider_address(PROVIDERSRCH)
            
                    # apply picked values back into allData
                    #apply_google_provider_result(allData, picked)
                    # REQUIRED by you: name the input PROVIDERSRCH

                    #PROVIDERSRCH = build_provider_search_text(allData)
            
                    # run google picker (user selects a result)
                    # -------------------------
                    # Test1: If validation missing provider details, call Google, then continue if fixed
                    # Minimal patch: if user clicks "No Accurate Result", DO NOT stop — proceed with remaining FCM code
                    # -------------------------
                    missing = get_validation_missing(allData)

                    if missing:
                        google_fixable = {
                            "Provider address",
                            "Provider city",
                            "Provider state",
                            "Provider ZIP",
                            # "Provider phone",
                        }
                        needs_google = any(m in google_fixable for m in missing)

                        if needs_google:
                            try:
                                from legacy.legacy_googlesearch import find_provider_address
                            except Exception as e:
                                notify("Google Import Error", f"Cannot import GoogleSearchV2.find_provider_address\n\n{e}")
                                botStop = True
                                sys.exit()

                            PROVIDERSRCH = build_provider_search_text(allData)

                            picked = find_provider_address(PROVIDERSRCH)
                            pick_err = (picked.get("error") or "").strip().lower()

                            # If user explicitly chooses "No Accurate Result" -> DO NOT stop, just continue.
                            if pick_err == "no_accurate_result":
                                print("[Google] User chose NO ACCURATE RESULT -> continuing without stopping.")
                            else:
                                # Apply Google result (only fills non-empty)
                                apply_google_provider_result(allData, picked)

                                # If user cancelled / blocked / exception -> stop as usual
                                if pick_err in ("cancelled", "google_blocked") or pick_err.startswith("exception"):
                                    ValidateInfo(allData)
                                    botStop = True
                                    sys.exit()

                                # Re-check after Google
                                missing_after = get_validation_missing(allData)
                                if missing_after:
                                    ValidateInfo(allData)
                                    botStop = True
                                    sys.exit()

                        else:
                            # Missing items not fixable by Google -> stop as usual
                            ValidateInfo(allData)
                            botStop = True
                            sys.exit()
                else:
                    # Missing items not fixable by Google (e.g. name/date/time rule) -> stop as usual
                    ValidateInfo(allData)
                    botStop = True
                    sys.exit()
            
            if not ValidateInfo(allData):
            # notify("Notice","For Manual Process. Missing critical information.") #migz
                botStop=True
                sys.exit()

        # pprint(allData)
        correctZipCode = allData['providerZip']
        ######
        #Insert CMS Reopen Checker
        print(f'Checking for Cases in CMS')
        ReopenResults = reopen_mod.MainReopenCheck(allData['claimNumber'])
        # pprint(ReopenResults)
        closedCases = []
        closedCases.clear()
        GYCTCasesLM = []
        GYCTCasesLM.clear()
        for r in ReopenResults:
            if r.get("case_status") == "C" or r.get("case_status") == "O":
                cmsCase = r.get("cms_caseNum")
                reopenMSG = r.get("message")
                caseType = r.get("caseType")
                closedCases.append((cmsCase,caseType,reopenMSG))
            if r.get("caseType") == "TCM" and r.get("case_status") == "O":
                cmsCase = r.get("cms_caseNum")
                reopenMSG = r.get("message")
                caseType = r.get("caseType")
                GYCTCasesLM.append((cmsCase,caseType,reopenMSG))
        # print(len(GYCTCasesLM))
        if len(GYCTCasesLM) > 0:
            if "goodyear tire" in allData['customer'].lower() or "cooper tire" in allData['customer'].lower():
                if "full case management" in allData['referralType'].lower() or "one-time rn visit" in allData['referralType'].lower():
                    notify("Goodyear/Cooper Tire Instruction For Open TCM","For Manual Process Bot Stop. Send an email to Jen.Herbert@enlyte.com and richard.castellini@enlyte.com\nRequest assistance to have the Full Med referral cancelled and resubmitted by Liberty as a One-Time RN Visit Provider")
                    sys.exit()

        if closedCases:
            lines = []
            for caseNum,caseTypes,Msg in closedCases:
                lines.append(f"{caseNum}\n{caseTypes}\n{Msg}")
            full_msg = "\n\n".join(lines)
            notify("Open/Closed Cases",full_msg)
            #--------------------ADD STOPPER!-------------------------
            def askUserContinue():
                import tkinter as tk
                from tkinter import messagebox
                root = tk.Tk()
                root.withdraw()
                root.attributes('-topmost',True)
                answer = messagebox.askyesno("User Confirmation","Select Yes if bot would proceed. \nSelect No if bot to stop process.")
                if answer:
                    return False
                else:
                    return True
                
            botStop = askUserContinue()
            if botStop:
                sys.exit()
            #--------------------------------------------------
        else:
            notify("Case Checker - CMS","No Open or Closed Cases Found in CMS")

        
        #Add Customer Checker
        
        # import CustomerCheckerV2
        print(f'Searching for Valid Employer Name in CMS')
        validCustomer = customer_mod.MainCustomerCheck(allData['customer'],allData['claimID'],allData['claimantFull'],app=app)
        if validCustomer:
            allData['customer'] = validCustomer
            print(f'Valid Employer: {validCustomer}')
            


        ######
        focus_control(emaiDisplay)
        print(f'Navigating to Subject Line Template Builder')
        SubjectLineBuilder = emaiDisplay.child_window(auto_id="btnSubjectSearch",control_type="Button").wrapper_object()
        focus_control(SubjectLineBuilder)
        safe_click(SubjectLineBuilder)
        # dlg.wait(timeout=3)

        # dlg = Desktop(backend="uia").window(auto_id="frmEmailDispaySubjectLineBuilder",control_type="Window")
        dlg = Desktop(backend="uia").window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
        dlg.wait("exists enabled visible ready",timeout=120,retry_interval=1)
        focus_control(dlg)

        # dlg.print_control_identifiers()
        tbCustomer = dlg.child_window(auto_id="tbCustomer", control_type="Edit").wrapper_object()
        tbCustomer.iface_value.SetValue("Liberty Mutual Commercial Market")

        
        # tbEmployer
        tbEmployer = dlg.child_window(auto_id="tbEmployer", control_type="Edit").wrapper_object()
        focus_control(tbEmployer)
        time.sleep(6)
        tbEmployer.iface_value.SetValue(allData['customer'])
        # # tbClaimantFirst
        tbClaimantFirst = dlg.child_window(auto_id="tbClaimantFirst", control_type="Edit").wrapper_object()
        tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
        
        # # tbClaimantLast
        tbClaimantLast = dlg.child_window(auto_id="tbClaimantLast", control_type="Edit").wrapper_object()
        tbClaimantLast.iface_value.SetValue(allData['claimantLast'])
        # checkFName = tbClaimantFirst.iface_value.CurrentValue
        # checkLName = tbClaimantLast.iface_value.CurrentValue
        # if checkFName:
        #     data['claimantFirst']  = checkFNameadjus
        # if checkLName:
        #     data['claimantLast'] =  checkLName

        # pbClaimant (Pane)
        pbClaimant = dlg.child_window(auto_id="pbClaimant",control_type="Pane").wrapper_object()

        #checker
        desk = Desktop(backend="uia")
        before_handles = {w.handle for w in desk.windows(top_level_only=True,visible_only=False)}

        focus_control(pbClaimant)
        safe_click(pbClaimant)
        print(f'Searching for Claimant')
        owner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
        owner.wait('exists',timeout=0.5)
        ucc = owner.child_window(auto_id="frmUnityClaimClaimant",control_type="Window")
        ucc.wait('exists',timeout=3)
        focus_control(ucc)


        #Add new logic here.
        def SearchForClaimant(FName,LName,DOB,ClaimNumber):
            # tbDOB
            # tbFirst
            # tbLast
            # tbClaim
            tbDOB = ucc.child_window(auto_id="tbDOB", control_type="Edit").wrapper_object()
            tbDOB.iface_value.SetValue(DOB)

            tbFirst = ucc.child_window(auto_id="tbFirst", control_type="Edit").wrapper_object()
            tbFirst.iface_value.SetValue(FName)

            tbLast = ucc.child_window(auto_id="tbLast", control_type="Edit").wrapper_object()
            tbLast.iface_value.SetValue(LName)

            tbClaim = ucc.child_window(auto_id="tbClaim", control_type="Edit").wrapper_object()
            tbClaim.iface_value.SetValue(ClaimNumber)

            SearchClaimant = ucc.child_window(auto_id="btnSearch",control_type="Button").wrapper_object()
            safe_click(SearchClaimant)

            lblStatusChecker = ucc.child_window(auto_id="lblStatus",control_type="Text")
            timeout = 300.0
            EndTime = time.monotonic() + timeout
            while time.monotonic() < EndTime:
                try:
                    lblStatus = lblStatusChecker.wrapper_object()
                    if (not lblStatus.is_visible()) or ((lblStatus.window_text() or "").strip()==""):
                        break
                except Exception:
                    break

            grid = ucc.child_window(auto_id="dgResults",control_type="Table").wrapper_object()
            AttachRow = grid.iface_grid.CurrentRowCount

            if AttachRow == 1:
                return True,"E"
            elif AttachRow > 1:
                return True,"G"
            elif AttachRow == 0:
                return False,"Z"


        def CheckClaimantResult():
            grid = ucc.child_window(auto_id="dgResults",control_type="Table")#.wrapper_object()
            AttachRow = grid.iface_grid.CurrentRowCount
            ClaimantSelected = False

            # el
            if AttachRow >= 1:
                # ---------- 1) Count EDI rows first ----------
                edi_rows = []
                for r in range(AttachRow):
                    # cell_el = grid.iface_grid.GetItem(r, 9)
                    # cell = UIAWrapper(UIAElementInfo(cell_el))
                    cell = grid.child_window(title=f"Source Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                    src = (cell.iface_value.CurrentValue or "").strip().upper()
                    if src == "EDI":
                        edi_rows.append(r)
            
                # ---------- 2) If multiple EDI -> notify, do NOT auto-pick ----------
                if len(edi_rows) > 1:
                    
                    notify("Notice", "Multiple Results! Kindly select appropriate Claimant\nClick Ok once Claim/Claimant is set")
            
                # ---------- 3) If exactly one EDI -> apply your match logic ----------
                elif len(edi_rows) == 1:
                    for r in range(AttachRow):
                        # cell_el = grid.iface_grid.GetItem(r, 6)
                        # cell = UIAWrapper(UIAElementInfo(cell_el))
                        cell = grid.child_window(title=f"DOB Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                        DOB = to_mmddyyyy(cell.iface_value.CurrentValue)
            
                        # cell_el = grid.iface_grid.GetItem(r, 9)
                        # cell = UIAWrapper(UIAElementInfo(cell_el))
                        cell = grid.child_window(title=f"Source Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                        Source = (cell.iface_value.CurrentValue or "").strip().upper()
            
                        # cell_el = grid.iface_grid.GetItem(r, 10)
                        # cell = UIAWrapper(UIAElementInfo(cell_el))
                        cell = grid.child_window(title=f"DOI Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                        DateOfInj = to_mmddyyyy(cell.iface_value.CurrentValue)
            
                        # cell_el = grid.iface_grid.GetItem(r, 11)
                        # cell = UIAWrapper(UIAElementInfo(cell_el))
                        cell = grid.child_window(title=f"LOI Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                        LineOfInsurance = (cell.iface_value.CurrentValue or "").replace(" ", "").replace("'", "").strip()
                        
                        # print(f"{allData.get('dob')} {allData.get('doi')} {(allData.get('claimType', '').replace(' ', '').lower())} {LineOfInsurance}")
                        if (Source == "EDI" and (allData.get("dob") == DOB) and (allData.get("doi") == DateOfInj) and (allData.get("claimType", "").replace(" ", "").lower() == LineOfInsurance.lower())):
                            # cell_el = grid.iface_grid.GetItem(r, 0)
                            # cell = UIAWrapper(UIAElementInfo(cell_el))
                            cell = grid.child_window(title=f"Select Row {r}", control_type="CheckBox").wrapper_object()

                            safe_click(cell)
                            ClaimantSelected = True
                            SelectClaimant = ucc.child_window(auto_id="btnSelect", control_type="Button").wrapper_object()
                            safe_click(SelectClaimant)
                            time.sleep(3)
                            return ClaimantSelected
                            break
                else:
                    SelectClaimant = ucc.child_window(auto_id="btnClose",control_type="Button").wrapper_object()
                    safe_click(SelectClaimant)
                    return ClaimantSelected

        SearchIndex = 1  
        WithPossibleSearchResult = False 
        ClaimantSelected = False
        for SearchIndex in range(1,4):
            if SearchIndex == 1:
                WithPossibleSearchResult,CountText = SearchForClaimant(None,None,allData['dob'],allData['claimNumber'])
                if WithPossibleSearchResult:
                    ClaimantSelected = CheckClaimantResult()
                    break
            elif SearchIndex == 2:
                WithPossibleSearchResult,CountText = SearchForClaimant(None,allData['claimantLast'][:3],None,allData['claimNumber'])
                if WithPossibleSearchResult:
                    ClaimantSelected = CheckClaimantResult()
                    break
            elif SearchIndex == 3:
                WithPossibleSearchResult,CountText = SearchForClaimant(allData['claimantFirst'][:3],allData['claimantLast'][:3],None,allData['claimNumber'])
                if WithPossibleSearchResult:
                    ClaimantSelected = CheckClaimantResult()
                    break

        try:
            if not ClaimantSelected or not WithPossibleSearchResult:
                SelectClaimant = ucc.child_window(auto_id="btnClose",control_type="Button").wrapper_object()
                safe_click(SelectClaimant)
        except:
            pass

        # tbClaimantFirst = dlg.child_window(auto_id="tbClaimantFirst", control_type="Edit").wrapper_object()
        # ClaimantFirstVal = (tbClaimantFirst.iface_value.CurrentValue or "").strip()
        # if ClaimantFirstVal and ClaimantSelected:
        #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
        #     allData['claimantFirst'] = ClaimantFirstVal

        # tbClaimantLast = dlg.child_window(auto_id="tbClaimantLast", control_type="Edit").wrapper_object()
        # tbClaimantLastVal = (tbClaimantLast.iface_value.CurrentValue or "").strip()
        # if tbClaimantLastVal and ClaimantSelected:
        #     # tbClaimantLast.iface_value.SetValue(allData['claimantLast'])
        #     allData['claimantLast'] = tbClaimantLastVal
        #Like item Search
        focus_control(dlg)
        print(f'Navigating to Like Item Search')
        time.sleep(3)
        btnLikeItemSearch = dlg.child_window(auto_id="btnLikeItemSearch",control_type="Button").wrapper_object()
        safe_click(btnLikeItemSearch)

        #frmLikeItemsSearch
        owner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
        owner.wait('exists',timeout=0.5)
        lis = owner.child_window(auto_id="frmLikeItemsSearch",control_type="Window")
        lis.wait('exists',timeout=3)
        time.sleep(3)
        #--------------------ADD STOPPER!-------------------------
        def askUserContinueLikeItemSearch():
            import tkinter as tk
            from tkinter import messagebox
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost',True)
            answer = messagebox.askyesno("User Confirmation","For Manual Process. Kindly validate result from Like Item Search\nSelect Yes if bot would proceed. \nSelect No if bot to stop process.")
            if answer:
                return False
            else:
                return True
        #-----------------------------------------------------------

        def LikeItemSearch(SearchCriteria : str):

            tbCriteria = lis.child_window(auto_id="tbCriteria", control_type="Edit").wrapper_object()
            tbCriteria.iface_value.SetValue(SearchCriteria)

            btnSearch = lis.child_window(auto_id="btnSearch", control_type="Button").wrapper_object()
            safe_click(btnSearch)
            time.sleep(3)

            lblStatusChecker = lis.child_window(auto_id="lblNoItem",control_type="Text")
            timeout = 300.0
            EndTime = time.monotonic() + timeout
            while time.monotonic() < EndTime:
                try:
                    lblStatus = lblStatusChecker.wrapper_object()
                    if (lblStatus.is_visible()) or ((lblStatus.window_text() == "No Items Returned For Search Criteria")):
                        break
                except Exception:
                    break

            grid = lis.child_window(auto_id="dgSearch",control_type="Table").wrapper_object()
            AttachRow = grid.iface_grid.CurrentRowCount
            # for r in range(0,AttachRow):
            if AttachRow >= 1:   
                # notify("Notice","For Manual Process. Kindly validate result from Like Item Search")  
                botStop = askUserContinueLikeItemSearch()
                if botStop:
                    sys.exit() 
            else:
                return 0
                # break

            # lblNoItem = lis.child_window(auto_id="lblNoItem", control_type="Text").wrapper_object()
            # "No Items Returned For Search Criteria"
        lisCount = LikeItemSearch(allData['claimNumber']) 
        if lisCount== 0:
            lisCount = LikeItemSearch(f"{allData['claimantFirst']} {allData['claimantLast']}")

        CloseWindowIns = lis.child_window(auto_id="btnClose",control_type="Button").wrapper_object()
        safe_click(CloseWindowIns)
        time.sleep(3)
        send_keys('{SPACE}')
        time.sleep(3)

        focus_control(dlg)
        time.sleep(3)



        # drpClaimJuris = dlg.child_window(title="LOADING PLEASE WAIT", control_type="Edit").wrapper_object()
        # drpClaimJurisVal = drpClaimJuris.iface_value.CurrentValue
        # if not drpClaimJurisVal:
        #     drpClaimJuris.iface_value.SetValue(allData['claimStateFull'])
        # # tbQue
        # tbQue = dlg.child_window(auto_id="tbQue", control_type="Edit").wrapper_object()
        # tbQue.iface_value.SetValue(allData['referralNumber']) 

        # tbClaimIdentifier = dlg.child_window(auto_id="tbClaimIdentifier", control_type="Edit").wrapper_object()
        # checktbClaimIdentifier = tbClaimIdentifier.iface_value.CurrentValue
        # if not checktbClaimIdentifier:
        #     tbClaimIdentifier.iface_value.SetValue(allData['claimID'])

        # tbClaimNumber = dlg.child_window(auto_id="tbClaimNumber", control_type="Edit").wrapper_object()

        # tbClaimNumber.set_focus()
        # send_keys('{TAB}')
        # send_keys('FULL MED' ,with_spaces=True)

        # # pbRefType (Pane)
        # pbRefType = dlg.child_window(auto_id="pbRefType",control_type="Pane").wrapper_object()
        # pbRefType.click_input()

        # owner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
        # owner.wait('exists',timeout=0.5)
        # uft = owner.child_window(auto_id="frmUnityReferralTypes",control_type="Window")
        # uft.wait('exists',timeout=0.5)
        # uft.set_focus()
        # # uft.print_control_identifiers()
        # # cb1 = uft.child_window(control_type="Edit",found_index=0).wrapper_object()

        # # "ServiceType":URTServiceType,
        # # "refType":URTrefType,
        # # "CaseObj":URTCaseObjective
        
        # cb1 = uft.child_window(auto_id="cbServiceType",control_type="ComboBox").wrapper_object()
        # # cb1.select("Medical Full") 
        # cb1.select(allData['ServiceType'])
        # time.sleep(.5)
        # cb1 = uft.child_window(auto_id="cbReferralType",control_type="ComboBox").wrapper_object()
        # # cb1.select("Medical") 
        # cb1.select(allData['refType']) 
        # time.sleep(.5)
        # cb1 = uft.child_window(auto_id="cbUnityObjectives",control_type="ComboBox").wrapper_object()
        # # cb1.select("Coordination of Care/Services") 
        # cb1.select(allData['CaseObj']) 
        # btnOK = uft.child_window(auto_id="btnOK",control_type="Button").wrapper_object()
        # btnOK.click_input()


        dlg = Desktop(backend="uia").window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
        dlg.wait("exists enabled visible ready",timeout=15,retry_interval=1)
        focus_control(dlg)

        tbClaimNumber = dlg.child_window(auto_id="tbClaimNumber", control_type="Edit").wrapper_object()
        focus_control(tbClaimNumber)
        #LOI

        # cbLOI = dlg.child_window(auto_id="cbLOI",control_type="ComboBox").wrapper_object()
        # cbLOIVal = cbLOI.iface_value.CurrentValue
        # if not cbLOIVal:
        #     # cbLOI.iface_value.SetValue(allData['claimType'])
        #     cbLOI.select(allData['claimType']) 
        # send_keys(allData['claimType'],with_spaces=True)
        send_keys('{TAB}')
        # rbClaimant
        rbClaimant = dlg.child_window(auto_id="rbClaimant",control_type="RadioButton").wrapper_object()
        rbClaimant.select()

        if ClaimantSelected == False:
            print(f'Populating Claimant Info')
            tbClaimNumber.iface_value.SetValue(allData['claimNumber'])
        # tbAddress1
            allData['addressLine1'] = titlecase(allData['addressLine1'], callback=address_callback)
            allData['addressLine2'] = titlecase(allData['addressLine2'], callback=address_callback)
            allData['city'] = titlecase(allData['city'], callback=address_callback)
            allData['state'] = titlecase(allData['state'], callback=address_callback)

            tbAddress1 = dlg.child_window(auto_id="tbAddress1", control_type="Edit").wrapper_object()
            tbAddress1.iface_value.SetValue(allData['addressLine1'])
            # tbAddress2
            tbAddress2 = dlg.child_window(auto_id="tbAddress2", control_type="Edit").wrapper_object()
            tbAddress2.iface_value.SetValue(allData['addressLine2'])
            # tbCity
            tbCity = dlg.child_window(auto_id="tbCity", control_type="Edit").wrapper_object()
            tbCity.iface_value.SetValue(allData['city'])
            # TAB enter for State
            send_keys('{TAB}')
            send_keys(allData['state'])
            # tbZip
            tbZip = dlg.child_window(auto_id="tbZip", control_type="Edit").wrapper_object()
            tbZip.iface_value.SetValue(allData['zip'])
            # tbPhone
            tbPhone = dlg.child_window(auto_id="tbPhone", control_type="Edit").wrapper_object()
            tbPhone.iface_value.SetValue(allData['phoneNumber'])
        # Gender (rbMale, rbFemale, rbUnknown)
        if allData['gender'].lower() == "female":
            rbGender = dlg.child_window(auto_id="rbFemale", control_type="RadioButton").wrapper_object()
        elif allData['gender'].lower() == "male":
            rbGender = dlg.child_window(auto_id="rbMale", control_type="RadioButton").wrapper_object()
        else:
            rbGender = dlg.child_window(auto_id="rbUnknown", control_type="RadioButton").wrapper_object()
        rbGender.select()
        # tbDOB
        tbDOB = dlg.child_window(auto_id="tbDOB", control_type="Edit").wrapper_object()
        tbDOB.iface_value.SetValue(allData['dob'])

        #######################################
        #Create If Statement for Case Manager#
        if allData['ncmContactName']:
        # #cbReqNCM
            cbReqNCM = dlg.child_window(auto_id="cbReqNCM", control_type="CheckBox").wrapper_object()
            cbReqNCM.iface_toggle.Toggle()
            # tbReqNcm
            tbReqNcm = dlg.child_window(auto_id="tbReqNcm", control_type="Edit").wrapper_object()
            tbReqNcm.type_keys(allData['ncmContactName'],with_spaces=True)
            time.sleep(3)
            send_keys('{DOWN}')
            send_keys('{ENTER}')
            # rbCustomer
            rbCustomer = dlg.child_window(auto_id="rbCustomer", control_type="RadioButton").wrapper_object()
            rbCustomer.select()

        #### GAA
        if allData['city'].lower() == "texarkana" and "goodyear" in allData['customer'].lower():
            cbGAAReq = dlg.child_window(auto_id="cbGAAReq", control_type="CheckBox").wrapper_object()
            cbGAAReq.iface_toggle.Toggle()

        ########################################

        
        # cbAppointment
        print(f'Populating Appointment Details')
        cbAppointment = dlg.child_window(auto_id="cbAppointment", control_type="CheckBox").wrapper_object()

        if cbAppointment.get_toggle_state() == 0:
            if allData['nextApptDate'] and allData['nextApptTime']:
                
                cbAppointment.iface_toggle.Toggle()
                cbDateOnly = dlg.child_window(auto_id="cbDateOnly", control_type="CheckBox").wrapper_object()
                cbDateOnly.iface_toggle.Toggle()
                cbDateOnly.iface_toggle.Toggle()
                # DatePicker
                DatePicker = dlg.child_window(auto_id="DatePicker", control_type="ComboBox").wrapper_object()
                DatePicker.iface_value.SetValue(allData['nextApptDate'])
                # DatePicker.set_focus()
                DatePicker.type_keys(allData['nextApptDate'],with_spaces=True)
                # TimePicker
                TimePicker = dlg.child_window(auto_id="TimePicker", control_type="ComboBox").wrapper_object()
                # print([c.friendly_class_name() for c in TimePicker.children()])
                # print([(d.control_type(),d.window_text()) for d in TimePicker.descendants()][:20])
                
                time_str = allData['nextApptTime']
                time_24 = datetime.strptime(time_str,'%I:%M %p').strftime('%H:%M')

                # TimePicker.type_keys(time_24,with_spaces=True)
                focus_control(TimePicker)
                send_keys(time_24)
            elif allData['nextApptDate'] and not allData['nextApptTime']:
                # cbAppointment = dlg.child_window(auto_id="cbAppointment", control_type="CheckBox").wrapper_object()
                cbAppointment.iface_toggle.Toggle()
                cbDateOnly = dlg.child_window(auto_id="cbDateOnly", control_type="CheckBox").wrapper_object()
                cbDateOnly.iface_toggle.Toggle()
                # DatePicker
                DatePicker = dlg.child_window(auto_id="DatePicker", control_type="ComboBox").wrapper_object()
                DatePicker.iface_value.SetValue(allData['nextApptDate'])
                # DatePicker.set_focus()
                DatePicker.type_keys(allData['nextApptDate'],with_spaces=True)
        # rbProvider
        print(f'Navigating to Provider Search')
        rbProvider = dlg.child_window(auto_id="rbProvider", control_type="RadioButton").wrapper_object()
        rbProvider.select()
        time.sleep(3)
        # pbProvider (Pane Magnifying Glass)
        pbProvider = dlg.child_window(auto_id="pbProvider",control_type="Pane").wrapper_object()
        safe_click(pbProvider)
        time.sleep(3)
        #ADD Searching here!
        owner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
        owner.wait('exists',timeout=10)
        ups = owner.child_window(auto_id="frmUnityProviderLM",control_type="Window")
        ups.wait('exists ready',timeout=10)
        # ups.print_control_identifiers()
        ###-----------------------------------------------------

        # ----------------- Inputs -----------------
        #"tbFirst","tbLast","tbFacility","tbPhone","tbCity","tbZip"
        # allData['providerLast'] = "Ansari"
        tbProviderLastExtract = ups.child_window(auto_id="tbProviderLastExtract", control_type="Edit").wrapper_object()
        tbProviderLastExtractVal = (tbProviderLastExtract.iface_value.CurrentValue or "").strip()

        tbLast = ups.child_window(auto_id="tbLast", control_type="Edit").wrapper_object()
        tbLastVal = (tbLast.iface_value.CurrentValue or "").strip()
        # if tbLastVal:
        #     allData['providerLast'] = tbLastVal
        # elif tbProviderLastExtractVal:
        #     allData['providerLast'] = tbProviderLastExtractVal
        
        tbProviderFirstExtract = ups.child_window(auto_id="tbProviderFirstExtract", control_type="Edit").wrapper_object()
        tbProviderFirstExtractVal = (tbProviderFirstExtract.iface_value.CurrentValue or "").strip()
        tbFirst = ups.child_window(auto_id="tbFirst", control_type="Edit").wrapper_object()
        tbFirstVal = (tbFirst.iface_value.CurrentValue or "").strip()
        # if tbFirstVal:
        #     allData['providerFirst'] = tbFirstVal
        # elif tbProviderFirstExtractVal:
        #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
        #     allData['providerFirst'] = tbProviderFirstExtractVal
        
        tbProviderNameExtract = ups.child_window(auto_id="tbProviderNameExtract", control_type="Edit").wrapper_object()
        tbProviderNameExtractVal = (tbProviderNameExtract.iface_value.CurrentValue or "").strip()
        tbFacility = ups.child_window(auto_id="tbFacility", control_type="Edit").wrapper_object()
        tbFacilityVal = (tbFacility.iface_value.CurrentValue or "").strip()
        # if tbFacilityVal:
        #     allData['providerName'] = tbFacilityVal
        # elif tbProviderNameExtractVal:
        #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
        #     allData['providerName'] = tbProviderNameExtractVal
        
        tbAddressExtract = ups.child_window(auto_id="tbAddressExtract", control_type="Edit").wrapper_object()
        tbAddressExtractVal = (tbAddressExtract.iface_value.CurrentValue or "").strip()
        # if tbAddressExtractVal:
        #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
        #     allData['providerAddr'] = tbAddressExtractVal
        
        tbPhoneExtract = ups.child_window(auto_id="tbPhoneExtract", control_type="Edit").wrapper_object()
        tbPhoneExtractVal = (tbPhoneExtract.iface_value.CurrentValue or "").strip()
        tbPhone = ups.child_window(auto_id="tbPhone", control_type="Edit").wrapper_object()
        tbPhoneVal = (tbPhone.iface_value.CurrentValue or "").strip()
        # if tbPhoneVal:
        #     allData['providerPhone'] = tbPhoneVal
        # elif tbPhoneExtractVal:
        #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
        #     allData['providerPhone'] = tbPhoneExtractVal
        
        tbCityExtract = ups.child_window(auto_id="tbCityExtract", control_type="Edit").wrapper_object()
        tbCityExtractVal = (tbCityExtract.iface_value.CurrentValue or "").strip()
        tbCity = ups.child_window(auto_id="tbCity", control_type="Edit").wrapper_object()
        tbCityVal = (tbCity.iface_value.CurrentValue or "").strip()
        # if tbCityVal:
        #     allData['providerCity'] = tbCityVal
        # elif tbCityExtractVal:
        #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
        #     allData['providerCity'] = tbCityExtractVal
        
        tbStateExtract = ups.child_window(auto_id="tbStateExtract", control_type="Edit").wrapper_object()
        tbStateExtractVal = (tbStateExtract.iface_value.CurrentValue or "").strip()
        cbState = ups.child_window(auto_id="cbState", control_type="ComboBox").wrapper_object()
        cbStateVal = (cbState.iface_value.CurrentValue or "").strip()
        # if cbStateVal:
        #     allData['providerState'] = cbStateVal
        # elif tbStateExtractVal:
        #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
        #     allData['providerState'] = tbStateExtractVal
        
        tbZipExtract = ups.child_window(auto_id="tbZipExtract", control_type="Edit").wrapper_object()
        tbZipExtractVal = (tbZipExtract.iface_value.CurrentValue or "").strip()
        tbZip = ups.child_window(auto_id="tbZip", control_type="Edit").wrapper_object()
        tbZipVal = (tbZip.iface_value.CurrentValue or "").strip()
        # if tbZipVal:
        #     allData['providerZip'] = tbZipVal
        # elif tbZipExtractVal:
        #     # tbClaimantFirst.iface_value.SetValue(allData['claimantFirst'])
        #     allData['providerZip'] = tbZipExtractVal
        
        # pprint(allData)

        FacilityLName = allData['providerLast']
        FacilityFName = allData['providerFirst']
        FacilityName  = allData['providerName']
        FacPhoneNumber = allData['providerPhone']
        FacCity = allData['providerCity']
        FacState = allData['providerState']
        FacZip = allData['providerZip']

        # print("--------START Unity Provider------------------")
        # ups.print_control_identifiers()
        # print("--------END Unity Provider------------------")

        # ----------------- Helpers -----------------
        def _get_edit(auto_id):
            return ups.child_window(auto_id=auto_id, control_type="Edit").wrapper_object()

        def _set_value(ctrl, value):
            try:
                ctrl.iface_value.SetValue("")   # clear
                if value:
                    ctrl.iface_value.SetValue(value)

            except Exception:
                # Fallback to type_keys if needed
                try:
                    focus_control(ctrl)
                    ctrl.type_keys("^a{BACKSPACE}")
                    if value:
                        ctrl.type_keys(value, with_spaces=True)
                except Exception:
                    pass

        def _select_state(value):
            try:
                cb = ups.child_window(auto_id="cbState", control_type="ComboBox").wrapper_object()
                if value:
                    cb.select(value)
                else:
                    try:
                        cb.select(0)
                    except Exception:
                        pass
            except Exception:
                pass

        def _click_search():
            btn = ups.child_window(auto_id="btnSearch", control_type="Button").wrapper_object()
            safe_click(btn)
            time.sleep(3)

        def _wait_search_idle(timeout=300.0):
            lbl = ups.child_window(auto_id="lblStatus", control_type="Text")
            end_t = time.monotonic() + timeout
            while time.monotonic() < end_t:
                try:
                    w = lbl.wrapper_object()
                    if (not w.is_visible()) or ((w.window_text() or "").strip() == ""):
                        return True
                except Exception:
                    return True
                time.sleep(0.2)
            return False

        def _results_count():
            try:
                grid = ups.child_window(auto_id="dgResults", control_type="Table").wrapper_object()
                return int(getattr(grid.iface_grid, "CurrentRowCount", 0) or 0)
            except Exception:
                return 0
            
        def _normalize_name(value: str) -> str:
            if not value:
                return ""
            value = value.upper().strip()
            value = re.sub(r'[^A-Z0-9\s]', ' ', value)
            value = re.sub(r'\s+', ' ', value).strip()
            return value
        
        def _facility_name_strong_match(facility_name: str, result_name: str) -> bool:
            fac = _normalize_name(facility_name)
            res = _normalize_name(result_name)
        
            if not fac or not res:
                return False
        
            # exact normalized match
            if fac == res:
                return True
        
            return False
        
        def _provider_name_strong_match(prov_first: str, prov_last: str, result_name: str) -> bool:
            first = _normalize_name(prov_first)
            last = _normalize_name(prov_last)
            res = _normalize_name(result_name)
        
            if not first or not last or not res:
                return False
        
            full1 = f"{first} {last}"
            full2 = f"{last} {first}"
            full3 = f"{last}, {first}"
        
            if res == full1 or res == full2 or res == full3:
                return True
        
            return False
        
        def _facility_name_loose_match(facility_name: str, result_name: str) -> bool:
            fac = _normalize_name(facility_name)
            res = _normalize_name(result_name)
        
            if not fac or not res:
                return False
        
            return fac in res or res in fac
        
        def _provider_name_loose_match(prov_first: str, prov_last: str, result_name: str) -> bool:
            first = _normalize_name(prov_first)
            last = _normalize_name(prov_last)
            res = _normalize_name(result_name)
        
            # if not first or not last or not res:
            #     return False
        
            return first in res and last in res
        
        def _click_select_row(grid, row_index):
            cell = grid.child_window(title=f"Select Row {row_index}", control_type="CheckBox").wrapper_object()
            safe_click(cell)

        def prvGoogleSearch(prvName):
            try:
                from legacy.legacy_googlesearch import find_provider_address
            except Exception as e:
                notify("Google Import Error", f"Cannot import GoogleSearchV2.find_provider_address\n\n{e}")
                botStop = True
                sys.exit()
            checkData = {}
            checkData["providerName"] = prvName
            checkData["providerAddr"] = allData["providerAddr"]
            checkData["providerCity"] = allData["providerCity"]
            checkData["providerState"] = allData["providerState"]
            checkData["providerZip"] = allData["providerZip"]
            checkData["providerPhone"] = allData["providerPhone"]

            PROVIDERSRCH = build_provider_search_text(checkData)
            print(PROVIDERSRCH)
            picked = find_provider_address(PROVIDERSRCH)
            pprint(picked)
            pick_err = (picked.get("error") or "").strip().lower()

            # If user explicitly chooses "No Accurate Result" -> DO NOT stop, just continue.
            if pick_err == "no_accurate_result":
                print("[Google] User chose NO ACCURATE RESULT -> continuing without stopping.")
            # else:
                # Apply Google result (only fills non-empty)
                # apply_google_provider_result(checkData, picked)

                # If user cancelled / blocked / exception -> stop as usual
                # if pick_err in ("cancelled", "google_blocked") or pick_err.startswith("exception"):
                    # ValidateInfo(allData)
                    # botStop = True
                    # sys.exit()


            return picked["providerName"]
                # Re-check after Google
                # missing_after = get_validation_missing(allData)
        
        



        def _fill_common_filters(phone, city, state, zip_):
            # Phone
            try: _set_value(_get_edit("tbPhone"), phone)
            except Exception: pass
            # City
            try: _set_value(_get_edit("tbCity"), city)
            except Exception: pass
            # State
            _select_state(state)
            # Zip
            try: _set_value(_get_edit("tbZip"), zip_)
            except Exception: pass

        def _clear_all_inputs():
            for eid in ("tbFirst","tbLast","tbFacility","tbPhone","tbCity","tbZip"):
                try:
                    _set_value(_get_edit(eid), "")
                except Exception:
                    pass
            _select_state(None)

        def _search_with(first=None, last=None, facility=None,
                        phone=None, city=None, state=None, zip_=None):
            """Populate fields, click Search, wait, return count."""
            _clear_all_inputs()
            # Names
            if first is not None:
                try: _set_value(_get_edit("tbFirst"), first)
                except Exception: pass
            if last is not None:
                try:
                    _set_value(_get_edit("tbLast"), last)
                    try:
                        _set_value(_get_edit("tbFacility"), "")
                    except Exception:
                        pass
                except Exception: pass

            # Facility
            if facility is not None:
                try: 
                    _set_value(_get_edit("tbFacility"), facility)
                    
                except Exception: pass
            # Common filters
            _fill_common_filters(phone, city, state, zip_)
            # Search
            _click_search()
            _wait_search_idle()

            return _results_count()

        def _close_dialog_if_needed():
            try:
                btn = ups.child_window(auto_id="btnClose", control_type="Button").wrapper_object()
                safe_click(btn)
            except Exception:
                pass

        # ----------------- Main search logic -----------------
        
        # ----------------- Execute -----------------
        rows_found = 0
        SelectedPrvResult = False
        rows_found,SelectedPrvResult = run_provider_search()
        # print(f"Rows found: {rows_found}")
        if not isManualEntered:
            print(f'Adding new provider information')
            # if rows_found == 0:
            btnAddNew= ups.child_window(auto_id="btnAddNew",control_type="Button").wrapper_object()
            safe_click(btnAddNew)
            time.sleep(2)
            upan = ups.child_window(auto_id="frmUnityProviderAddNew",control_type="Window")
            upan.wait('exists',timeout=5)
            # upan.print_control_identifiers()
            time.sleep(2)
            if allData['providerName'] and (not allData['providerLast'].strip() or not allData['providerFirst'].strip()):
                tbFacilityName = upan.child_window(auto_id="tbFacilityName", control_type="Edit").wrapper_object()
                tbFacilityName.iface_value.SetValue(allData['providerName'])
            ProfName = f"{allData['providerFirst']} {allData['providerLast']}"
            if ProfName: #ProfName.strip() != allData['providerName'].strip()
                tbName = upan.child_window(auto_id="tbName", control_type="Edit").wrapper_object()
                tbName.iface_value.SetValue(ProfName)
                
            tbAddress1= upan.child_window(auto_id="tbAddress1", control_type="Edit").wrapper_object()
            tbAddress1.iface_value.SetValue(allData['providerAddr'])
            # tbAddress2= upan.child_window(auto_id="tbName", control_type="Edit").wrapper_object()
            # tbName.iface_value.SetValue(allData['providerName'])
            tbCity= upan.child_window(auto_id="tbCity", control_type="Edit").wrapper_object()
            tbCity.iface_value.SetValue(allData['providerCity'])
            cbState = upan.child_window(auto_id="cbState",control_type="ComboBox").wrapper_object()
            cbState.select(allData['providerState']) 
            tbZip= upan.child_window(auto_id="tbZip", control_type="Edit").wrapper_object()
            tbZip.iface_value.SetValue(allData['providerZip'])
            tbPhone= upan.child_window(auto_id="tbPhone", control_type="Edit").wrapper_object()
            tbPhone.iface_value.SetValue(allData['providerPhone'])

            notify("Notice","Please validate if provider information is correct. User to Manual Click on Add Provider.")
                # btnAddPrv = dlg.child_window(auto_id="btnAdd",control_type="Button").wrapper_object()
                # btnAddPrv.click_input()
                

            # else:
            #     btnSelect = ups.child_window(auto_id="btnSelect",control_type="Button").wrapper_object()
            #     btnSelect.click_input()
            #     _close_dialog_if_needed()
        ###------------------------------------------------------------------


        # tbDOIdt
        tbDOIdt = dlg.child_window(auto_id="tbDOIdt", control_type="Edit").wrapper_object()
        tbDOIdtVal = tbDOIdt.iface_value.CurrentValue
        if not tbDOIdtVal:
            tbDOIdt.iface_value.SetValue(allData['doi'])

        # tbSSN
        # tbSSN = dlg.child_window(auto_id="tbSSN", control_type="Edit").wrapper_object()
        # tbSSN.iface_value.SetValue(allData['claimantSSN'])
        # tbReferralSource
        tbReferralSource = dlg.child_window(auto_id="tbReferralSource", control_type="Edit").wrapper_object()
        tbReferralSource.iface_value.SetValue(allData['refSource'])
        # tbAdjuster
        print(f'Navigating to Adjust Search')
        AdjusterName = allData['adjuster']
        tbAdjuster = dlg.child_window(auto_id="tbAdjuster", control_type="Edit").wrapper_object()
        tbAdjuster.iface_value.SetValue(AdjusterName)
        # pbAdjusterSearch (Pane)
        pbAdjusterSearch = dlg.child_window(auto_id="pbAdjusterSearch",control_type="Pane").wrapper_object()
        safe_click(pbAdjusterSearch)
        
        owner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
        owner.wait('exists',timeout=5)
        uas = owner.child_window(auto_id="frmUnityAdjusterSearch",control_type="Window")
        uas.wait('exists',timeout=5)

        parts = AdjusterName.strip().split()
        # Initialize defaults
        AdjusterFirstName = ""
        AdjusterLastName = ""
        if len(parts) == 1:
            AdjusterFirstName = parts[0]
        elif len(parts) == 2:
            AdjusterFirstName, AdjusterLastName = parts
        else:
            # For names with middle name or initial
            AdjusterFirstName = parts[0]
            AdjusterLastName = parts[-1]

        SearchAdjuster = uas.child_window(auto_id="btnSearch",control_type="Button").wrapper_object()
        safe_click(SearchAdjuster)

        # lblStatus.Name ="Searching please wait"
        lblStatusChecker = uas.child_window(auto_id="lblProgress",control_type="Text")
        timeout = 300.0
        EndTime = time.monotonic() + timeout
        while time.monotonic() < EndTime:
            try:
                lblStatus = lblStatusChecker.wrapper_object()
                if (not lblStatus.is_visible()) or ((lblStatus.window_text() or "").strip()==""):
                    break
            except Exception:
                break


        grid = uas.child_window(auto_id="dgAdjusters",control_type="Table")#.wrapper_object()
        AttachRow = grid.iface_grid.CurrentRowCount

        # el
        AdjusterFName = ""
        AdjusterLName= ""
        if AttachRow >= 1:
            validateAdjuster = False
            for r in range(AttachRow):
                print(f'Validating Adjuster information. {r} of {AttachRow}')
                # cell_el = grid.iface_grid.GetItem(r,3)
                # cell = UIAWrapper(UIAElementInfo(cell_el))
                # adjusterPhone = cell.iface_value.CurrentValue

                # cell_el = grid.iface_grid.GetItem(r,5)
                # cell = UIAWrapper(UIAElementInfo(cell_el))
                cell = grid.child_window(title=f"Physical Address Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                adjPhyAddr = cell.iface_value.CurrentValue
                adjAddrFull = f"{allData['adjAddrFull']}"
                adjAddrMail = f"{allData['adjAddrMail']}"

                TempAddress = expand_suffix_long(adjPhyAddr)
                RealAddress = expand_suffix_long(adjAddrFull)
                ComparisonResult = addressMatch(TempAddress,RealAddress)

                TempAddress1 = expand_suffix_long(adjPhyAddr)
                RealAddress1 = expand_suffix_long(adjAddrMail)
                ComparisonResult1 = addressMatch(TempAddress1, RealAddress1)
                
                if ComparisonResult or ComparisonResult1:
                    # cell_el = grid.iface_grid.GetItem(r,0)
                    # cell = UIAWrapper(UIAElementInfo(cell_el))
                    
                    cell = grid.child_window(title=f"Select Row {r}", control_type="CheckBox").wrapper_object()
                    safe_click(cell)
                    time.sleep(2)
                    SelectAdjuster = uas.child_window(auto_id="btnSetAdjuster",control_type="Button").wrapper_object()
                    safe_click(SelectAdjuster)

                    validateAdjuster = True
                    break
            # if not validateAdjuster:
            #     notify("Notice",f"Multiple Results. Please select correct Adjuster Information from the Results.\nAdjuster Full Address:{allData['adjAddrFull']}\nClick Ok once Adjuster is selected and set.")

        elif AttachRow == 0:
            print(f'Add New Adjuster Info')
            btnAddNew = uas.child_window(auto_id="btnAddNew",control_type="Button").wrapper_object()
            safe_click(btnAddNew)

            newowner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
            newowner.wait('exists',timeout=5.0)
            uaan = newowner.child_window(auto_id="frmUnityAdjusterAddNew",control_type="Window")
            uaan.wait('exists',timeout=5.0)

            RawAdjusterName = allData['adjuster'].strip()

            if "," in RawAdjusterName:
                AdjusterLName,restName = [name.strip() for name in RawAdjusterName.split(",", 1)]
                AdjusterFName = restName.split()[0]
            else:
                adjusterNameSplit = RawAdjusterName.split()
                AdjusterFName = adjusterNameSplit[0] if adjusterNameSplit else ""
                AdjusterLName = adjusterNameSplit[-1] if len(adjusterNameSplit)>1 else ""
            
            
            tbFirst = uaan.child_window(auto_id="tbFirst", control_type="Edit").wrapper_object()
            tbFirst.iface_value.SetValue(AdjusterFName)

            tbLast = uaan.child_window(auto_id="tbLast", control_type="Edit").wrapper_object()
            tbLast.iface_value.SetValue(AdjusterLName)

            tbEmail = uaan.child_window(auto_id="tbEmail", control_type="Edit").wrapper_object()
            tbEmail.iface_value.SetValue(allData['adjEmail'])

            tbPhone = uaan.child_window(auto_id="tbPhone", control_type="Edit").wrapper_object()
            tbPhone.iface_value.SetValue(allData['adjPhone'])

            # tbFax = uaan.child_window(auto_id="tbFirst", control_type="Edit").wrapper_object()
            # tbFax.iface_value.SetValue(AdjusterFName)

            tbAddress = uaan.child_window(auto_id="tbAddress", control_type="Edit").wrapper_object()
            tbAddress.iface_value.SetValue(allData['adjAddr1'])

            tbAddress2 = uaan.child_window(auto_id="tbAddress2", control_type="Edit").wrapper_object()
            tbAddress2.iface_value.SetValue(allData['adjAddr2'])

            tbCity = uaan.child_window(auto_id="tbCity", control_type="Edit").wrapper_object()
            tbCity.iface_value.SetValue(allData['adjCity'])

            cbState = uaan.child_window(auto_id="cbState",control_type="ComboBox").wrapper_object()
            cbState.select(allData['adjState']) 

            tbZip = uaan.child_window(auto_id="tbZip", control_type="Edit").wrapper_object()
            tbZip.iface_value.SetValue(allData['adjZip'])

            notify("Notice","Please validate the Adjuster Information and Manually Click Add the Adjuster then Click Ok.")
            # btnAddAdjuster = uaan.child_window(auto_id="btnAddAdjuster",control_type="Button").wrapper_object()
            # btnAddAdjuster.click_input()
            # CloseWindowAdjuster = uas.child_window(auto_id="btnClose",control_type="Button").wrapper_object()
            # CloseWindowAdjuster.click_input()

        if not validateAdjuster:
            print(f'Add New Adjuster Info')
            btnAddNew = uas.child_window(auto_id="btnAddNew",control_type="Button").wrapper_object()
            safe_click(btnAddNew)

            newowner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
            newowner.wait('exists',timeout=5.0)
            uaan = newowner.child_window(auto_id="frmUnityAdjusterAddNew",control_type="Window")
            uaan.wait('exists',timeout=5.0)

            RawAdjusterName = allData['adjuster'].strip()

            if "," in RawAdjusterName:
                AdjusterLName,restName = [name.strip() for name in RawAdjusterName.split(",", 1)]
                AdjusterFName = restName.split()[0]
            else:
                adjusterNameSplit = RawAdjusterName.split()
                AdjusterFName = adjusterNameSplit[0] if adjusterNameSplit else ""
                AdjusterLName = adjusterNameSplit[-1] if len(adjusterNameSplit)>1 else ""
            
            
            tbFirst = uaan.child_window(auto_id="tbFirst", control_type="Edit").wrapper_object()
            tbFirst.iface_value.SetValue(AdjusterFName)

            tbLast = uaan.child_window(auto_id="tbLast", control_type="Edit").wrapper_object()
            tbLast.iface_value.SetValue(AdjusterLName)

            tbEmail = uaan.child_window(auto_id="tbEmail", control_type="Edit").wrapper_object()
            tbEmail.iface_value.SetValue(allData['adjEmail'])

            tbPhone = uaan.child_window(auto_id="tbPhone", control_type="Edit").wrapper_object()
            tbPhone.iface_value.SetValue(allData['adjPhone'])

            # tbFax = uaan.child_window(auto_id="tbFirst", control_type="Edit").wrapper_object()
            # tbFax.iface_value.SetValue(AdjusterFName)

            tbAddress = uaan.child_window(auto_id="tbAddress", control_type="Edit").wrapper_object()
            tbAddress.iface_value.SetValue(allData['adjAddr1'])

            tbAddress2 = uaan.child_window(auto_id="tbAddress2", control_type="Edit").wrapper_object()
            tbAddress2.iface_value.SetValue(allData['adjAddr2'])

            tbCity = uaan.child_window(auto_id="tbCity", control_type="Edit").wrapper_object()
            tbCity.iface_value.SetValue(allData['adjCity'])

            cbState = uaan.child_window(auto_id="cbState",control_type="ComboBox").wrapper_object()
            cbState.select(allData['adjState']) 

            tbZip = uaan.child_window(auto_id="tbZip", control_type="Edit").wrapper_object()
            tbZip.iface_value.SetValue(allData['adjZip'])

            # notify("Notice","Please validate the Adjuster Information and Manually Add the Adjuster.")
            notify("Notice","Please validate the Adjuster Information and Manually Click Add the Adjuster then Click Ok.")

        # tbClaimantAttorney
        # pbClaimantAttorney (Pane)
        
        time.sleep(2)
        ClaimAtty = allData['claimantAttyName']
        ClaimantPhone = allData['claimantAttyPhone']
        ClaimantState = allData['claimantAttyState']
        ClaimantZip = allData['claimantAttyZip']
        ClaimantCity = allData['claimantAttyCity']

        if " " in ClaimantPhone:
            ClaimantPhoneParts = ClaimantPhone.split()
            if len(ClaimantPhoneParts) > 1:
                ClaimantPhone = ClaimantPhoneParts[0]

        if ClaimAtty:
            print(f'Navigating to Attorney Search')
            
            try:
                pbClaimantAttorney = dlg.child_window(auto_id="pbClaimantAttorney",control_type="Pane").wrapper_object()
                safe_click(pbClaimantAttorney)
            except:   
                pbClaimantAttorney = dlg.child_window(auto_id="pbClaimantAttorneyCheckmark",control_type="Pane").wrapper_object()
                safe_click(pbClaimantAttorney)

            owner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
            owner.wait('exists',timeout=5)
            uads = owner.child_window(auto_id="frmUnityAttorneySearch",control_type="Window")
            uads.wait('exists',timeout=5)


            #frmUnityAttorneySearch
            parts = ClaimAtty.strip().split()
            # Initialize defaults
            ClaimAttyFirstName = ""
            ClaimAttyLastName = ""
            if len(parts) == 1:
                ClaimAttyFirstName = parts[0]
            elif len(parts) == 2:
                ClaimAttyFirstName, ClaimAttyLastName = parts
            else:
                # For names with middle name or initial
                ClaimAttyFirstName = parts[0]
                ClaimAttyLastName = parts[-1]

            # ClaimAttyFirstName = "Andrew"
            # ClaimAttyLastName = "Creech"    
            # print("First Name:", ClaimAttyFirstName)
            # print("Last Name:", ClaimAttyLastName)

            def _get_edit_Atty(auto_id):
                return uads.child_window(auto_id=auto_id, control_type="Edit").wrapper_object()
            
            def _select_stateAtty(value):
                try:
                    cb = uads.child_window(auto_id="cbState", control_type="ComboBox").wrapper_object()
                    if value:
                        cb.select(value)
                    else:
                        try:
                            cb.select(0)
                        except Exception:
                            pass
                except Exception:
                    pass

            def _fill_common_filters_Atty(phone, city, state, zip_):
            # Phone
                try: _set_value(_get_edit_Atty("tbPhone"), phone)
                except Exception: pass
                # City
                try: _set_value(_get_edit_Atty("tbCity"), city)
                except Exception: pass
                # Zip
                try: _set_value(_get_edit_Atty("tbZip"), zip_)
                except Exception: pass
                # State
                try:
                    _select_stateAtty(state)
                except Exception: pass
                
            
            def _wait_search_idle_atty(timeout=300.0):
                lbl = uads.child_window(auto_id="lblStatus", control_type="Text")
                end_t = time.monotonic() + timeout
                while time.monotonic() < end_t:
                    try:
                        w = lbl.wrapper_object()
                        if (not w.is_visible()) or ((w.window_text() or "").strip() == ""):
                            return True
                    except Exception:
                        return True
                    time.sleep(0.2)
                return False
            
            def _select_state_atty(value):
                try:
                    cb = uads.child_window(auto_id="cbState", control_type="ComboBox").wrapper_object()
                    if value:
                        cb.select(value)
                    else:
                        try:
                            cb.select(0)
                        except Exception:
                            pass
                except Exception:
                    pass

            def _results_count_atty():
                try:
                    grid = uads.child_window(auto_id="dgResults", control_type="Table").wrapper_object()
                    return int(getattr(grid.iface_grid, "CurrentRowCount", 0) or 0)
                except Exception:
                    return 0
            def _clear_all_inputs_atty():
                for eid in ("tbFirst","tbLast","tbFirmName","tbPhone","tbCity","tbZip"):
                    try:
                        _set_value(_get_edit_Atty(eid), "")
                    except Exception:
                        pass
                _select_state_atty(None)
            
            
            def _search_with_atty(first=None, last=None, FirmName=None,
                        phone=None, city=None, state=None, zip_=None):
                """Populate fields, click Search, wait, return count."""
                _clear_all_inputs_atty()
                # Names
                if first is not None:
                    try: _set_value(_get_edit_Atty("tbFirst"), first)
                    except Exception: pass
                if last is not None:
                    try: _set_value(_get_edit_Atty("tbLast"), last)
                    except Exception: pass

                # FirmName
                if FirmName is not None:
                    try: _set_value(_get_edit_Atty("tbFirmName"), FirmName)
                    except Exception: pass
                # Common filters
                _fill_common_filters_Atty(phone, city, state, zip_)
                time.sleep(3)
                # Search
                btn = uads.child_window(auto_id="btnSearch", control_type="Button").wrapper_object()
                safe_click(btn)
                # _click_search()
                _wait_search_idle_atty()

                return _results_count_atty()
            

            ##########ATTY TEST
            def _norm(s) -> str:
                return " ".join((s or "").replace("\n", " ").strip().split()).lower()

            def _in_match(needle: str, hay: str) -> bool:
                needle_n = _norm(needle)
                if not needle_n:
                    return True
                return needle_n in _norm(hay)

            def _addr_match(addr_expected: str, addr_row: str) -> bool:
                if not _norm(addr_expected):
                    return True
                return bool(addressMatch(addr_expected, addr_row))

            def _base_phone(phone: str) -> str:
                parts = (phone or "").strip().split()
                return parts[0] if parts else ""

            def _phone_match(expected_phone: str, row_phone: str) -> bool:
                if not (expected_phone or "").strip():
                    return True

                expected_base = _base_phone(expected_phone)

                try:
                    realPhone = (format_phone_us(expected_base) or "").strip()
                except Exception:
                    realPhone = expected_base.strip()

                try:
                    tempPhone = (format_phone_us(row_phone) or "").strip()
                except Exception:
                    tempPhone = (row_phone or "").strip()

                # fallback if formatter returns empty
                if not realPhone or not tempPhone:
                    digits_real = re.sub(r"\D+", "", expected_base)
                    digits_temp = re.sub(r"\D+", "", row_phone or "")
                    return (not digits_real) or (digits_real in digits_temp)

                return realPhone == tempPhone

            def select_matching_attorney_row_using_your_format() -> bool:
                ClaimAtty = allData['claimantAttyName']
                ClaimantAttyPhone = allData['claimantAttyPhone']
                ClaimantAttyState = allData['claimantAttyState']
                ClaimantAttyZip = allData['claimantAttyZip']
                ClaimantAttyCity = allData['claimantAttyCity']
                # ClaimAttyFirstName = "Andrew"
                # ClaimAttyLastName = "Creech"
                claimantAttyFirm = ""
                ClaimantAttyAddr1 = allData['claimantAttyAddr1']
                ClaimantAttyAddr2 = allData['claimantAttyAddr2']

                if " " in ClaimantAttyPhone:
                    ClaimantPhoneParts = ClaimantAttyPhone.split()
                    if len(ClaimantPhoneParts) > 1:
                        ClaimantAttyPhone = ClaimantPhoneParts[0]
                try:
                    grid = uads.child_window(auto_id="dgResults", control_type="Table")#.wrapper_object()
                    rows = int(getattr(grid.iface_grid, "CurrentRowCount", 0) or 0)
                    if rows <= 0:
                        return False

                    # Start at 1 if row 0 is header; if not, change to 0
                    start_row = 0

                    for r in range(start_row, rows):
                        print(f'Validating Result. {r} out of {rows}')
                        rule_A =False
                        rule_B =False
                        rule_C=False
                        # --- READ VALUES USING YOUR FORMAT ---
                        # col 2 First Name
                        # cell_el = grid.iface_grid.GetItem(r, 2)
                        # cell = UIAWrapper(UIAElementInfo(cell_el))
                        cell = grid.child_window(title=f"First Name Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                        rAttyFName = (cell.iface_value.CurrentValue or "").strip()

                        # col 3 Last Name
                        # cell_el = grid.iface_grid.GetItem(r, 3)
                        # cell = UIAWrapper(UIAElementInfo(cell_el))
                        cell = grid.child_window(title=f"Last Name Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                        rAttyLName = (cell.iface_value.CurrentValue or "").strip()

                        # col 4 Firm Name
                        # cell_el = grid.iface_grid.GetItem(r, 4)
                        # cell = UIAWrapper(UIAElementInfo(cell_el))
                        cell = grid.child_window(title=f"Firm Name Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                        rFirmName = (cell.iface_value.CurrentValue or "").strip()

                        # col 5 Address (single string)
                        # cell_el = grid.iface_grid.GetItem(r, 5)
                        # cell = UIAWrapper(UIAElementInfo(cell_el))
                        cell = grid.child_window(title=f"Address Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                        rAttyAddr = (cell.iface_value.CurrentValue or "").strip()

                        # col 6 Phone
                        # cell_el = grid.iface_grid.GetItem(r, 6)
                        # cell = UIAWrapper(UIAElementInfo(cell_el))
                        cell = grid.child_window(title=f"Phone Row {r}, Not sorted.", control_type="Edit").wrapper_object()
                        rAttyPhone = (cell.iface_value.CurrentValue or "").strip()

                        # --- OPTIONAL EXTRA FILTERS (only if you provided them) ---
                        # If you pass these, we require them regardless of rule A/B/C.
                        if claimantAttyFirm and not _in_match(claimantAttyFirm, rFirmName):
                            continue
                        if ClaimantAttyCity and not _in_match(ClaimantAttyCity, rAttyAddr):
                            continue
                        if ClaimantAttyState and not _in_match(ClaimantAttyState, rAttyAddr):
                            continue
                        if ClaimantAttyZip and not _in_match(ClaimantAttyZip, rAttyAddr):
                            continue

                        # --- BASE NAME MATCH FLAGS ---
                        f_ok = _in_match(ClaimAttyFirstName, rAttyFName) if (ClaimAttyFirstName or "").strip() else False
                        l_ok = _in_match(ClaimAttyLastName, rAttyLName) if (ClaimAttyLastName or "").strip() else False

                        # A) First OR Last
                        

                        # For B/C you said: "first name and last name and ..."
                        have_both_names = bool((ClaimAttyFirstName or "").strip()) and bool((ClaimAttyLastName or "").strip())
                        names_both_ok = have_both_names and _in_match(ClaimAttyFirstName, rAttyFName) and _in_match(ClaimAttyLastName, rAttyLName)

                        # B) First + Last + Address (Addr1/Addr2; match either)
                        have_any_addr = bool(_norm(ClaimantAttyAddr1) or _norm(ClaimantAttyAddr2))
                        addr_ok = True
                        if have_any_addr:
                            addr_ok = (
                                _addr_match(ClaimantAttyAddr1, rAttyAddr) or
                                _addr_match(ClaimantAttyAddr2, rAttyAddr)
                            )
                        rule_B = names_both_ok and have_any_addr and addr_ok

                        # C) First + Last + Phone
                        have_phone = bool((ClaimantAttyPhone or "").strip())
                        phone_ok = _phone_match(ClaimantAttyPhone, rAttyPhone) if have_phone else False
                        rule_C = names_both_ok and have_phone and phone_ok

                        # ---- SELECT IF ANY RULE HIT ----
                        if f_ok and l_ok:
                            if not (rule_B or rule_C):
                                continue
                        else:
                            # if f_ok or l_ok:
                            if not ClaimAttyFirstName or not ClaimAttyLastName:
                                rule_A = f_ok or l_ok
                                if not (rule_A):
                                    continue

                        # Click checkbox column 0
                        # print(f'testing  {rAttyFName} {rAttyLName} {rAttyAddr}')
                        # print(f'test {rule_A} {rule_B} {rule_C}')
                        if rule_A or rule_B or rule_C:
                            try:
                                print(f'Selecting {rAttyFName} {rAttyLName} {rAttyAddr}')
                                # select_el = grid.iface_grid.GetItem(r, 0)
                                # select_cell = UIAWrapper(UIAElementInfo(select_el))
                                select_cell = grid.child_window(title=f"Select Row {r}", control_type="CheckBox").wrapper_object()
                                safe_click(select_cell)
                                time.sleep(2)
                                SelectAtty = uads.child_window(auto_id="btnSelect",control_type="Button").wrapper_object()
                                safe_click(SelectAtty)
                                return True
                            except Exception:
                                try:
                                    # select_el = grid.iface_grid.GetItem(r, 0)
                                    # select_cell = UIAWrapper(UIAElementInfo(select_el))
                                    select_cell = grid.child_window(title=f"Select Row {r}", control_type="CheckBox").wrapper_object()
                                    kids = select_cell.children()
                                    if kids:
                                        safe_click(kids[0])
                                        return True
                                except Exception:
                                    pass
                    return False

                except Exception:
                    return False






            #############
            def _click_first_row_if_any_atty():
                global isManualEnteredAtty
                try:
                    grid = uads.child_window(auto_id="dgResults", control_type="Table").wrapper_object()
                    rows = int(getattr(grid.iface_grid, "CurrentRowCount", 0) or 0)

                    for r in range(rows):
                        if rows >= 1:
                            cell_el = grid.iface_grid.GetItem(r,2)
                            cell = UIAWrapper(UIAElementInfo(cell_el))
                            rAttyFName = cell.iface_value.CurrentValue

                            cell_el = grid.iface_grid.GetItem(r,3)
                            cell = UIAWrapper(UIAElementInfo(cell_el))
                            rAttyLName = cell.iface_value.CurrentValue

                            cell_el = grid.iface_grid.GetItem(r,4)
                            cell = UIAWrapper(UIAElementInfo(cell_el))
                            rFirmName = cell.iface_value.CurrentValue

                            cell_el = grid.iface_grid.GetItem(r,5)
                            cell = UIAWrapper(UIAElementInfo(cell_el))
                            rAttyAddr = cell.iface_value.CurrentValue

                            cell_el = grid.iface_grid.GetItem(r,6)
                            cell = UIAWrapper(UIAElementInfo(cell_el))
                            rAttyPhone = cell.iface_value.CurrentValue

                            print(f"{rAttyFName}\n{rAttyLName}\n{rFirmName}\n{rAttyAddr}\n{rAttyPhone}")

                            atyAddrFull = f"{allData['claimantAttyAddr1']} {allData['claimantAttyAddr2']} {allData['claimantAttyCity']} {allData['claimantAttyState']} {allData['claimantAttyZip']}"

                            TempAddress = expand_suffix_long(rAttyAddr)
                            RealAddress = expand_suffix_long(atyAddrFull)
                            ComparisonResult = addressMatch(TempAddress,RealAddress)

                            if ClaimAttyFirstName and ClaimAttyFirstName in rAttyFName:
                                checkAttyFName = True
                            if ClaimAttyLastName and ClaimAttyLastName in rAttyLName:
                                checkAttyLName = True
                            

                            if ComparisonResult:
                                cell_el = grid.iface_grid.GetItem(r,0)
                                cell = UIAWrapper(UIAElementInfo(cell_el))
                                safe_click(cell)
                                time.sleep(2)
                                SelectAtty = uads.child_window(auto_id="btnSelect",control_type="Button").wrapper_object()
                                safe_click(SelectAtty)
                                validateAdjuster = True
                            # notify("Notice",f"Please Select the correct information from the Attorney Results. \n\nAttorney Name: {allData['claimantAttyName']} \nProvider Address: {allData['claimantAttyAddr1']} {allData['claimantAttyAddr2']} {allData['claimantAttyCity']} {allData['claimantAttyState']} {allData['claimantAttyZip']}\nProvider Phone: {allData['claimantAttyPhone']}")
                            isManualEnteredAtty = True
                            return True
                except Exception:
                    pass
                return False
                
                # ClaimAtty = allData['claimantAttyName']
                # ClaimantPhone = allData['claimantAttyPhone']
                # ClaimantState = allData['claimantAttyState']
                # allData['claimantAttyAddr1']
                # ClaimantZip = allData['claimantAttyZip']

            def run_atty_search():
                firstAttempAtty = False
                global isManualEntered 
                isManualEntered = False
                if ClaimantPhone:
                    count = _search_with_atty(first=None, last=None, FirmName=None,
                                        phone=ClaimantPhone, city=None, state=None, zip_=None)
                    if count <= 10:
                        firstAttempAtty = select_matching_attorney_row_using_your_format()
                        if firstAttempAtty:
                            isManualEntered= firstAttempAtty
                            return count
                if not firstAttempAtty:    
                    if ClaimAttyLastName or ClaimAttyFirstName or (ClaimantZip and ClaimantState and ClaimantCity) :
                        count = _search_with_atty(first=ClaimAttyFirstName[:3], last=ClaimAttyLastName, FirmName=None,
                                            phone=None, city=ClaimantCity, state=ClaimantState, zip_=ClaimantZip)
                        if count > 0:
                            
                            isManualEntered= select_matching_attorney_row_using_your_format()
                            return count
                    isManualEntered = False    
                    return 0
                
            attyCount = run_atty_search()
            # print(f"Atty search: {isManualEntered}")
            if not isManualEntered:
                print(f'Adding new attorney info')
                if attyCount == 0:
                    btnAddNew= uads.child_window(auto_id="btnAddNew",control_type="Button").wrapper_object()
                    safe_click(btnAddNew)

                    uatan = uads.child_window(auto_id="frmUnityAttorneyAddNew",control_type="Window")
                    uatan.wait('exists',timeout=3)
                    tbFirst = uatan.child_window(auto_id="tbFirst", control_type="Edit").wrapper_object()
                    tbFirst.iface_value.SetValue(ClaimAttyFirstName)
                    tbLast = uatan.child_window(auto_id="tbLast", control_type="Edit").wrapper_object()
                    tbLast.iface_value.SetValue(ClaimAttyLastName)
                    tbAddress1= uatan.child_window(auto_id="tbAddress1", control_type="Edit").wrapper_object()
                    tbAddress1.iface_value.SetValue(allData['claimantAttyAddr1'])
                    tbAddress2= uatan.child_window(auto_id="tbAddress2", control_type="Edit").wrapper_object()
                    tbAddress2.iface_value.SetValue(allData['claimantAttyAddr2'])
                    tbCity= uatan.child_window(auto_id="tbCity", control_type="Edit").wrapper_object()
                    tbCity.iface_value.SetValue(allData['claimantAttyCity'])
                    cbState = uatan.child_window(auto_id="cbState",control_type="ComboBox").wrapper_object()
                    cbState.select(allData['claimantAttyState']) 
                    tbZip= uatan.child_window(auto_id="tbZip", control_type="Edit").wrapper_object()
                    tbZip.iface_value.SetValue(allData['claimantAttyZip'])
                    tbPhone= uatan.child_window(auto_id="tbPhone", control_type="Edit").wrapper_object()
                    tbPhone.iface_value.SetValue(ClaimantPhone) # migz

                    notify("Notice","Please validate if Attorney information is correct. User to Manual Click on Add Attorney.")


            # tbFirst = uads.child_window(auto_id="tbFirst", control_type="Edit").wrapper_object()
            # tbFirst.iface_value.SetValue(ClaimAttyFirstName[:3])
            # tbLast = uads.child_window(auto_id="tbLast", control_type="Edit").wrapper_object()
            # tbLast.iface_value.SetValue(ClaimAttyLastName)
            # tbPhone = uads.child_window(auto_id="tbPhone", control_type="Edit").wrapper_object()
            # tbPhone.iface_value.SetValue(ClaimantPhone)
            # cbState = uads.child_window(auto_id="cbState",control_type="ComboBox").wrapper_object()
            # cbState.select(ClaimantState) 
            # tbZip = uads.child_window(auto_id="tbZip", control_type="Edit").wrapper_object()
            # tbZip.iface_value.SetValue(ClaimantZip)

            # SearchAdjuster = uads.child_window(auto_id="btnSearch",control_type="Button").wrapper_object()
            # SearchAdjuster.click_input()

            # lblStatus.Name ="Searching please wait"
            # lblStatusChecker = uads.child_window(auto_id="lblProgress",control_type="Text")
            # timeout = 300.0
            # EndTime = time.monotonic() + timeout
            # while time.monotonic() < EndTime:
            #     try:
            #         lblStatus = lblStatusChecker.wrapper_object()
            #         if (not lblStatus.is_visible()) or ((lblStatus.window_text() or "").strip()==""):
            #             break
            #     except Exception:
            #         break


            # grid = uads.child_window(auto_id="dgResults",control_type="Table").wrapper_object()
            # AttachRow = grid.iface_grid.CurrentRowCount

            # for r in range(0,AttachRow):   
            #     cell_el = grid.iface_grid.GetItem(r,0)
            #     cell = UIAWrapper(UIAElementInfo(cell_el))
            #     cell.click_input()

            #     SelectAtty = uads.child_window(auto_id="btnSelect",control_type="Button").wrapper_object()
            #     SelectAtty.click_input()
            #     break

            # if AttachRow == 0:
            #     CloseWindowAtty = uads.child_window(auto_id="btnClose",control_type="Button").wrapper_object()
            #     CloseWindowAtty.click_input()

        # pbDefeneseAttorney(Pane)
        defAtty= ""
        if defAtty:
            pbDefeneseAttorney = dlg.child_window(auto_id="pbDefeneseAttorney",control_type="Pane").wrapper_object()
            safe_click(pbDefeneseAttorney)
            time.sleep(5)
            owner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
            owner.wait('exists',timeout=0.5)
            uads = owner.child_window(auto_id="frmUnityAttorneySearch",control_type="Window")
            uads.wait('exists',timeout=0.5)


            CloseWindowAdjuster = uads.child_window(auto_id="btnClose",control_type="Button").wrapper_object()
            safe_click(CloseWindowAdjuster)
        #frmUnityAttorneySearch
        #btnClose

        #Unity Body Part Name:"Provider Info"
        cbBodyPart = dlg.child_window(title="Provider Info", control_type="Edit").wrapper_object()
        cbBodyPartVal = cbBodyPart.iface_value.CurrentValue
        if not cbBodyPartVal:
            cbBodyPart.iface_value.SetValue(allData['bodyPart']) 

        # TAB Injury Type
        cbInjuryType = dlg.child_window(auto_id="cbInjuryType", control_type="ComboBox").wrapper_object()
        cbInjTypeVal = cbInjuryType.iface_value.CurrentValue
        

        # cbInjuryType.iface_value.SetValue(allData['injuryType'])
        # send_keys('{TAB}')
        # send_keys(allData['injuryType'],with_spaces=True) #All Other Specific Injuries, NOC
        # cbInjuryType cbInjuryCause

        # TAB Injury Cause
        cbInjuryCause = dlg.child_window(auto_id="cbInjuryCause", control_type="ComboBox").wrapper_object()
        cbInjCauseVal = cbInjuryCause.iface_value.CurrentValue

        if not cbInjTypeVal or not cbInjCauseVal:
            notify("Notice",f"Injury Type: {allData['injuryType']}\nInjury Cause: {allData['injuryCause']}")
        # cbInjuryCause.iface_value.SetValue(allData['injuryCause'])
        # send_keys('{TAB}')
        # send_keys(allData['injuryCause'],with_spaces=True) #'Other - Miscellaneous, NOC
        

        # btnSpecialInstrucitons
        # SpecialInstBtn = dlg.child_window(auto_id="btnSpecialInstrucitons",control_type="Button").wrapper_object()
        # SpecialInstBtn.click_input()
        # time.sleep(3)

        # owner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
        # owner.wait('exists',timeout=0.5)
        # specIns = owner.child_window(auto_id="frmSpecialInstructions",control_type="Window")
        # specIns.wait('exists',timeout=0.5)
        # time.sleep(3)
        # rtcInstruction = specIns.child_window(auto_id="rtbSpecialInstructions", control_type="Document").wrapper_object()
        # SpecialInstructionText = allData['specialInstructions']
        # rtcInstruction.iface_value.SetValue(SpecialInstructionText)
        # CloseWindowIns = specIns.child_window(auto_id="btnOk",control_type="Button").wrapper_object()
        # CloseWindowIns.click_input()

        cbClaimType = dlg.child_window(auto_id="cbClaimType", control_type="ComboBox").wrapper_object()
        if CTMedOnly:
            cbClaimType.iface_value.SetValue("Medical Only")
        else:
            cbClaimType.iface_value.SetValue("Lost Time")
        # cbClaimType.select("Lost Time") 
        time.sleep(3)
        focus_control(dlg)


        #     # owner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
        #     # owner.wait('exists',timeout=0.5)
        #     # ucc = owner.child_window(auto_id="frmUnityClaimClaimant",control_type="Window")
        #     # ucc.wait('exists',timeout=3)
        #     # ucc.set_focus()

        #     owner = desk.window(title_re=r"Subject Line Builder Email Display.*", control_type="Window")
        #     owner.wait("exists", timeout=10)

        #     def get_urs(timeout=20):
        #         def _locate():
        #             hwnd = find_windows(
        #                 title="Unity Referral Source",
        #                 class_name="WindowsForms10.Window.8.app.0.6255dd_r8_ad1",
        #                 top_level_only=False,
        #                 visible_only=False,
        #                 enabled_only=False
        #             )[0]

        #             app = Application(backend="uia").connect(handle=hwnd)
        #             w = app.window(handle=hwnd)
        #             w.wait("exists", timeout=1)
        #             return w

        #         urs = wait_until_passes(timeout, 0.2, _locate)
        #         urs.wait("ready", timeout=timeout)
        #         try:
        #             urs.set_focus()
        #         except Exception:
        #             pass
        #         return urs

        #     urs = get_urs(timeout=20)


        #     # owner = desk.window(title_re=r"Subject Line Builder Email Display*",control_type="Window")
        #     # owner.wait('exists',timeout=3)
        #     # urs = Desktop(backend="uia").window(title_re=r"Unity Referral Source*",control_type="Window")
        #     # urs = Desktop(backend="uia").window(auto_id="frmUnityReferralSource",control_type="Window")
            

        #     # urs.wait('visible ready exists',timeout=3) 
        #     time.sleep(3)
        #     #Validate CB Values
        #     cbContactType = urs.child_window(auto_id="cbContactType", control_type="ComboBox").wrapper_object()
        #     cbContactType.iface_value.SetValue("Customer TCM")
        #     notify("Notice","Please validate if Selected dropdown for Contact Type is Correct.")
        #     time.sleep(3)
        #     rbRefSource = dlg.child_window(auto_id="rbRefSource", control_type="RadioButton").wrapper_object()
        #     rbRefSource.select()
        # else:
        #     rbAdjuster = dlg.child_window(auto_id="rbAdjuster", control_type="RadioButton").wrapper_object()
        #     rbAdjuster.select()

       # assumes you already have: import time, ctypes
# and: from pywinauto import Desktop
# and: from pywinauto.timings import wait_until_passes

        def msgbox(text, title="Debug"):
            ctypes.windll.user32.MessageBoxW(0, str(text), str(title), 0)

        desk = Desktop(backend="uia")

        owner = desk.window(title="Subject Line Builder Email Display", control_type="Window")
        owner.wait("exists ready", timeout=10)
        time.sleep(0.2)

        # rbAdjuster
        if allData['refSource']:
            # IMPORTANT: CLICK instead of Toggle (more like manual)
            try:
                cbReferralSource = owner.child_window(auto_id="cbReferralSource", control_type="CheckBox").wrapper_object()

                # click only if unchecked (avoid accidentally unchecking)
                try:
                    if cbReferralSource.get_toggle_state() != 1:
                        safe_click(cbReferralSource)
                except Exception:
                    safe_click(cbReferralSource)

            except Exception as e:
                msgbox(f"Failed to click cbReferralSource.\n{type(e).__name__}: {e}", "RRS UIA")
                raise

            # let WinForms spawn popup + UIA tree update
            time.sleep(0.6)

            def find_urs_under_owner():
                # Re-acquire owner each retry to avoid stale UIA subtree
                o = desk.window(title="Subject Line Builder Email Display", control_type="Window")

                # small settle each retry
                time.sleep(0.15)

                # Search deeper; filter by control_type in element_info
                for w in o.descendants():
                    try:
                        ei = w.element_info
                        if ei.control_type != "Window":
                            continue

                        title = (w.window_text() or "").strip()
                        autoid = (ei.automation_id or "").strip()

                        if autoid == "frmUnityReferralSource" or title == "Unity Referral Source":
                            return desk.window(handle=w.handle)
                    except Exception:
                        pass

                raise RuntimeError("Unity Referral Source not found under owner yet")

            try:
                urs = wait_until_passes(30, 0.3, find_urs_under_owner)
            except Exception as e:
                # Show what UIA sees under owner (Window nodes only)
                lines = []
                try:
                    o = desk.window(title="Subject Line Builder Email Display", control_type="Window")
                    for w in o.descendants():
                        try:
                            ei = w.element_info
                            if ei.control_type == "Window":
                                lines.append(
                                    f"{hex(w.handle)} | {repr(w.window_text())} | auto_id={ei.automation_id!r} | class={ei.class_name!r}"
                                )
                        except Exception:
                            pass
                except Exception as ee:
                    lines.append(f"<descendants failed: {ee!r}>")

                msgbox(
                    "FAILED to find 'Unity Referral Source' under owner.\n\n"
                    f"Error: {type(e).__name__}: {e}\n\n"
                    "Owner subtree windows:\n" + "\n".join(lines[:80]),
                    "RRS Popup Debug (UIA)"
                )
                raise

            # Wait exists then ready (plus small sleeps)
            urs.wait("exists", timeout=10)
            time.sleep(0.3)
            urs.wait("ready", timeout=10)
            time.sleep(0.2)

            # Now your original UIA code should work:
            cbContactType = urs.child_window(auto_id="cbContactType", control_type="ComboBox").wrapper_object()
            cbContactType.iface_value.SetValue("Customer TCM")
            time.sleep(0.15)

            tbFirstName = urs.child_window(auto_id="tbFirstName", control_type="Edit").wrapper_object()
            tbFirstName.iface_value.SetValue(AdjusterFName)
            time.sleep(0.10)

            tbLastName = urs.child_window(auto_id="tbLastName", control_type="Edit").wrapper_object()
            tbLastName.iface_value.SetValue(AdjusterLName)
            time.sleep(0.10)

            tbEmail = urs.child_window(auto_id="tbEmail", control_type="Edit").wrapper_object()
            tbEmail.iface_value.SetValue(allData['adjEmail'])
            time.sleep(0.10)

            tbPhone = urs.child_window(auto_id="tbPhone", control_type="Edit").wrapper_object()
            tbPhone.iface_value.SetValue(allData['adjPhone'])
            
            btnOK = urs.child_window(auto_id="btnOK", control_type="Button").wrapper_object()
            safe_click(btnOK)

            #Validate CB Values

            rbRefSource = dlg.child_window(auto_id="rbRefSource", control_type="RadioButton").wrapper_object()
            rbRefSource.select()
        else:
            rbAdjuster = dlg.child_window(auto_id="rbAdjuster", control_type="RadioButton").wrapper_object()
            rbAdjuster.select()

        # do no edit beyond this code   
        # rbMed,rbVoc
        if 'full case management' in allData['referralType'].lower() or "one-time rn visit" in allData['referralType'].lower():
            rbMed = dlg.child_window(auto_id="rbMed", control_type="RadioButton").wrapper_object()
            rbMed.select()
        elif "voc" in allData['referralType'].lower():
            rbVoc = dlg.child_window(auto_id="rbVoc", control_type="RadioButton").wrapper_object()
            rbVoc.select()

        # # tbDiagCode - edited
        # tbDiagCode = dlg.child_window(auto_id="tbDiagCode", control_type="Edit").wrapper_object()
        # tbDiagCode.iface_value.SetValue(allData['dxCode'])

        notify("Script Paused","Review SBL fields to ensure accuracy and completeness.\nOnce done, Click OK to build.")
        print(f'Submit Subject Template Builder')
        # btnCreate
        btnCreate = dlg.child_window(auto_id="btnCreate",control_type="Button").wrapper_object()
        safe_click(btnCreate)
        time.sleep(10)

        #Email Display Exit
        focus_control(emaiDisplay)
        btnExit = emaiDisplay.child_window(auto_id="btnExit",control_type="Button").wrapper_object()
        safe_click(btnExit)
        time.sleep(3)
        botStop = False
        # pprint(allData)
    except Exception as ex:
        # print("error occurred: ", ex)
        exc_type, exc_obj, tb = sys.exc_info()
        line_number = tb.tb_lineno
        # driver.execute_script("h$(0);")
        print(f"error occurred: {ex} (line {line_number})")
        botStop=True
        sys.exit()


def CompleteTriageAndExportAttachment():
    global correctZipCode
    global botStop
    #Complete Triage Process
    main = Desktop(backend="uia").window(title_re=r"Referral Routing System*",control_type="Window")
    pid = main.element_info.process_id
    focus_control(main)
    grid = main.child_window(auto_id="dgWorkWindow",control_type="Table")#.wrapper_object()
    # cell_el = grid.iface_grid.GetItem(0,1)
    # cell = UIAWrapper(UIAElementInfo(cell_el))
    cell = grid.child_window(title=f"Select Row 0", control_type="CheckBox").wrapper_object()
    safe_click(cell)

    #btnTriage
    btnTriage = main.child_window(auto_id="btnTriage",control_type="Button").wrapper_object()
    safe_click(btnTriage)
    time.sleep(3)

    
    try:
    #Zip Code?
        # print(correctZipCode)
        app = Application(backend="win32").connect(process=pid)
        prompt = app.window(title_re=r"Zip Code\?")
        prompt.wait("ready", 10)
        prompt.Edit.set_edit_text(correctZipCode)
        time.sleep(2)
        # prompt.Edit.set_edit_text("90017")
        safe_click(prompt.OK)
    except Exception:
        pass

    time.sleep(5)
    try:
    #Information Complete Dialogbox/Prompter
        infoCompletePrompter = Desktop(backend="win32").window(class_name="#32770",title_re="Information Complete?") #control_type="Window"
        infoCompletePrompter.wait("visible ready",timeout=5)
        safe_click(infoCompletePrompter["Yes"])
    except Exception:
        pass
    # infoCompletePrompter.child_window(title="&Yes",control_type="Button").click()
    time.sleep(5)

    try:
        reTriggerPrompter = Desktop(backend="win32").window(class_name="#32770",title_re="Unity ReTrigger")
        reTriggerPrompter.wait("visible ready",timeout=5)
        safe_click(reTriggerPrompter["Yes"])
    except Exception:
        pass
    time.sleep(5)

    # Check Customer
    try:
        checkCustomerPrompter = Desktop(backend="win32").window(class_name="#32770",title_re="Customer Check") #control_type="Window"
        checkCustomerPrompter.wait("visible ready",timeout=5)
        safe_click(checkCustomerPrompter["Yes"])
    except Exception:
        pass
    time.sleep(5)
    #Create If statment if existing
    try:
        reTriggerPrompter = Desktop(backend="win32").window(class_name="#32770",title_re="Unity ReTrigger")
        reTriggerPrompter.wait("visible ready",timeout=5)
        safe_click(reTriggerPrompter["Yes"])
    except Exception:
        pass
    time.sleep(5)
    # ReferralRoutingSystem
    try:
        ReferralRoutingSystemPrompter = Desktop(backend="win32").window(class_name="#32770",title_re="ReferralRoutingSystem")
            # ReferralRoutingSystemPrompter.print_control_identifiers()
        ReferralRoutingSystemPrompter.wait("visible ready",timeout=5)
        focus_control(ReferralRoutingSystemPrompter)
        okButton = ReferralRoutingSystemPrompter.child_window(title="OK",class_name="Button").wrapper_object()
        safe_click(okButton)
        time.sleep(5)
    except: pass
    try:
        ReferralRoutingSystemPrompter = Desktop(backend="win32").window(class_name="#32770",title_re="ReferralRoutingSystem")
        ReferralRoutingSystemPrompter.wait("visible ready",timeout=15)
        focus_control(ReferralRoutingSystemPrompter)
        okButton = ReferralRoutingSystemPrompter.child_window(title="OK",class_name="Button").wrapper_object()
        safe_click(okButton)
        time.sleep(5)
    except: pass

    #Export Attachement
    UnityEntry = Desktop(backend="uia").window(title_re=r"Unity Entry*",control_type="Window")
    UnityEntry.wait("visible ready",timeout=15)
    focus_control(UnityEntry)
    # UnityEntry.print_control_identifiers()
    exportBtn = UnityEntry.child_window(auto_id="btnDownload",control_type="Button").wrapper_object()
    focus_control(exportBtn)
    send_keys('{SPACE}')
 
    exportPrompter = Desktop(backend="win32").window(class_name="#32770",title_re="ReferralRoutingSystem")
    exportPrompter.wait("visible ready",timeout=300)
    safe_click(exportPrompter["Ok"])
    time.sleep(5)

    focus_control(UnityEntry)
    exportExitBtn = UnityEntry.child_window(auto_id="btnExit",control_type="Button").wrapper_object()
    safe_click(exportExitBtn)

def InsertIntoTable(Claimant,ClaimNumber):
    global connection_string
    with pyodbc.connect(connection_string) as conn:
        with conn.cursor() as cursor:
            sql = """INSERT INTO [dbo].[RRS_APP_Sagility_Bot_Assignment_Log] ([Claimant],[ClaimNumber],[isAssigned]) VALUES (?,?,?)"""
            values = (Claimant,ClaimNumber,0)
            cursor.execute(sql,values)
            conn.commit()
            print("Successfully Added")


def OpenUnity():
    global data
    print(f"{data['claimantFirst']} {data['claimantLast']}")
    # notify("Script Paused",f"The script is paused. Click OK to continue.") #{data['customer']} {data['claimantFull']}
    # if "," in data["claimantFull"]:
    #     last,first = [x.strip() for x in data["claimantFull"].split(",",1)]
    #     claimantFull = f"{first} {last}"
    claimantFull =f"{data['claimantFirst']} {data['claimantLast']}"
    # import UnityAssignment
    # UnityAssignment.main(data["customer"],claimantFull,data)

    # import UnityAssignmentV2
    import legacy.legacy_unity
    try:
        legacy.legacy_unity.main(data["customer"],claimantFull,data)
    
    except Exception as e:
        print(f"Error retrieving Assign/Unity status: {e}")
        traceback.print_exc()
    finally:
        InsertIntoTable(claimantFull, data["claimNumber"])
        
        # CheckAdvNotice(claimantFull, data["claimNumber"])

def CheckAdvNotice(claimant,claim_number):
    global connection_string
    try:
        with pyodbc.connect(connection_string) as conn:
            with conn.cursor() as cursor:
                # Stored procedure call; "EXEC proc ?, ?" also works if preferred
                cursor.execute("{CALL RRS_Sagility_Get_Assign_Unity_Status(?,?,?)}", claimant, claim_number,"Liberty Mutual Commercial Market")
                row = cursor.fetchone()

                if row:
                    assign_value = row[0] if row[0] is not None else ""
                    unity_value = row[1] if row[1] is not None else ""
                    return assign_value,unity_value

                return None
    except Exception as e:
        print(f"Error retrieving Assign/Unity status: {e}")
        return None
    
def notify(title: str, text: str):
    """Show a messagebox if Tk is available; otherwise print."""
    try:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost',True)
        messagebox.showinfo(title, text,parent=root)
        root.destroy()
    except Exception:
        print(f"[{title}] {text}")

def CompleteAssignment(clmNumber):
    dlg = Desktop(backend="uia").window(title_re=r"Referral Routing System*",control_type="Window")
    dlg.wait("exists enabled visible ready",timeout=120,retry_interval=1)
    focus_control(dlg)

    btnAssign = dlg.child_window(auto_id="btnAssign", control_type="Button").wrapper_object()
    safe_click(btnAssign)
    time.sleep(2)
    
    CompleteAssignmentForm = dlg.child_window(auto_id="frmInputBox",control_type="Window")
    CompleteAssignmentForm.wait('exists',timeout=5.0)
    focus_control(CompleteAssignmentForm)

    tbInput = CompleteAssignmentForm.child_window(auto_id="tbInput", control_type="Edit").wrapper_object()
    tbInput.iface_value.SetValue("19")
    time.sleep(2)

    btnOk = CompleteAssignmentForm.child_window(auto_id="btnOk", control_type="Button").wrapper_object()
    safe_click(btnOk)
    time.sleep(2)

    try:
        UnityCompletePrompter = Desktop(backend="win32").window(class_name="#32770",title_re=r"Unity Complete \?")
        UnityCompletePrompter.wait("visible ready",timeout=15)
        focus_control(UnityCompletePrompter)
        safe_click(UnityCompletePrompter["Yes"])
        time.sleep(5)
    except:
        pass

    

    #------------------------------------------------------------
    CMS_LOGIN_URL = CONFIG_CMS_LOGIN_URL
    EDGE_PATH = CONFIG_EDGE_PATH
    IE_DRIVER_PATH = CONFIG_IE_DRIVER_PATH
    from selenium import webdriver
    from selenium.webdriver.ie.service import Service as IeService
    from selenium.webdriver.ie.options import Options as IeOptions
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.keys import Keys
    from selenium.common.exceptions import WebDriverException,NoSuchWindowException,StaleElementReferenceException
    import os

    global driver

    def ensure_readable_path(path: str, description: str = ""):
        if not os.path.exists(path):
            raise FileNotFoundError(f"{description or 'Path'} not found: {path}")
        if not os.access(path, os.R_OK):
            raise PermissionError(f"Access denied (no read permission) for {description or path}")
 
 
    def ensure_executable_path(path: str, description: str = ""):
        ensure_readable_path(path, description)
        if not os.access(path, os.X_OK):
            print(f"Warning: {description or path} might not be executable. On Windows this is usually fine.")
        
    def legacy_safe_type(by, sel, text, timeout=20, click_first=True):
        global driver
        wait = WebDriverWait(driver, timeout)
        el = wait.until(EC.presence_of_element_located((by, sel)))
    
        driver.execute_script("arguments[0].scrollIntoView(true);", el)
        if click_first:
            try:
                safe_click(el)
            except WebDriverException:
                pass
    
        try:
            el.send_keys(Keys.CONTROL, "a")
            el.send_keys(Keys.DELETE)
            el.send_keys(text)
            if el.get_attribute("value") == text:
                return True
        except WebDriverException:
            pass
    
        try:
            ok = driver.execute_script(r"""
                var el = arguments[0], val = arguments[1];
                try { el.focus && el.focus(); } catch(e){}
                try { el.value = ''; } catch(e){}
                try { el.value = val; } catch(e){}
    
                if (document.createEventObject && el.fireEvent) {
                    try {
                        var ev = document.createEventObject();
                        ev.propertyName = 'value';
                        el.fireEvent('onpropertychange', ev);
                    } catch(e){}
                    try { el.fireEvent('oninput', document.createEventObject()); } catch(e){}
                    try { el.fireEvent('onchange', document.createEventObject()); } catch(e){}
                    try { el.fireEvent('onkeyup', document.createEventObject()); } catch(e){}
                    try { el.fireEvent('onblur', document.createEventObject()); } catch(e){}
                    return el.value === val;
                } else if (document.createEvent) {
                    var types = ['input','change','keyup','keydown','blur'];
                    for (var i=0;i<types.length;i++){
                        try {
                            var evt = document.createEvent('HTMLEvents');
                            evt.initEvent(types[i], true, false);
                            el.dispatchEvent(evt);
                        } catch(e){}
                    }
                    return el.value === val;
                } else {
                    return el.value === val;
                }
            """, el, text)
            if ok:
                return True
        except WebDriverException:
            pass
    
        return False
    
    def element_check(el):
        global driver
        return driver.execute_script(r"""
            (function(el){
                el.checked = true;
                if (document.createEventObject && el.fireEvent) {
                    el.fireEvent('onchange', document.createEventObject());
                } else if (document.createEvent) {
                    var e = document.createEvent('Event');
                    e.initEvent('change', true, false);
                    el.dispatchEvent(e);
                }
            })(arguments[0]);
        """, el)
    
    def element_click(el):
        global driver
        return driver.execute_script(r"""
            (function(el){
                function idToName(id){ return id.replace(/_/g,'$'); }
                try {
                    if (el && el.click) { el.click(); return; }
                } catch(e){}
    
                if (typeof window.__doPostBack === 'function') {
                    window.__doPostBack(idToName(el.id), '');
                    return;
                }
                if (el.form) {
                    try { el.form.submit(); return; } catch(e){}
                }
                if (document.createEventObject && el.fireEvent) {
                    try {
                        var ev = document.createEventObject();
                        el.fireEvent('onmousedown', ev);
                        el.fireEvent('onmouseup', ev);
                        el.fireEvent('onclick', ev);
                        return;
                    } catch(e){}
                }
                try { if (typeof el.onclick === 'function') el.onclick(); } catch(e){}
            })(arguments[0]);
        """, el)

    def elementExist(by, value, timeout=10):
        global driver
        try:
            WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))
            return True
        except Exception:
            return False
    

    import CMS
    SelectedCaseID = CMS.CheckCaseUnity(data["claimNumber"])
    # if SelectedCaseID:
    #     driver.quit()
    #-----------------------------------------------------------
    try:
    #Case Code? ADD LOGIC OF PUTTING CASE ID
    #---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        print(f'Case Selected: {SelectedCaseID}')
        main = Desktop(backend="uia").window(title_re=r"Referral Routing System*",control_type="Window")
        pid = main.element_info.process_id
        app = Application(backend="win32").connect(process=pid)
        prompt = app.window(title_re=r"Case Id")
        prompt.wait("ready", 10)
        prompt.Edit.set_edit_text(SelectedCaseID)
        time.sleep(2)
        safe_click(prompt.OK)
    except Exception:
        pass

    # Unity Complete ?
    try:
        UnityCompletePrompter = Desktop(backend="win32").window(class_name="#32770",title_re=r"Unity Complete \?")
        UnityCompletePrompter.wait("visible ready",timeout=5)
        focus_control(UnityCompletePrompter)
        safe_click(UnityCompletePrompter["Yes"])
        time.sleep(5)
    
    except:
        pass


    try:
        ReferralRoutingSystemPrompter = Desktop(backend="win32").window(class_name="#32770",title_re="ReferralRoutingSystem")
            # ReferralRoutingSystemPrompter.print_control_identifiers()
        ReferralRoutingSystemPrompter.wait("visible ready",timeout=5)
        focus_control(ReferralRoutingSystemPrompter)
        okButton = ReferralRoutingSystemPrompter.child_window(title="OK",class_name="Button").wrapper_object()
        safe_click(okButton)
        time.sleep(5)
    except: pass

    
    # frmInputBox
    # tbInput
    # btnOk


    # btnAssign

def main(app=None,context=None):
    global botStop
    global data
    try:
        CreateSubjectLineBuilder(app=app)
        if not botStop:
            CompleteTriageAndExportAttachment()
        if not botStop:
            OpenUnity()
            notify("Complete Assignment Stopper","Click OK Once Referral is Approved in ICasseManager to Proceed Complete Assisgnment")
            CompleteAssignment(data["claimNumber"])

        # notify("Complete Assignment Stopper","Click OK Once Referral is Approved in ICasseManager to Proceed Complete Assisgnment")
        # CompleteAssignment()    
        
    except Exception as ex:
        # print("error occurred: ", ex)
        exc_type, exc_obj, tb = sys.exc_info()
        line_number = tb.tb_lineno
        print(f"error occurred: {ex} (line {line_number})")
        
    finally:
        time.sleep(10)

if __name__ == "__main__":
    main()



